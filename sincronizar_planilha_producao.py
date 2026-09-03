"""Sincroniza a Gestão Produção (`Pedido`/`ItemPedido`) com a aba "GERAL TESTE"
de uma planilha de controle mais recente — pedido do Bruno em 25/08/2026,
revisado em 03/09/2026 pra uma nova versão da planilha ("03_09 CONTROLE
PRODUCAO_V1.xlsx") com o layout de colunas deslocado em relação à de 25/08
(ver `_validar_cabecalhos` — a sincronização recusa rodar, só loga e sai, se
os cabeçalhos não baterem com o esperado, em vez de importar dado no campo
errado).

Diferente de `importar_gestao_operacao.py` (que só roda sobre uma tabela
ainda vazia), esta sincronização roda POR CIMA de dados de produção que já
existem — por isso é protegida por um marcador em `ControleSistema` (ver
`_sincronizar_planilha_producao_25_08_2026`/`_sincronizar_planilha_producao_03_09_2026`
em app.py), não por uma checagem de "tabela vazia". Cada chave de
`ControleSistema` roda só UMA vez, mas o módulo é reaproveitado a cada nova
planilha que o Bruno manda — por isso o COL abaixo reflete sempre a versão
MAIS RECENTE da planilha (a sincronização antiga, de 25/08, já rodou e nunca
mais roda de novo, então mudar o mapeamento aqui não afeta ela).

Regras combinadas com o Bruno (25/08/2026, ainda valendo):
  1. Pedidos/itens que já existem no site são ATUALIZADOS com os valores
     atuais da planilha; pedidos que estão na planilha mas não existem no
     site são CRIADOS.
  2. Planejamento semanal também é trazido da planilha (antes era só
     manual, mas a planilha já vem preenchida — o Bruno confirmou trazer).
  3. Pedidos que já existem no site mas NÃO aparecem na planilha ficam
     intocados — a sincronização só toca no que está listado na planilha.

Casamento por `pedido_venda` (== "Pedido de Venda" da planilha), exato só —
depois da consolidação de duplicados (fase anterior) cada pedido_venda
corresponde a NO MÁXIMO UM `Pedido`. Dentro de um pedido, cada linha da
planilha (um produto) casa com um `ItemPedido` existente pela descrição do
produto normalizada (maiúsculas, sem espaço nas pontas); sem casar, vira um
item novo. Itens que já existem no site mas não aparecem na planilha (para
aquele pedido) ficam intocados, pelo mesmo motivo do item 3 acima.

Campos deliberadamente NÃO importados desta planilha:
  - LT COMERCIAL / TEMPO DE ESPERA / LT PRODUÇÃO: métricas calculadas
    (diferença de datas) só para uso interno da própria planilha — o site
    já calcula os equivalentes dele mesmo (lt_comercial_dias etc.).
  - VENDA TOTAL ITEM / TOTAL PEDIDO: não são campos guardados no banco,
    são propriedades calculadas (quantidade × custo_unitario, e a soma dos
    itens) — importar o valor unitário (VENDA UNIDADE ITEM) já basta.
  - DESVIO / QTDE PEÇAS / DESCRIÇÃO / RNC: a versão 25/08 não tinha essas
    colunas preenchidas; a versão 03/09 nem tem mais a coluna RNC.

Campo novo a partir de 03/09/2026:
  - LIBERAÇÃO REAL: a planilha de 25/08 não tinha equivalente (por isso
    ficava só manual no site); a de 03/09 passou a ter, então agora é
    importado pra `ItemPedido.liberacao_real`. Junto, `termino_inspecao`
    passa a ser preenchido com o mesmo valor de LIBERAÇÃO DE FATURAMENTO —
    mesmo par que o campo único "Conclusão produção" da tela de edição já
    grava (ver `editar_pedido` em app.py) — pra que lt_producao_dias e
    demais métricas calculadas fiquem corretas nos pedidos importados.
"""

import collections
import re
from datetime import date, datetime

import openpyxl

from extensions import db
from models import ESTACOES, FRETE_OPCOES, PRIORIDADE_OPCOES, STATUS_OPCOES, ItemPedido, Pedido, PedidoOperacao

SHEET_NAME = "GERAL TESTE"
HEADER_ROW = 2
FIRST_ROW = 3

# Mapeamento válido pra planilha "03_09 CONTROLE PRODUCAO_V1.xlsx" (pedido do
# Bruno, 03/09/2026) — layout deslocado em relação à versão de 25/08
# (ex.: "PRIORIDADE" era coluna 5, agora é 4; não existe mais coluna RNC).
COL = {
    "data_cliente": 1,
    "data_inclusao": 2,
    "prioridade": 4,
    "cliente": 5,
    "cnpj": 6,
    "cidade": 7,
    "estado_pais": 8,
    "frete": 9,
    "vendedor": 10,
    "pedido_venda": 11,
    "descricao_produto": 12,
    "quantidade": 13,
    "inicio_producao": 14,
    "liberacao_faturamento": 15,
    "status_producao": 17,
    "estacao": 18,
    "venda_unidade": 19,
    "obs": 22,
    "liberacao_prevista": 23,
    "liberacao_real": 24,
    "planejamento_semanal": 25,
}

# Texto (normalizado: maiúsculo, quebras de linha viram espaço, espaços
# duplicados colapsados) que cada coluna do dict acima DEVE conter na linha
# de cabeçalho — checado por `_validar_cabecalhos` antes de importar
# qualquer linha, pra nunca gravar um campo no lugar errado silenciosamente
# se o Bruno mandar uma planilha com colunas rearranjadas de novo.
_CABECALHOS_ESPERADOS = {
    "data_cliente": "DATA SOLICITADA",
    "data_inclusao": "INCLUS",
    "prioridade": "PRIORIDADE",
    "cliente": "CLIENTE",
    "cnpj": "CNPJ",
    "cidade": "CIDADE",
    "estado_pais": "ESTADO",
    "frete": "FRETE",
    "vendedor": "VENDEDOR",
    "pedido_venda": "PEDIDO DE VENDA",
    "descricao_produto": "DESCRI",
    "quantidade": "QUANTIDADE",
    "inicio_producao": "INICIO",
    "liberacao_faturamento": "FATURAMENTO",
    "status_producao": "STATUS",
    "estacao": "ESTA",
    "venda_unidade": "UNIDADE",
    "obs": "OBS",
    "liberacao_prevista": "PREVISTA",
    "liberacao_real": "REAL",
}

_MESES_ABREV_PCP = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

_UFS_VALIDAS = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG",
    "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
}

# Só mapeamos país quando a planilha diz uma coisa específica o suficiente
# pra não arriscar sobrescrever um país já certo no site com um chute. "EXP"
# sozinho não diz QUAL país — fica de fora de propósito (ver docstring).
_PAISES_RECONHECIDOS = {
    "CHILE": "Chile",
    "BOLIVIA": "Bolívia",
    "BOLÍVIA": "Bolívia",
    "COL": "Colômbia",
    "EQUADOR": "Equador",
}

_PRIORIDADE_VALIDA = set(PRIORIDADE_OPCOES)
_STATUS_VALIDO = set(STATUS_OPCOES)
_ESTACAO_VALIDA = set(ESTACOES)


def _cell(ws, row, campo):
    return ws.cell(row=row, column=COL[campo]).value


def _parse_texto(v, max_len=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if max_len:
        s = s[:max_len]
    return s


def _parse_data(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_numero(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _normaliza_pedido_venda(v):
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip() or None


def _normaliza_prioridade(v):
    s = _parse_texto(v)
    if not s:
        return None
    s = s.strip().upper()
    return s if s in _PRIORIDADE_VALIDA else None


def _normaliza_status(v):
    s = _parse_texto(v)
    if not s:
        return None
    s = s.strip().upper()
    return s if s in _STATUS_VALIDO else None


def _normaliza_estacao(v):
    s = _parse_texto(v)
    if not s:
        return None
    s = s.strip().upper()
    # "SOBRESSALENTE PU" (planilha) -> "PU" (site): decisão do Bruno
    # (03/09/2026, AskUserQuestion) — a estação "PU" já é a que
    # RDIM_ESTACOES_OPCOES (Qualidade, que não pode ser tocado) espera pra
    # liberar inspeção RDIM, então mantemos o nome "PU" em vez de renomear
    # a estação pra "SOBRESSALENTE PU" (ver ESTACOES em models.py).
    if s == "SOBRESSALENTE PU":
        return "PU"
    if s in _ESTACAO_VALIDA:
        return s
    if s == "OUTRO":
        return "OUTROS"
    return None


def _normaliza_frete(v):
    s = _parse_texto(v)
    if not s:
        return None
    s = s.strip().upper()
    if s in ("FOB",):
        return "FOB"
    if s in ("CIF",):
        return "CIF"
    if s in ("SEM", "SEM FRETE"):
        return "SEM FRETE"
    if s in ("EXP", "EXPORTAÇÃO", "EXPORTACAO"):
        return "EXPORTAÇÃO"
    return None if s not in FRETE_OPCOES else s


def _normaliza_estado_pais(v):
    """Devolve (estado, pais) — ou (None, None) quando o valor da planilha
    não é confiável o bastante pra mexer no que já está cadastrado (ver
    docstring do módulo)."""
    s = _parse_texto(v)
    if not s:
        return None, None
    s = s.strip().upper()
    if s in _UFS_VALIDAS:
        return s, "Brasil"
    if s in _PAISES_RECONHECIDOS:
        return None, _PAISES_RECONHECIDOS[s]
    return None, None


_RE_SEMANA_PLANILHA = re.compile(
    r"^SEM(?:ANA)?\s*0*(\d{1,2})\s*/\s*([A-ZÇÃÕ]{3})\s*/\s*(\d{4})$"
)


def _normaliza_semana(v):
    """Converte "SEM 04 / JUN / 2026" (planilha) -> "SEMANA 04 / JUN / 2026"
    (formato usado pelo site, gerar_semanas_pcp). Rejeita qualquer coisa que
    não bata no formato esperado (datas soltas, anos claramente errados por
    arrasto de célula etc.) em vez de tentar adivinhar."""
    s = _parse_texto(v)
    if not s:
        return None
    s = s.strip().upper()
    m = _RE_SEMANA_PLANILHA.match(s)
    if not m:
        return None
    semana, mes, ano = m.group(1), m.group(2), m.group(3)
    if mes not in _MESES_ABREV_PCP:
        return None
    try:
        semana_n = int(semana)
        ano_n = int(ano)
    except ValueError:
        return None
    # Faixa de ano deliberadamente estreita (não um range genérico) — a
    # planilha tem um punhado de linhas com "SEM 04 / MAR / 2027..2032"
    # (erro de arrastar célula no Excel, 1 linha cada) que passariam num
    # range mais largo; todo valor legítimo observado nesta planilha é 2026
    # (com folga pra 2025, por segurança).
    if not (1 <= semana_n <= 5) or ano_n not in (2025, 2026):
        return None
    return f"SEMANA {semana_n:02d} / {mes} / {ano_n}"


def _normaliza_descricao_chave(v):
    s = _parse_texto(v) or ""
    return re.sub(r"\s+", " ", s.strip().upper())


def _normaliza_cabecalho(v):
    s = _parse_texto(v) or ""
    return re.sub(r"\s+", " ", s.replace("\n", " ")).strip().upper()


def _validar_cabecalhos(ws):
    """Confere se a linha de cabeçalho (`HEADER_ROW`) da planilha bate com o
    que `COL` espera pra cada campo — comparação por "contém o texto chave"
    (ex.: campo "descricao_produto" só precisa achar "DESCRI" em algum lugar
    do cabeçalho, cobre tanto "DESCRIÇÃO DO PRODUTO" quanto variações de
    acentuação/quebra de linha). Devolve a lista de divergências (texto
    pronto pra log) — lista vazia significa "pode importar com segurança".

    Chamada ANTES de importar qualquer linha (ver `sincronizar_planilha_producao`)
    pra nunca gravar dado no campo errado silenciosamente se o Bruno mandar
    uma planilha com colunas rearranjadas de novo (foi exatamente o que
    aconteceu entre a versão de 25/08 e a de 03/09 — ver docstring do
    módulo). Nunca lança exceção: se `COL[campo]` apontar pra uma coluna
    fora do intervalo da planilha, isso também conta como divergência
    (planilha mais curta que o esperado), não como erro de programação —
    quem chama decide o que fazer (nunca derruba o boot da aplicação)."""
    divergencias = []
    for campo, esperado in _CABECALHOS_ESPERADOS.items():
        coluna = COL[campo]
        try:
            bruto = ws.cell(row=HEADER_ROW, column=coluna).value
        except Exception:
            bruto = None
        real = _normaliza_cabecalho(bruto)
        if esperado not in real:
            divergencias.append(
                f'{campo} (coluna {coluna}): esperado conter "{esperado}", '
                f'achou "{real or "(vazio)"}"'
            )
    return divergencias


def _ler_linhas(ws, diagnostico):
    """Devolve a lista de linhas (dict) já parseadas, na ordem da planilha,
    ignorando linhas totalmente em branco (sem pedido_venda nem cliente).

    `diagnostico` recebe uma linha por valor bruto que TINHA algo escrito na
    planilha mas não bateu com nenhuma opção válida do site (ex.: uma
    estação/frete/prioridade/semana escrita diferente do esperado) — usado
    só pra relatar ao Bruno o que não foi trazido automaticamente."""
    linhas = []
    for row in range(FIRST_ROW, ws.max_row + 1):
        pedido_venda = _normaliza_pedido_venda(_cell(ws, row, "pedido_venda"))
        cliente = _parse_texto(_cell(ws, row, "cliente"), 200)
        if not pedido_venda and not cliente:
            continue

        estado, pais = _normaliza_estado_pais(_cell(ws, row, "estado_pais"))
        prioridade = _normaliza_prioridade(_cell(ws, row, "prioridade"))
        frete = _normaliza_frete(_cell(ws, row, "frete"))
        status_producao = _normaliza_status(_cell(ws, row, "status_producao"))
        estacao = _normaliza_estacao(_cell(ws, row, "estacao"))
        semana = _normaliza_semana(_cell(ws, row, "planejamento_semanal"))

        for campo, normalizado, bruto in (
            ("prioridade", prioridade, _cell(ws, row, "prioridade")),
            ("frete", frete, _cell(ws, row, "frete")),
            ("status_producao", status_producao, _cell(ws, row, "status_producao")),
            ("estacao", estacao, _cell(ws, row, "estacao")),
            ("planejamento_semanal", semana, _cell(ws, row, "planejamento_semanal")),
        ):
            if normalizado is None and bruto not in (None, ""):
                diagnostico[campo][str(bruto).strip()].append(row)

        linhas.append(
            {
                "linha_planilha": row,
                "pedido_venda": pedido_venda,
                "cliente": cliente,
                "cnpj": _parse_texto(_cell(ws, row, "cnpj"), 30),
                "cidade": _parse_texto(_cell(ws, row, "cidade"), 120),
                "vendedor": _parse_texto(_cell(ws, row, "vendedor"), 120),
                "data_cliente": _parse_data(_cell(ws, row, "data_cliente")),
                "data_inclusao": _parse_data(_cell(ws, row, "data_inclusao")),
                "prioridade": prioridade,
                "frete": frete,
                "estado": estado,
                "pais": pais,
                "obs": _parse_texto(_cell(ws, row, "obs")),
                "descricao_produto": _parse_texto(_cell(ws, row, "descricao_produto"), 300),
                "quantidade": _parse_numero(_cell(ws, row, "quantidade")),
                "venda_unidade": _parse_numero(_cell(ws, row, "venda_unidade")),
                "status_producao": status_producao,
                "estacao": estacao,
                "liberacao_prevista": _parse_data(_cell(ws, row, "liberacao_prevista")),
                "liberacao_faturamento": _parse_data(_cell(ws, row, "liberacao_faturamento")),
                "inicio_producao": _parse_data(_cell(ws, row, "inicio_producao")),
                "planejamento_semanal": semana,
                "liberacao_real": _parse_data(_cell(ws, row, "liberacao_real")),
            }
        )
    return linhas


def _agrupar_por_pedido(linhas):
    grupos = {}
    ordem = []
    for linha in linhas:
        pv = linha["pedido_venda"]
        if not pv:
            continue  # sem número de pedido não dá pra casar nem criar com segurança
        if pv not in grupos:
            grupos[pv] = []
            ordem.append(pv)
        grupos[pv].append(linha)
    return ordem, grupos


def _aplicar_campos_pedido(pedido, rep, stats):
    """Atualiza os campos de IDENTIDADE do pedido a partir da linha
    representante (primeira linha do grupo) — só sobrescreve quando a
    planilha tem valor (célula em branco nunca apaga o que já está no
    site)."""
    campos = [
        ("data_cliente", rep["data_cliente"]),
        ("data_inclusao_pedido", rep["data_inclusao"]),
        ("prioridade", rep["prioridade"]),
        ("cliente", rep["cliente"]),
        ("cnpj", rep["cnpj"]),
        ("cidade", rep["cidade"]),
        ("vendedor", rep["vendedor"]),
        ("frete", rep["frete"]),
    ]
    if rep["estado"]:
        campos.append(("estado", rep["estado"]))
        campos.append(("pais", rep["pais"]))
    elif rep["pais"]:
        campos.append(("pais", rep["pais"]))

    for campo, valor in campos:
        if valor is None:
            continue
        if getattr(pedido, campo) != valor:
            setattr(pedido, campo, valor)
            stats["campos_pedido_atualizados"] += 1

    if rep["obs"] and not pedido.obs:
        pedido.obs = rep["obs"]


def _aplicar_campos_item(item, linha, stats):
    campos = [
        ("quantidade", linha["quantidade"]),
        ("custo_unitario", linha["venda_unidade"]),
        ("liberacao_prevista", linha["liberacao_prevista"]),
        ("liberacao_faturamento", linha["liberacao_faturamento"]),
        ("inicio_producao", linha["inicio_producao"]),
        ("planejamento_semanal", linha["planejamento_semanal"]),
        ("liberacao_real", linha["liberacao_real"]),
    ]
    for campo, valor in campos:
        if valor is None:
            continue
        if getattr(item, campo) != valor:
            setattr(item, campo, valor)
            stats["campos_item_atualizados"] += 1

    # termino_inspecao acompanha liberacao_faturamento — mesmo par que o
    # campo único "Conclusão produção" da tela de edição grava (editar_pedido
    # em app.py), pra lt_producao_dias e afins ficarem corretos nos pedidos
    # importados (ver docstring do módulo).
    if linha["liberacao_faturamento"] and item.termino_inspecao != linha["liberacao_faturamento"]:
        item.termino_inspecao = linha["liberacao_faturamento"]
        stats["campos_item_atualizados"] += 1

    if linha["status_producao"]:
        # status_manual precisa ficar sincronizado com "é EM TRATATIVA?" mesmo
        # quando status_producao em si já estava certo (senão um item que já
        # chegou EM TRATATIVA antes desta sincronização fica sem a trava que
        # impede o recálculo automático de sobrescrevê-lo depois).
        novo_manual = linha["status_producao"] == "EM TRATATIVA"
        if item.status_producao != linha["status_producao"]:
            item.status_producao = linha["status_producao"]
            stats["campos_item_atualizados"] += 1
        if item.status_manual != novo_manual:
            item.status_manual = novo_manual
            stats["campos_item_atualizados"] += 1

    if linha["estacao"] and item.estacao != linha["estacao"]:
        item.estacao = linha["estacao"]
        stats["campos_item_atualizados"] += 1


def _criar_item(pedido, linha, stats):
    item = ItemPedido(
        pedido=pedido,
        descricao_produto=linha["descricao_produto"] or "(sem descrição)",
        quantidade=linha["quantidade"] or 0,
        custo_unitario=linha["venda_unidade"] or 0,
        estacao=linha["estacao"],
        status_producao=linha["status_producao"] or "PENDENTE",
        status_manual=(linha["status_producao"] == "EM TRATATIVA"),
        liberacao_prevista=linha["liberacao_prevista"],
        liberacao_faturamento=linha["liberacao_faturamento"],
        inicio_producao=linha["inicio_producao"],
        planejamento_semanal=linha["planejamento_semanal"],
        liberacao_real=linha["liberacao_real"],
        # termino_inspecao acompanha liberacao_faturamento — ver
        # _aplicar_campos_item / docstring do módulo.
        termino_inspecao=linha["liberacao_faturamento"],
    )
    db.session.add(item)
    stats["itens_criados"] += 1
    return item


def sincronizar_planilha_producao(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    divergencias_cabecalho = _validar_cabecalhos(ws)
    if divergencias_cabecalho:
        # Nunca importa com o mapeamento de colunas incerto — devolve só o
        # diagnóstico. Quem chama (app.py) loga um erro e NÃO marca o
        # ControleSistema como concluído, pra poder tentar de novo depois
        # que o COL for corrigido (ver docstring do módulo / _validar_cabecalhos).
        return {
            "cabecalhos_invalidos": divergencias_cabecalho,
            "linhas_lidas": 0,
            "pedidos_na_planilha": 0,
            "pedidos_atualizados": 0,
            "pedidos_criados": 0,
            "itens_atualizados": 0,
            "itens_criados": 0,
            "campos_pedido_atualizados": 0,
            "campos_item_atualizados": 0,
            "pedidos_sem_cliente_ignorados": [],
            "valores_nao_reconhecidos": {},
        }

    diagnostico = {
        campo: collections.defaultdict(list)
        for campo in ("prioridade", "frete", "status_producao", "estacao", "planejamento_semanal")
    }

    linhas = _ler_linhas(ws, diagnostico)
    ordem, grupos = _agrupar_por_pedido(linhas)

    pedidos_existentes = {}
    for p in Pedido.query.filter(Pedido.pedido_venda.isnot(None)).all():
        pv = p.pedido_venda.strip()
        if pv and pv not in pedidos_existentes:  # já consolidado — 1 registro por pedido_venda
            pedidos_existentes[pv] = p

    stats = {
        "linhas_lidas": len(linhas),
        "pedidos_na_planilha": len(ordem),
        "pedidos_atualizados": 0,
        "pedidos_criados": 0,
        "itens_atualizados": 0,
        "itens_criados": 0,
        "campos_pedido_atualizados": 0,
        "campos_item_atualizados": 0,
        "pedidos_sem_cliente_ignorados": [],
    }

    for pv in ordem:
        grupo = grupos[pv]
        rep = grupo[0]

        pedido = pedidos_existentes.get(pv)
        if pedido is None:
            if not rep["cliente"]:
                stats["pedidos_sem_cliente_ignorados"].append(pv)
                continue
            pedido = Pedido(
                pedido_venda=pv,
                cliente=rep["cliente"],
                prioridade=rep["prioridade"] or "MÉDIA",
            )
            db.session.add(pedido)
            _aplicar_campos_pedido(pedido, rep, stats)
            stats["pedidos_criados"] += 1
            pedidos_existentes[pv] = pedido
        else:
            _aplicar_campos_pedido(pedido, rep, stats)
            stats["pedidos_atualizados"] += 1

        # -- casamento de itens dentro do pedido, por descrição normalizada --
        itens_disponiveis = list(pedido.itens)
        usados = set()
        for linha in grupo:
            chave = _normaliza_descricao_chave(linha["descricao_produto"])
            item_casado = None
            for item in itens_disponiveis:
                if item.id in usados:
                    continue
                if _normaliza_descricao_chave(item.descricao_produto) == chave:
                    item_casado = item
                    break
            if item_casado is not None:
                usados.add(item_casado.id)
                antes = stats["campos_item_atualizados"]
                _aplicar_campos_item(item_casado, linha, stats)
                if stats["campos_item_atualizados"] > antes:
                    stats["itens_atualizados"] += 1
            else:
                novo_item = _criar_item(pedido, linha, stats)
                db.session.flush()  # garante item.id pra não ser reusado no mesmo grupo
                itens_disponiveis.append(novo_item)
                usados.add(novo_item.id)

    db.session.commit()

    stats["valores_nao_reconhecidos"] = {
        campo: {valor: linhas_ref for valor, linhas_ref in valores.items()}
        for campo, valores in diagnostico.items()
        if valores
    }
    return stats


def seed_pedidos_operacao_basico(xlsx_path):
    """Cria um `PedidoOperacao` "básico" pra cada pedido desta mesma planilha
    de Produção que AINDA NÃO existe em Gestão Operação — pedido do Bruno
    (03/09/2026, `AskUserQuestion`, opção "Criar registros básicos em
    Operação também"): esta planilha ("CONTROLE PRODUCAO") não tem os dados
    de OTD/PCP/logística que normalmente vêm de uma planilha separada
    ("Gestão de Fluxo Produtivo", importada por `importar_gestao_operacao.py`)
    — então só preenche os campos comerciais que ELA tem
    (pedido_venda/cliente/vendedor/data_inclusao_pedido/prioridade/frete/
    pais/estado/cidade) e deixa todo o resto (go_*) em branco, pro Bruno
    completar manualmente depois.

    Casamento por `pedido_venda` (trim) contra os `PedidoOperacao` já
    existentes — pedido QUE JÁ TEM registro em Operação (vindo do backfill
    de Fase 13 ou do import da planilha "Gestão de Fluxo Produtivo") fica
    intocado, só os que faltam ganham um registro novo. Não precisa de
    `ControleSistema` própria: é chamada pela mesma função guardada de
    app.py que já protege `sincronizar_planilha_producao` (mesma chave,
    mesmo commit único)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    divergencias_cabecalho = _validar_cabecalhos(ws)
    if divergencias_cabecalho:
        return {
            "cabecalhos_invalidos": divergencias_cabecalho,
            "pedidos_operacao_criados": 0,
            "pedidos_operacao_ja_existentes": 0,
        }

    diagnostico = {
        campo: collections.defaultdict(list)
        for campo in ("prioridade", "frete", "status_producao", "estacao", "planejamento_semanal")
    }
    linhas = _ler_linhas(ws, diagnostico)
    ordem, grupos = _agrupar_por_pedido(linhas)

    pedidos_operacao_existentes = set()
    for pv, in PedidoOperacao.query.with_entities(PedidoOperacao.pedido_venda).filter(
        PedidoOperacao.pedido_venda.isnot(None)
    ).all():
        pv = (pv or "").strip()
        if pv:
            pedidos_operacao_existentes.add(pv)

    stats = {"pedidos_operacao_criados": 0, "pedidos_operacao_ja_existentes": 0}

    for pv in ordem:
        if pv in pedidos_operacao_existentes:
            stats["pedidos_operacao_ja_existentes"] += 1
            continue
        rep = grupos[pv][0]
        if not rep["cliente"]:
            continue  # mesma trava de _sincronizar_planilha_producao: sem cliente não cria

        db.session.add(
            PedidoOperacao(
                pedido_venda=pv,
                cliente=rep["cliente"],
                vendedor=rep["vendedor"],
                data_inclusao_pedido=rep["data_inclusao"],
                prioridade=rep["prioridade"] or "MÉDIA",
                frete=rep["frete"],
                pais=rep["pais"] or "Brasil",
                estado=rep["estado"],
                cidade=rep["cidade"],
            )
        )
        pedidos_operacao_existentes.add(pv)
        stats["pedidos_operacao_criados"] += 1

    db.session.commit()
    return stats
