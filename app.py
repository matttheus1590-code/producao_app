import csv
import io
import os
import re
from calendar import monthrange
from datetime import date, datetime, timedelta
from itertools import zip_longest

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, session, url_for
from flask_login import current_user, login_required, login_user, logout_user
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import and_, extract, false, func, inspect, or_, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from extensions import db, login_manager
from models import (
    ESTACOES,
    ESTACOES_GRUPOS_MONITORAMENTO,
    FRETE_OPCOES,
    GO_OTD_META_PERCENTUAL,
    GO_STATUS_PEDIDO_INFO_CORES,
    GO_STATUS_PEDIDO_INFO_OPCOES,
    GO_TIPO_PEDIDO_OPCOES,
    PD_CATEGORIA_OPCOES,
    PD_ETAPA_CORES,
    PD_ETAPA_OPCOES,
    PD_RESULTADO_ESPERADO_OPCOES,
    PD_TESTE_RESULTADO_INFO,
    PD_TESTE_RESULTADO_OPCOES,
    PD_TIPO_EVENTO_OPCOES,
    PRAZO_ALERTA_DIAS,
    PRIORIDADE_CORES,
    PRIORIDADE_OPCOES,
    RDIM_CATEGORIA_DESVIO_OPCOES,
    RDIM_ESTACOES_OPCOES,
    RDIM_GRANDEZAS_PADRAO,
    RDIM_INSPECAO_VISUAL_OPCOES,
    RDIM_RESULTADO_CORES,
    RDIM_RESULTADO_LABELS,
    RDIM_RESULTADO_OPCOES,
    RDIM_SUBCATEGORIA_DESVIO_OPCOES,
    REGIAO_POR_UF,
    REGIOES_OPCOES,
    RNC_DISPOSICAO_OPCOES,
    RNC_EFICACIA_CORES,
    RNC_EFICACIA_OPCOES,
    RNC_EMITENTE_OPCOES,
    RNC_FERRAMENTA_ANALISE_OPCOES,
    RNC_LOCAL_SETOR_OPCOES,
    RNC_ORIGEM_OPCOES,
    RNC_SETOR_OPCOES,
    RNC_SEVERIDADE_CORES,
    RNC_SEVERIDADE_OPCOES,
    RNC_SIM_NAO_OPCOES,
    RNC_STATUS_ACAO_OPCOES,
    RNC_STATUS_GERAL_ABERTOS,
    RNC_STATUS_GERAL_CORES,
    RNC_STATUS_GERAL_OPCOES,
    RNC_TIPO_NC_OPCOES,
    SEMAFORO_CORES,
    SEMAFORO_LABELS,
    STATUS_CHAO_CORES,
    STATUS_CHAO_LABELS,
    STATUS_CHAO_OPCOES,
    STATUS_CORES,
    STATUS_OPCOES,
    ControleSistema,
    Estacao,
    HistoricoAlteracao,
    InspecaoFinal,
    ItemPedido,
    Pedido,
    PedidoOperacao,
    Programacao,
    ProjetoPD,
    RdimMedicao,
    RdimPecaDesvio,
    RncQualidade,
    TesteProjetoPD,
    Transportadora,
    Usuario,
    VisitaReuniaoPD,
    gerar_semanas_pcp,
    rotulo_estacao,
)
from permissoes import ROLES, ROLES_LABELS, pode_editar_estacao, requer_role

MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PAGE_SIZE = 25

# Campos "importantes" que geram uma linha no histórico de alterações quando
# mudam de valor (não registramos tudo, só o que realmente importa acompanhar).
CAMPOS_HISTORICO_PEDIDO = ["cliente", "prioridade", "vendedor", "frete"]
CAMPOS_HISTORICO_ITEM = [
    "estacao",
    "status_producao",
    "inicio_producao",
    "inicio_inspecao",
    "termino_inspecao",
    "liberacao_faturamento",
    "liberacao_prevista",
    "liberacao_real",
    "planejamento_semanal",
    "rnc",
    "numero_nota_fiscal",
    "valor_faturado",
    "transportadora_id",
    "data_envio",
]

# Campos de Gestão Operação (Fase 13) que geram histórico de alteração —
# igual em espírito a CAMPOS_HISTORICO_PEDIDO/ITEM, só que namespaced com
# "go_" (não precisa de tabela nova: usa o mesmo HistoricoAlteracao de sempre).
CAMPOS_HISTORICO_GESTAO_OPERACAO = [
    "go_tipo_pedido",
    "go_status_pedido_info",
    "go_previsao_liberacao_pcp",
    "go_data_efetiva_liberacao_pcp",
    "go_status_logistica",
    "go_data_pedido_expedido",
    "go_transportadora_id",
    "go_data_real_entrega",
    "go_otd_realizado",
    "go_data_entregue_cliente",
    "go_status_final_alinhamento",
]

# Campos de P&D (Fase 14) que geram histórico de alteração — em especial
# etapa_atual, pra que toda troca de etapa (inclusive "andar pra trás", ex.:
# Teste -> Reprovado -> Desenvolvimento) fique registrada, como o Bruno pediu.
CAMPOS_HISTORICO_PD = [
    "etapa_atual",
    "percentual_conclusao",
    "prioridade",
    "responsavel",
    "data_prevista_conclusao",
    "data_real_conclusao",
    "custo_realizado",
    "investimento_realizado",
    "economia_realizada",
]

# Pedido do Bruno (31/08/2026): a tela de edição de Gestão Operação deixa de
# mostrar TODOS os blocos (Comercial/PCP/Logística/Resultados) de uma vez —
# cada aba só edita os campos da própria área. GO_CAMPOS_POR_SECAO é a única
# fonte de verdade de "quais campos pertencem a cada aba", usada tanto pra
# decidir o que o template desenha quanto pra decidir o que a rota lê do
# formulário — assim um campo fora da seção atual nunca é tocado (nem lido,
# nem zerado) ao salvar.
GO_SECOES = ("comercial", "pcp", "logistica", "resultados")
GO_CAMPOS_POR_SECAO = {
    "comercial": [
        "go_tipo_pedido", "go_contrato", "go_pedido_compra_cliente", "go_proposta",
        "go_data_solicitada_entrega", "go_status_pedido_info", "go_valor_pedido_operacao",
    ],
    "pcp": [
        "go_previsao_liberacao_pcp", "go_data_efetiva_liberacao_pcp",
        "go_data_solicitada_cliente_retira", "go_custo_producao_real", "go_termino_semanal_pcp",
    ],
    "logistica": [
        "go_data_emissao_nf", "go_valor_nf_emitida", "go_numero_nf", "go_status_logistica",
        "go_data_pedido_expedido", "go_transportadora_id", "go_custo_frete_previsto",
        "go_custo_frete_final", "go_custo_frete_sobre_nota", "go_data_prevista_entrega",
        "go_data_real_entrega",
    ],
    "resultados": [
        "go_otd_realizado", "go_data_solicitada_cliente_final", "go_data_entregue_cliente",
        "go_obs_operacao", "go_status_final_alinhamento",
    ],
}
GO_SECAO_ENDPOINT = {
    # "comercial" aponta pra Listagem Geral (pedido do Bruno, 01/09/2026: a
    # aba/lista "Comercial" separada foi apagada porque a Listagem Geral já
    # mostra as mesmas informações — só o FORMULÁRIO de edição da seção
    # Comercial continua existindo, dentro de gestao_operacao_editar).
    "comercial": "gestao_operacao_listagem_geral",
    "pcp": "gestao_operacao_pcp",
    "logistica": "gestao_operacao_logistica",
    "resultados": "gestao_operacao_resultados",
}
GO_SECAO_LABEL = {
    "comercial": "Comercial", "pcp": "PCP", "logistica": "Logística / NF", "resultados": "Resultados / OTD",
}
_GO_CAMPOS_DATA = {
    "go_data_solicitada_entrega", "go_previsao_liberacao_pcp", "go_data_efetiva_liberacao_pcp",
    "go_data_solicitada_cliente_retira", "go_data_emissao_nf", "go_data_pedido_expedido",
    "go_data_prevista_entrega", "go_data_real_entrega", "go_data_solicitada_cliente_final",
    "go_data_entregue_cliente",
}
_GO_CAMPOS_FLOAT = {
    "go_valor_pedido_operacao", "go_custo_producao_real", "go_valor_nf_emitida",
    "go_custo_frete_previsto", "go_custo_frete_final", "go_custo_frete_sobre_nota",
}


def _parse_campo_go(campo, f):
    """Lê e converte UM campo go_* do formulário — usado pela edição
    seccionada de Gestão Operação, campo por campo, só para os campos da
    seção que está sendo salva."""
    if campo == "go_transportadora_id":
        valor = f.get("go_transportadora_id", "")
        return int(valor) if valor.strip().isdigit() else None
    if campo in _GO_CAMPOS_DATA:
        return _parse_data_form(f.get(campo))
    if campo in _GO_CAMPOS_FLOAT:
        valor = f.get(campo, "")
        return _parse_float_form(valor, default=None) if valor.strip() else None
    return f.get(campo, "").strip() or None


def _resolve_database_uri():
    """Usa DATABASE_URL (Postgres do Render) quando existir; senão, SQLite local."""
    url = os.environ.get("DATABASE_URL")
    if url:
        # Render/Heroku às vezes fornecem "postgres://", mas o SQLAlchemy 2.x exige "postgresql://"
        if url.startswith("postgres://"):
            url = url.replace("postgres://", "postgresql://", 1)
        return url
    return "sqlite:///" + os.path.join(BASE_DIR, "instance", "pedidos.db")


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")
    app.config["SQLALCHEMY_DATABASE_URI"] = _resolve_database_uri()
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    db.init_app(app)
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(user_id):
        return db.session.get(Usuario, int(user_id))

    with app.app_context():
        os.makedirs(os.path.join(BASE_DIR, "instance"), exist_ok=True)
        db.create_all()
        _migrar_itens_pedido(app)
        _migrar_producao_para_itens(app)
        _migrar_usuarios_role(app)
        _migrar_estacoes_tabela(app)
        _atualizar_estacoes_planilha_03_09_2026(app)
        # Depende da reconciliação acima já ter rodado (garante que Reforma/
        # Revenda existem no cadastro pra serem excluídas, e que Manutenção /
        # Devolução já existe pra ser renomeada).
        _organizar_estacoes_03_09_2026(app)
        _migrar_faturamento_itens(app)
        _migrar_logistica_itens(app)
        _migrar_planejamento_semanal_itens(app)
        _migrar_liberacao_real_itens(app)
        _migrar_atualizado_em_itens(app)
        _migrar_gestao_operacao_pedidos(app)
        _migrar_dados_go_para_pedidos_operacao(app)
        # Só depois de TODAS as colunas de pedidos/itens existirem de verdade
        # (senão o ORM tenta selecionar coluna que ainda não foi criada nesta
        # execução, em bancos antigos que ainda não passaram pelas migrações
        # acima) — consolida os pedidos fragmentados em vários registros.
        _consolidar_pedidos_duplicados(app)
        _seed_inicial(app)
        _importar_gestao_operacao(app)
        # Roda por último: depende de Pedido/ItemPedido já existirem com todas
        # as colunas migradas, e de _consolidar_pedidos_duplicados já ter
        # deixado (no máximo) 1 Pedido por pedido_venda.
        _sincronizar_planilha_producao_25_08_2026(app)
        # Idem para Gestão Operação: depende de _importar_gestao_operacao já
        # ter rodado (banco com PedidoOperacao populado) antes de sincronizar
        # por cima com a planilha mais nova.
        _sincronizar_gestao_operacao_28_08_2026(app)
        # Correção pontual pedida pelo Bruno depois de conferir os números da
        # semana 04/Ago manualmente — precisa rodar depois da sincronização
        # acima (que é quem trouxe o pedido 835 com a semana errada).
        _corrigir_semana_pcp_morken_835(app)
        _seed_rnc_qualidade(app)
        _backfill_go_data_solicitada_cliente_retira(app)
        _migrar_rdim_inspecao_final(app)
        _migrar_rdim_pecas_desvio(app)
        # Roda por último de todos: depende de tudo acima (Pedido/ItemPedido
        # com todas as colunas migradas, PedidoOperacao já existindo).
        _sincronizar_planilha_producao_03_09_2026(app)
        # Mesmo motivo do sync 28/08 acima, só que com a planilha mais nova
        # que o Bruno mandou em 03/09 (mesmo layout de coluna, bloco
        # "Resultados" novo) — depende só de PedidoOperacao já existir.
        _sincronizar_gestao_operacao_03_09_2026(app)

    @app.context_processor
    def inject_globals():
        estacoes_ativas = [e.nome for e in Estacao.query.filter_by(ativo=True).order_by(Estacao.ordem_exibicao).all()]
        return dict(
            ESTACOES=estacoes_ativas or ESTACOES,
            STATUS_OPCOES=STATUS_OPCOES,
            PRIORIDADE_OPCOES=PRIORIDADE_OPCOES,
            FRETE_OPCOES=FRETE_OPCOES,
            STATUS_CORES=STATUS_CORES,
            PRIORIDADE_CORES=PRIORIDADE_CORES,
            ROLES=ROLES,
            ROLES_LABELS=ROLES_LABELS,
            REGIOES_OPCOES=REGIOES_OPCOES,
            SEMAFORO_CORES=SEMAFORO_CORES,
            SEMAFORO_LABELS=SEMAFORO_LABELS,
            PRAZO_ALERTA_DIAS=PRAZO_ALERTA_DIAS,
            STATUS_CHAO_OPCOES=STATUS_CHAO_OPCOES,
            STATUS_CHAO_LABELS=STATUS_CHAO_LABELS,
            STATUS_CHAO_CORES=STATUS_CHAO_CORES,
            GO_OTD_META_PERCENTUAL=GO_OTD_META_PERCENTUAL,
            GO_SEMANAS_PCP=gerar_semanas_pcp(),
            GO_TIPO_PEDIDO_OPCOES=GO_TIPO_PEDIDO_OPCOES,
            GO_STATUS_PEDIDO_INFO_OPCOES=GO_STATUS_PEDIDO_INFO_OPCOES,
            GO_STATUS_PEDIDO_INFO_CORES=GO_STATUS_PEDIDO_INFO_CORES,
            RNC_EMITENTE_OPCOES=RNC_EMITENTE_OPCOES,
            RNC_SETOR_OPCOES=RNC_SETOR_OPCOES,
            RNC_ORIGEM_OPCOES=RNC_ORIGEM_OPCOES,
            RNC_LOCAL_SETOR_OPCOES=RNC_LOCAL_SETOR_OPCOES,
            RNC_TIPO_NC_OPCOES=RNC_TIPO_NC_OPCOES,
            RNC_SEVERIDADE_OPCOES=RNC_SEVERIDADE_OPCOES,
            RNC_FERRAMENTA_ANALISE_OPCOES=RNC_FERRAMENTA_ANALISE_OPCOES,
            RNC_DISPOSICAO_OPCOES=RNC_DISPOSICAO_OPCOES,
            RNC_STATUS_ACAO_OPCOES=RNC_STATUS_ACAO_OPCOES,
            RNC_EFICACIA_OPCOES=RNC_EFICACIA_OPCOES,
            RNC_SIM_NAO_OPCOES=RNC_SIM_NAO_OPCOES,
            RNC_STATUS_GERAL_OPCOES=RNC_STATUS_GERAL_OPCOES,
            RNC_SEVERIDADE_CORES=RNC_SEVERIDADE_CORES,
            RNC_STATUS_GERAL_CORES=RNC_STATUS_GERAL_CORES,
            RNC_EFICACIA_CORES=RNC_EFICACIA_CORES,
            PD_CATEGORIA_OPCOES=PD_CATEGORIA_OPCOES,
            PD_ETAPA_OPCOES=PD_ETAPA_OPCOES,
            PD_ETAPA_CORES=PD_ETAPA_CORES,
            PD_RESULTADO_ESPERADO_OPCOES=PD_RESULTADO_ESPERADO_OPCOES,
            PD_TIPO_EVENTO_OPCOES=PD_TIPO_EVENTO_OPCOES,
            PD_TESTE_RESULTADO_OPCOES=PD_TESTE_RESULTADO_OPCOES,
            PD_TESTE_RESULTADO_INFO=PD_TESTE_RESULTADO_INFO,
            RDIM_ESTACOES_OPCOES=RDIM_ESTACOES_OPCOES,
            RDIM_RESULTADO_OPCOES=RDIM_RESULTADO_OPCOES,
            RDIM_RESULTADO_LABELS=RDIM_RESULTADO_LABELS,
            RDIM_RESULTADO_CORES=RDIM_RESULTADO_CORES,
            RDIM_INSPECAO_VISUAL_OPCOES=RDIM_INSPECAO_VISUAL_OPCOES,
            RDIM_CATEGORIA_DESVIO_OPCOES=RDIM_CATEGORIA_DESVIO_OPCOES,
            RDIM_SUBCATEGORIA_DESVIO_OPCOES=RDIM_SUBCATEGORIA_DESVIO_OPCOES,
            RDIM_GRANDEZAS_PADRAO=RDIM_GRANDEZAS_PADRAO,
            hoje_iso=date.today().isoformat(),
        )

    register_routes(app)
    return app


def _migrar_itens_pedido(app):
    """Migra bancos criados antes de pedidos aceitarem vários produtos.

    Antes, cada Pedido tinha um único produto (colunas descricao_produto,
    quantidade e custo_unitario direto na tabela "pedidos"). Agora esses
    dados moram na tabela "itens_pedido" (um pedido -> vários itens). Esta
    função roda sozinha a cada start do site e só faz alguma coisa se
    detectar o formato antigo — não apaga nenhum pedido, só reorganiza os
    dados de produto para o novo formato.
    """
    inspector = inspect(db.engine)
    if "pedidos" not in inspector.get_table_names():
        return  # banco novo — db.create_all() já cuidou de tudo

    colunas_pedidos = {c["name"] for c in inspector.get_columns("pedidos")}
    formato_antigo = {"descricao_produto", "quantidade", "custo_unitario"}.issubset(colunas_pedidos)
    if not formato_antigo:
        return  # já está no formato novo

    with db.engine.begin() as conn:
        conn.execute(
            text(
                "INSERT INTO itens_pedido (pedido_id, descricao_produto, quantidade, custo_unitario) "
                "SELECT id, descricao_produto, quantidade, custo_unitario FROM pedidos"
            )
        )
        conn.execute(text("ALTER TABLE pedidos DROP COLUMN descricao_produto"))
        conn.execute(text("ALTER TABLE pedidos DROP COLUMN quantidade"))
        conn.execute(text("ALTER TABLE pedidos DROP COLUMN custo_unitario"))
    app.logger.info("Migração automática: produtos dos pedidos movidos para itens_pedido com sucesso.")


_COLUNAS_PRODUCAO = [
    "estacao",
    "status_producao",
    "status_manual",
    "inicio_producao",
    "inicio_inspecao",
    "termino_inspecao",
    "liberacao_faturamento",
    "liberacao_prevista",
    "rnc",
]

_TIPOS_COLUNAS_PRODUCAO = {
    "estacao": "VARCHAR(40)",
    "status_producao": "VARCHAR(20)",
    "status_manual": "BOOLEAN",
    "inicio_producao": "DATE",
    "inicio_inspecao": "DATE",
    "termino_inspecao": "DATE",
    "liberacao_faturamento": "DATE",
    "liberacao_prevista": "DATE",
    "rnc": "VARCHAR(120)",
}


def _migrar_producao_para_itens(app):
    """Migra bancos criados antes da produção (estação, status, datas, RNC) virar
    algo por ITEM em vez de por pedido inteiro.

    Como cada pedido tinha exatamente 1 item na época em que essas colunas ainda
    ficavam em "pedidos", a migração é uma cópia direta: cada item recebe os
    dados de produção do pedido que o originou. Não apaga nenhum dado.
    """
    inspector = inspect(db.engine)
    if "pedidos" not in inspector.get_table_names():
        return

    colunas_pedidos = {c["name"] for c in inspector.get_columns("pedidos")}
    formato_antigo = set(_COLUNAS_PRODUCAO).issubset(colunas_pedidos)
    if not formato_antigo:
        return  # já está no formato novo

    # db.create_all() não altera tabelas já existentes — "itens_pedido" pode já
    # existir (criada pela migração anterior) sem essas colunas novas ainda.
    colunas_itens = {c["name"] for c in inspector.get_columns("itens_pedido")}

    with db.engine.begin() as conn:
        for coluna, tipo in _TIPOS_COLUNAS_PRODUCAO.items():
            if coluna not in colunas_itens:
                conn.execute(text(f"ALTER TABLE itens_pedido ADD COLUMN {coluna} {tipo}"))
        for coluna in _COLUNAS_PRODUCAO:
            conn.execute(
                text(
                    f"UPDATE itens_pedido SET {coluna} = ("
                    f"SELECT p.{coluna} FROM pedidos p WHERE p.id = itens_pedido.pedido_id"
                    f")"
                )
            )
        for coluna in _COLUNAS_PRODUCAO:
            conn.execute(text(f"ALTER TABLE pedidos DROP COLUMN {coluna}"))
    app.logger.info("Migração automática: dados de produção movidos dos pedidos para os itens com sucesso.")


def _consolidar_pedidos_duplicados(app):
    """Corrige um problema histórico do import original: pedidos com mais de
    um produto foram importados como VÁRIOS registros `Pedido` separados (um
    por linha da planilha, cada um com 1 item só) em vez de 1 `Pedido` com
    vários `ItemPedido` dentro. Isso fazia "Venda total pedido" somar só 1
    produto por vez, e a tela de editar não mostrar todos os produtos de um
    pedido juntos (ex.: CATTALINI 728 = 4 registros `Pedido` separados, cada
    um com 1 item, em vez de 1 registro com 4 itens).

    Roda uma vez: agrupa os `Pedido` que compartilham o mesmo `pedido_venda`,
    escolhe o de menor id como "principal", move todos os itens — e o
    histórico de alterações — dos outros pra ele, e apaga os registros que
    sobraram vazios. Não perde nenhum item nem histórico, só reorganiza quem
    é o dono (pedido_id). Idempotente: numa segunda execução não encontra
    mais grupos com mais de 1 registro por pedido_venda, então não faz nada."""
    grupos = (
        db.session.query(Pedido.pedido_venda)
        .filter(Pedido.pedido_venda.isnot(None), Pedido.pedido_venda != "")
        .group_by(Pedido.pedido_venda)
        .having(func.count(Pedido.id) > 1)
        .all()
    )
    if not grupos:
        return

    total_pedidos_removidos = 0
    total_itens_movidos = 0

    for (pedido_venda,) in grupos:
        registros = Pedido.query.filter(Pedido.pedido_venda == pedido_venda).order_by(Pedido.id).all()
        if len(registros) < 2:
            continue  # já foi consolidado nesta mesma rodada (não deveria acontecer, mas por segurança)

        primario, duplicados = registros[0], registros[1:]

        datas_inclusao = [r.data_inclusao_pedido for r in registros if r.data_inclusao_pedido]
        if datas_inclusao:
            primario.data_inclusao_pedido = min(datas_inclusao)

        # Todos os outros campos escalares do pedido (comerciais + os go_*
        # legados de Gestão Operação, que ainda vivem fisicamente na tabela
        # mesmo sem serem mais lidos pelo app) — mantém o valor do principal
        # se já tiver algo, senão pega o primeiro valor não vazio encontrado
        # entre os duplicados. Genérico de propósito, pra não perder nenhum
        # dado esquecido numa lista manual de campos.
        campos_genericos = [
            c.name for c in Pedido.__table__.columns if c.name not in ("id", "pedido_venda", "data_inclusao_pedido")
        ]
        for campo in campos_genericos:
            if not getattr(primario, campo):
                for dup in duplicados:
                    valor = getattr(dup, campo)
                    if valor:
                        setattr(primario, campo, valor)
                        break

        for dup in duplicados:
            for item in list(dup.itens):
                item.pedido = primario  # via relationship, não só pedido_id — mantém o cascade consistente
                total_itens_movidos += 1
            HistoricoAlteracao.query.filter_by(pedido_id=dup.id).update(
                {"pedido_id": primario.id}, synchronize_session=False
            )
            db.session.flush()
            db.session.delete(dup)
            total_pedidos_removidos += 1

    db.session.commit()
    app.logger.info(
        "Migração automática: %d pedidos duplicados consolidados (%d registros removidos, %d itens reagrupados).",
        len(grupos),
        total_pedidos_removidos,
        total_itens_movidos,
    )


def _migrar_usuarios_role(app):
    """Adiciona os campos de papel de acesso (role/setor/ativo) em usuários
    criados antes do Dashboard Gerencial de PCP existir.

    Roda sozinha a cada início do site e só faz alguma coisa se detectar que
    essas colunas ainda não existem — não apaga nem altera nenhum usuário além
    de garantir que o "admin" original continue com acesso total (ADMIN) e que
    todo o resto continue podendo fazer o que já fazia hoje (PCP).
    """
    inspector = inspect(db.engine)
    if "usuarios" not in inspector.get_table_names():
        return  # banco novo — db.create_all() já cuidou de tudo

    colunas = {c["name"] for c in inspector.get_columns("usuarios")}
    faltando = [c for c in ("role", "setor", "ativo") if c not in colunas]
    if not faltando:
        return  # já está no formato novo

    with db.engine.begin() as conn:
        if "role" not in colunas:
            conn.execute(text("ALTER TABLE usuarios ADD COLUMN role VARCHAR(20)"))
        if "setor" not in colunas:
            conn.execute(text("ALTER TABLE usuarios ADD COLUMN setor VARCHAR(40)"))
        if "ativo" not in colunas:
            conn.execute(text("ALTER TABLE usuarios ADD COLUMN ativo BOOLEAN"))

        conn.execute(text("UPDATE usuarios SET role = 'PCP' WHERE role IS NULL"))
        conn.execute(text("UPDATE usuarios SET role = 'ADMIN' WHERE username = 'admin'"))
        conn.execute(text("UPDATE usuarios SET ativo = TRUE WHERE ativo IS NULL"))
    app.logger.info("Migração automática: usuários existentes receberam papel de acesso (role).")


def _migrar_estacoes_tabela(app):
    """Cadastra as estações como registros de verdade na tabela "estacoes", com
    os mesmos nomes e a mesma ordem da lista fixa ESTACOES (models.py).

    Como itens continuam guardando a estação pelo nome (texto), nenhum pedido
    existente precisa mudar — só passa a existir um registro correspondente
    pra cada nome, que os cadastros/tela de Estações usam."""
    if Estacao.query.count() > 0:
        return  # já foi semeado (ou o usuário já está gerenciando pelo cadastro)
    for i, nome in enumerate(ESTACOES):
        db.session.add(Estacao(nome=nome, ordem_exibicao=i, ativo=True))
    db.session.commit()
    app.logger.info("Migração automática: %d estações cadastradas como tabela.", len(ESTACOES))


_CHAVE_ESTACOES_PLANILHA_03_09_2026 = "atualizacao_estacoes_planilha_03_09_2026"
_ESTACOES_NOVAS_03_09_2026 = ["SOBRESSALENTE METAL MECANICA", "SOBRESSALENTE BORRACHA", "MANUTENÇÃO / DEVOLUÇÃO"]
_ESTACOES_DESATIVADAS_03_09_2026 = ["REFORMA", "REVENDA"]


def _atualizar_estacoes_planilha_03_09_2026(app):
    """Reconcilia a tabela `estacoes` (cadastro) com a nova lista ESTACOES —
    pedido do Bruno (03/09/2026): "considere exatamente as mesmas estações da
    planilha... revenda não vai existir mais, sobressalentes de borracha e
    metal mecânica passarão a existir". `_migrar_estacoes_tabela` só semeia
    a tabela quando ela está VAZIA, então em produção (já populada desde o
    primeiro boot) mudar a constante ESTACOES sozinha não reflete nos
    registros existentes — precisa desta migração à parte, guardada por
    ControleSistema pra rodar exatamente uma vez.

    Desativa (ativo=False) em vez de apagar — mesmo critério já usado no
    cadastro manual de estações (não existe rota de exclusão física, só
    ativar/desativar), então nenhum pedido antigo que porventura já tenha
    usado "Reforma"/"Revenda" fica com referência quebrada."""
    if ControleSistema.query.filter_by(chave=_CHAVE_ESTACOES_PLANILHA_03_09_2026).first() is not None:
        return

    for nome in _ESTACOES_DESATIVADAS_03_09_2026:
        estacao = Estacao.query.filter_by(nome=nome).first()
        if estacao is not None and estacao.ativo:
            estacao.ativo = False

    maior_ordem = db.session.query(func.max(Estacao.ordem_exibicao)).scalar() or 0
    for i, nome in enumerate(_ESTACOES_NOVAS_03_09_2026):
        if Estacao.query.filter_by(nome=nome).first() is None:
            maior_ordem += 1
            db.session.add(Estacao(nome=nome, ordem_exibicao=maior_ordem, ativo=True))

    db.session.add(ControleSistema(chave=_CHAVE_ESTACOES_PLANILHA_03_09_2026))
    db.session.commit()
    app.logger.info(
        "Migração automática: estações atualizadas conforme planilha 03/09/2026 — "
        "%s desativadas, %s criadas.",
        ", ".join(_ESTACOES_DESATIVADAS_03_09_2026), ", ".join(_ESTACOES_NOVAS_03_09_2026),
    )


_CHAVE_ORGANIZAR_ESTACOES_03_09_2026 = "organizar_estacoes_03_09_2026"
_ESTACOES_EXCLUIDAS_03_09_2026 = ["REFORMA", "REVENDA"]
_ESTACAO_RENOMEADA_03_09_2026 = ("MANUTENÇÃO / DEVOLUÇÃO", "MANUTENÇÃO")


def _organizar_estacoes_03_09_2026(app):
    """Segunda rodada de ajuste no cadastro de Estações, mesmo dia da
    reconciliação com a planilha (`_atualizar_estacoes_planilha_03_09_2026`),
    agora a pedido direto do Bruno na tela de Estações:
      1. Exclui de vez "Reforma" e "Revenda" — antes só tinham sido
         desativadas (não existia pedido pra excluir ainda). Como
         `ItemPedido.estacao` é texto solto, sem chave estrangeira pra
         `Estacao` (ver docstring do model), apagar a linha do cadastro é
         seguro — não quebra nenhum pedido antigo que porventura tenha usado
         esses nomes, só deixa de aparecer no cadastro/dropdowns.
      2. Renomeia "Manutenção / Devolução" pra só "Manutenção", tanto no
         cadastro quanto em qualquer ItemPedido já gravado com o nome antigo
         (pra não sobrar pedido usando um nome de estação que não existe
         mais em lugar nenhum do site).

    Guardado por ControleSistema pra rodar exatamente uma vez, mesmo padrão
    de todas as outras migrações de estações."""
    if ControleSistema.query.filter_by(chave=_CHAVE_ORGANIZAR_ESTACOES_03_09_2026).first() is not None:
        return

    excluidas = []
    for nome in _ESTACOES_EXCLUIDAS_03_09_2026:
        estacao = Estacao.query.filter_by(nome=nome).first()
        if estacao is not None:
            db.session.delete(estacao)
            excluidas.append(nome)

    nome_antigo, nome_novo = _ESTACAO_RENOMEADA_03_09_2026
    estacao_manutencao = Estacao.query.filter_by(nome=nome_antigo).first()
    if estacao_manutencao is not None:
        estacao_manutencao.nome = nome_novo

    itens_renomeados = ItemPedido.query.filter_by(estacao=nome_antigo).update(
        {"estacao": nome_novo}, synchronize_session=False
    )

    db.session.add(ControleSistema(chave=_CHAVE_ORGANIZAR_ESTACOES_03_09_2026))
    db.session.commit()
    app.logger.info(
        "Migração automática: estações organizadas (pedido Bruno 03/09/2026) — "
        "excluídas: %s | renomeada: '%s' -> '%s' (%d item(ns) atualizado(s)).",
        ", ".join(excluidas) or "nenhuma", nome_antigo, nome_novo, itens_renomeados,
    )


def _migrar_faturamento_itens(app):
    """Adiciona os campos de faturamento (número da nota fiscal e valor
    faturado) nos itens criados antes dessa fase existir.

    São campos 100% novos — a planilha original não tinha essa informação —
    então não há nada pra migrar/copiar: só ficam em branco nos itens antigos
    e passam a ser preenchidos dali pra frente."""
    inspector = inspect(db.engine)
    if "itens_pedido" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("itens_pedido")}
    faltando = [c for c in ("numero_nota_fiscal", "valor_faturado") if c not in colunas]
    if not faltando:
        return

    with db.engine.begin() as conn:
        if "numero_nota_fiscal" not in colunas:
            conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN numero_nota_fiscal VARCHAR(30)"))
        if "valor_faturado" not in colunas:
            conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN valor_faturado FLOAT"))
    app.logger.info("Migração automática: campos de faturamento (NF e valor faturado) adicionados aos itens.")


def _migrar_logistica_itens(app):
    """Adiciona os campos de logística (transportadora e data de envio) nos
    itens criados antes dessa fase existir.

    Assim como faturamento, são campos 100% novos — a planilha original não
    tinha essa informação — então ficam em branco nos itens antigos e passam
    a ser preenchidos dali pra frente. "transportadora_id" não usa FK de
    verdade (ver comentário no models.py), então a coluna é só um INTEGER."""
    inspector = inspect(db.engine)
    if "itens_pedido" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("itens_pedido")}
    faltando = [c for c in ("transportadora_id", "data_envio") if c not in colunas]
    if not faltando:
        return

    with db.engine.begin() as conn:
        if "transportadora_id" not in colunas:
            conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN transportadora_id INTEGER"))
        if "data_envio" not in colunas:
            conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN data_envio DATE"))
    app.logger.info("Migração automática: campos de logística (transportadora e data de envio) adicionados aos itens.")


def _migrar_planejamento_semanal_itens(app):
    """Adiciona o campo de planejamento semanal (preenchido manualmente pelo
    PCP, junto com a liberação prevista) nos itens criados antes dessa fase
    existir.

    É um campo 100% novo — não deriva de nenhum dado existente — então fica
    em branco nos itens antigos e passa a ser preenchido dali pra frente."""
    inspector = inspect(db.engine)
    if "itens_pedido" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("itens_pedido")}
    if "planejamento_semanal" in colunas:
        return

    with db.engine.begin() as conn:
        conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN planejamento_semanal VARCHAR(40)"))
    app.logger.info("Migração automática: campo de planejamento semanal adicionado aos itens.")


def _migrar_liberacao_real_itens(app):
    """Adiciona o campo de liberação real (preenchido manualmente, pra
    comparar com a liberação prevista) nos itens criados antes dessa fase
    existir.

    Campo 100% novo — não deriva de nenhum dado existente — fica em branco
    nos itens antigos e passa a ser preenchido dali pra frente."""
    inspector = inspect(db.engine)
    if "itens_pedido" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("itens_pedido")}
    if "liberacao_real" in colunas:
        return

    with db.engine.begin() as conn:
        conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN liberacao_real DATE"))
    app.logger.info("Migração automática: campo de liberação real adicionado aos itens.")


def _migrar_atualizado_em_itens(app):
    """Adiciona um carimbo de "última alteração" (atualizado_em) em cada item
    — o SQLAlchemy atualiza sozinho (onupdate) toda vez que o item é salvo,
    seja editando o pedido ou clicando em "Avançar" no Kanban das Estações.
    Usado pra ordenar o Kanban das Estações sempre com os itens mais
    novos/recém movimentados no topo de cada coluna (pedido do Bruno,
    25/08/2026).

    Como o campo não existia antes, faz um backfill único pros itens já
    existentes: usa a data mais avançada que o item já tem registrada
    (término de produção > início de produção > inclusão do pedido) como
    aproximação de "última atividade conhecida" — sem isso, todo item legado
    nasceria com o mesmo timestamp (o momento do deploy) e a ordenação
    ficaria arbitrária entre eles."""
    inspector = inspect(db.engine)
    if "itens_pedido" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("itens_pedido")}
    if "atualizado_em" in colunas:
        return

    with db.engine.begin() as conn:
        # TIMESTAMP, não DATETIME: "DATETIME" é aceito pelo SQLite (que ignora
        # o nome do tipo) mas não existe no Postgres de produção — quebrou o
        # primeiro deploy dessa migração (psycopg2.errors.UndefinedObject).
        conn.execute(text("ALTER TABLE itens_pedido ADD COLUMN atualizado_em TIMESTAMP"))

    itens = ItemPedido.query.options(selectinload(ItemPedido.pedido)).all()
    agora = datetime.utcnow()
    for item in itens:
        melhor_data = (
            item.termino_inspecao
            or item.liberacao_faturamento
            or item.inicio_producao
            or (item.pedido.data_inclusao_pedido if item.pedido else None)
        )
        item.atualizado_em = datetime.combine(melhor_data, datetime.min.time()) if melhor_data else agora
    db.session.commit()
    app.logger.info("Migração automática: campo atualizado_em adicionado e preenchido em %d itens.", len(itens))


# Colunas novas da Fase 13 (Gestão Operação) e seu tipo SQL — todas opcionais,
# nenhuma substitui nada que já existe em Pedido.
_COLUNAS_GESTAO_OPERACAO = {
    # Comercial
    "go_tipo_pedido": "VARCHAR(60)",
    "go_contrato": "VARCHAR(60)",
    "go_pedido_compra_cliente": "VARCHAR(60)",
    "go_proposta": "VARCHAR(60)",
    "go_data_solicitada_entrega": "DATE",
    "go_status_pedido_info": "VARCHAR(120)",
    "go_valor_pedido_operacao": "FLOAT",
    # PCP
    "go_previsao_liberacao_pcp": "DATE",
    "go_data_efetiva_liberacao_pcp": "DATE",
    "go_data_solicitada_cliente_retira": "DATE",
    "go_custo_producao_real": "FLOAT",
    "go_termino_semanal_pcp": "VARCHAR(40)",
    # Logística / NF
    "go_data_emissao_nf": "DATE",
    "go_valor_nf_emitida": "FLOAT",
    "go_numero_nf": "VARCHAR(30)",
    "go_status_logistica": "VARCHAR(60)",
    "go_data_pedido_expedido": "DATE",
    "go_transportadora_id": "INTEGER",
    "go_custo_frete_previsto": "FLOAT",
    "go_custo_frete_final": "FLOAT",
    "go_custo_frete_sobre_nota": "FLOAT",
    "go_data_prevista_entrega": "DATE",
    "go_data_real_entrega": "DATE",
    # Resultados / OTD
    "go_otd_realizado": "VARCHAR(10)",
    "go_data_solicitada_cliente_final": "DATE",
    "go_data_entregue_cliente": "DATE",
    "go_obs_operacao": "TEXT",
    "go_status_final_alinhamento": "VARCHAR(60)",
}


def _migrar_gestao_operacao_pedidos(app):
    """Adiciona as colunas novas da Fase 13 (Gestão Operação: Comercial / PCP /
    Logística-NF / Resultados-OTD) em pedidos criados antes dessa fase existir.

    Mesmo padrão das migrações anteriores: só adiciona o que ainda não existe,
    roda sozinha a cada início do site, não apaga nem altera nenhum pedido."""
    inspector = inspect(db.engine)
    if "pedidos" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("pedidos")}
    faltando = [c for c in _COLUNAS_GESTAO_OPERACAO if c not in colunas]
    if not faltando:
        return

    with db.engine.begin() as conn:
        for coluna in faltando:
            tipo = _COLUNAS_GESTAO_OPERACAO[coluna]
            conn.execute(text(f"ALTER TABLE pedidos ADD COLUMN {coluna} {tipo}"))
    app.logger.info("Migração automática: %d campos de Gestão Operação adicionados aos pedidos.", len(faltando))


def _migrar_dados_go_para_pedidos_operacao(app):
    """Backfill único: cria em `pedidos_operacao` (tabela nova, independente de
    `pedidos`) uma linha por pedido comercial, a partir dos dados de Gestão
    Operação que já existem nos `Pedido` legados (tabela de Gestão Produção).

    Motivo: até agora Gestão Operação vivia dentro da tabela `pedidos`, com o
    problema de que um mesmo pedido comercial podia corresponder a vários
    registros Pedido legados (o import histórico original criava 1 Pedido por
    LINHA de planilha, não por pedido comercial — por isso hoje existem 1146+
    Pedido mas só ~385 pedido_venda únicos). A partir de agora Gestão Operação
    e Gestão Produção são independentes: este backfill roda uma única vez
    (protegido por `PedidoOperacao.query.count() == 0`) e agrupa os Pedido
    legados por pedido_venda, usando o de menor id como representante — os
    campos go_* já estavam sincronizados entre eles (ver
    importar_gestao_operacao.py da Fase 13), então não há perda de dado.

    Os campos go_* continuam fisicamente na tabela `pedidos` depois disso
    (nada é apagado) — só o app para de lê-los/escrevê-los por ali."""
    if PedidoOperacao.query.count() > 0:
        return

    pedidos_com_dado_go = (
        Pedido.query.filter(
            or_(*(getattr(Pedido, campo).isnot(None) for campo in _COLUNAS_GESTAO_OPERACAO))
        )
        .order_by(Pedido.id)
        .all()
    )
    if not pedidos_com_dado_go:
        return

    representantes = {}
    ordem = []
    for p in pedidos_com_dado_go:
        chave = p.pedido_venda.strip() if p.pedido_venda else f"__pedido_{p.id}"
        if chave not in representantes:
            representantes[chave] = p  # primeiro da lista (menor id) = representante
            ordem.append(chave)

    for chave in ordem:
        rep = representantes[chave]
        db.session.add(
            PedidoOperacao(
                pedido_venda=rep.pedido_venda,
                cliente=rep.cliente,
                vendedor=rep.vendedor,
                data_inclusao_pedido=rep.data_inclusao_pedido,
                prioridade=rep.prioridade,
                frete=rep.frete,
                pais=rep.pais,
                estado=rep.estado,
                cidade=rep.cidade,
                **{campo: getattr(rep, campo) for campo in _COLUNAS_GESTAO_OPERACAO},
            )
        )

    db.session.commit()
    app.logger.info(
        "Backfill Gestão Operação -> tabela própria (pedidos_operacao): %d pedidos "
        "comerciais migrados (a partir de %d registros legados em `pedidos`).",
        len(ordem), len(pedidos_com_dado_go),
    )


# Marcador de "o Bruno zerou os dados de propósito, pra testar manualmente" —
# ver rota /admin/zerar-dados. Sem isso, tanto _seed_inicial (Pedido vazio)
# quanto _importar_gestao_operacao (PedidoOperacao sem go_tipo_pedido) veriam
# a tabela vazia depois do zerar e reimportariam a planilha antiga sozinhos no
# deploy seguinte — desfazendo a limpeza sem o Bruno pedir de novo.
_CHAVE_DADOS_ZERADOS_MANUALMENTE = "dados_pedidos_zerados_manualmente_28_08_2026"


def _seed_inicial(app):
    """Cria o usuário admin padrão e importa a planilha na primeira execução."""
    if Usuario.query.count() == 0:
        admin = Usuario(nome="Administrador", username="admin", role="ADMIN")
        admin.set_senha("admin123")
        db.session.add(admin)
        db.session.commit()
        app.logger.info("Usuário admin criado (login: admin / senha: admin123 — troque depois!)")

    zerado_manualmente = (
        ControleSistema.query.filter_by(chave=_CHAVE_DADOS_ZERADOS_MANUALMENTE).first() is not None
    )
    if Pedido.query.count() == 0 and not zerado_manualmente:
        xlsx_path = os.path.join(BASE_DIR, "data", "controle_producao_base.xlsx")
        if os.path.exists(xlsx_path):
            from seed import importar_planilha

            total = importar_planilha(xlsx_path)
            app.logger.info(f"{total} pedidos importados da planilha original.")


def _importar_gestao_operacao(app):
    """Importa (uma única vez) a planilha "Gestão de Fluxo Produtivo" pra dentro
    de `pedidos_operacao` (Gestão Operação — tabela própria, independente de
    `pedidos`/Gestão Produção), casando por pedido_venda.

    Roda DEPOIS de `_migrar_dados_go_para_pedidos_operacao` (o backfill a
    partir dos dados legados) — na prática isso quase sempre já deixa
    `go_tipo_pedido` preenchido em `pedidos_operacao`, então esta função só
    chega a importar de verdade da planilha se o backfill não tiver rodado
    (ex.: banco novo, sem nenhum pedido legado com dado de Gestão Operação)."""
    if PedidoOperacao.query.filter(PedidoOperacao.go_tipo_pedido.isnot(None)).first() is not None:
        return
    if ControleSistema.query.filter_by(chave=_CHAVE_DADOS_ZERADOS_MANUALMENTE).first() is not None:
        return

    xlsx_path = os.path.join(BASE_DIR, "data", "gestao_fluxo_2026.xlsx")
    if not os.path.exists(xlsx_path):
        return

    from importar_gestao_operacao import importar_gestao_operacao

    resultado = importar_gestao_operacao(xlsx_path)
    app.logger.info(
        "Importação Gestão Operação: %d linhas | %d exatas | %d aproximadas | "
        "%d pedidos novos | %d transportadoras.",
        resultado["total"],
        resultado["exato"],
        resultado["aproximado"],
        resultado["novos"],
        len(resultado["transportadoras_canonicas"]),
    )


_CHAVE_SINCRONIZACAO_25_08_2026 = "sincronizacao_planilha_producao_25_08_2026"


def _sincronizar_planilha_producao_25_08_2026(app):
    """Sincroniza Gestão Produção (Pedido/ItemPedido) com a aba "GERAL TESTE"
    da planilha enviada pelo Bruno em 25/08/2026 — atualiza pedidos/itens que
    já existem e cria os que estão na planilha mas ainda não existem no site.
    Protegido por `ControleSistema` (chave abaixo) porque, ao contrário das
    outras importações, roda por cima de dados que já existem — precisa
    rodar exatamente uma vez, mesmo com o banco de produção já povoado. Ver
    sincronizar_planilha_producao.py para as regras completas."""
    if ControleSistema.query.filter_by(chave=_CHAVE_SINCRONIZACAO_25_08_2026).first() is not None:
        return

    xlsx_path = os.path.join(BASE_DIR, "data", "sincronizacao_25_08_2026.xlsx")
    if not os.path.exists(xlsx_path):
        return

    from sincronizar_planilha_producao import sincronizar_planilha_producao

    stats = sincronizar_planilha_producao(xlsx_path)
    db.session.add(ControleSistema(chave=_CHAVE_SINCRONIZACAO_25_08_2026))
    db.session.commit()
    app.logger.info(
        "Sincronização planilha 25/08/2026: %d linhas | %d pedidos atualizados | "
        "%d pedidos criados | %d itens atualizados | %d itens criados | "
        "%d pedidos sem cliente ignorados.",
        stats["linhas_lidas"],
        stats["pedidos_atualizados"],
        stats["pedidos_criados"],
        stats["itens_atualizados"],
        stats["itens_criados"],
        len(stats["pedidos_sem_cliente_ignorados"]),
    )


_CHAVE_SINCRONIZACAO_03_09_2026 = "sincronizacao_planilha_producao_03_09_2026"


def _sincronizar_planilha_producao_03_09_2026(app):
    """Sincroniza Gestão Produção (Pedido/ItemPedido) com a aba "GERAL TESTE"
    da nova planilha enviada pelo Bruno em 03/09/2026 ("03_09 CONTROLE
    PRODUCAO_V1.xlsx", layout de colunas diferente da de 25/08 — ver
    sincronizar_planilha_producao.py) e, na sequência, cria registros básicos
    em Gestão Operação (PedidoOperacao) pros pedidos desta planilha que ainda
    não têm nenhum lá — pedido do Bruno (03/09/2026, AskUserQuestion):
    "extraia todos os dados da planilha anexa... distribua em todo o
    aplicativo, devidamente para cada area (produção e operação)", com o
    reforço explícito de NÃO tocar em Qualidade nem P&D (por isso esta função
    só chama os dois sincronizadores de Produção/Operação, nada de RNC/RDIM/
    ProjetoPD).

    Protegida por `ControleSistema` (mesmo padrão de
    `_sincronizar_planilha_producao_25_08_2026`) — roda por cima de dados que
    já existem, precisa rodar exatamente uma vez. Se a validação de
    cabeçalhos da planilha falhar (`stats["cabecalhos_invalidos"]`), NÃO
    marca o ControleSistema como concluído — loga um erro e sai, pra dar
    outra chance de rodar automaticamente depois que o mapeamento de colunas
    (COL) for corrigido, sem precisar mexer manualmente no banco. Nunca
    lança exceção: um erro aqui não pode derrubar o boot do site inteiro."""
    if ControleSistema.query.filter_by(chave=_CHAVE_SINCRONIZACAO_03_09_2026).first() is not None:
        return

    xlsx_path = os.path.join(BASE_DIR, "data", "sincronizacao_producao_03_09_2026.xlsx")
    if not os.path.exists(xlsx_path):
        return

    from sincronizar_planilha_producao import seed_pedidos_operacao_basico, sincronizar_planilha_producao

    try:
        stats = sincronizar_planilha_producao(xlsx_path)
        if stats.get("cabecalhos_invalidos"):
            app.logger.error(
                "Sincronização planilha 03/09/2026 ABORTADA — cabeçalhos da planilha não "
                "batem com o mapeamento esperado (COL): %s",
                " | ".join(stats["cabecalhos_invalidos"]),
            )
            return

        stats_operacao = seed_pedidos_operacao_basico(xlsx_path)
        if stats_operacao.get("cabecalhos_invalidos"):
            # Não deveria acontecer (mesma planilha já validou acima), mas se
            # acontecer não descarta o que a sincronização de Produção já
            # commitou — só loga e segue sem os registros de Operação.
            app.logger.error(
                "Seed de PedidoOperacao (planilha 03/09/2026) ABORTADO — cabeçalhos não bateram: %s",
                " | ".join(stats_operacao["cabecalhos_invalidos"]),
            )
            stats_operacao = {"pedidos_operacao_criados": 0, "pedidos_operacao_ja_existentes": 0}
    except Exception:
        db.session.rollback()
        app.logger.exception("Sincronização planilha 03/09/2026 falhou com erro inesperado.")
        return

    db.session.add(ControleSistema(chave=_CHAVE_SINCRONIZACAO_03_09_2026))
    db.session.commit()
    app.logger.info(
        "Sincronização planilha 03/09/2026: %d linhas | %d pedidos atualizados | "
        "%d pedidos criados | %d itens atualizados | %d itens criados | "
        "%d pedidos sem cliente ignorados | Operação: %d PedidoOperacao criados, "
        "%d já existentes.",
        stats["linhas_lidas"],
        stats["pedidos_atualizados"],
        stats["pedidos_criados"],
        stats["itens_atualizados"],
        stats["itens_criados"],
        len(stats["pedidos_sem_cliente_ignorados"]),
        stats_operacao["pedidos_operacao_criados"],
        stats_operacao["pedidos_operacao_ja_existentes"],
    )
    if stats["valores_nao_reconhecidos"]:
        app.logger.warning(
            "Sincronização planilha 03/09/2026 — valores não reconhecidos (não importados "
            "automaticamente, checar manualmente): %s",
            {campo: list(valores.keys()) for campo, valores in stats["valores_nao_reconhecidos"].items()},
        )


_CHAVE_SINCRONIZACAO_GO_28_08_2026 = "sincronizacao_gestao_operacao_28_08_2026"


def _sincronizar_gestao_operacao_28_08_2026(app):
    """Sincroniza Gestão Operação (PedidoOperacao) com a planilha "28_08 Gestão
    de Fluxo Produtivo 2026" enviada pelo Bruno — atualiza pedidos que já
    existem e cria os que estão na planilha mas ainda não existem no site.
    Protegido por `ControleSistema` (mesmo padrão de
    `_sincronizar_planilha_producao_25_08_2026`) porque roda por cima de dados
    que já existem — precisa rodar exatamente uma vez, mesmo com o banco de
    produção já povoado. Ver sincronizar_gestao_operacao.py para as regras
    completas."""
    if ControleSistema.query.filter_by(chave=_CHAVE_SINCRONIZACAO_GO_28_08_2026).first() is not None:
        return

    xlsx_path = os.path.join(BASE_DIR, "data", "sincronizacao_gestao_operacao_28_08_2026.xlsx")
    if not os.path.exists(xlsx_path):
        return

    from sincronizar_gestao_operacao import sincronizar_gestao_operacao

    stats = sincronizar_gestao_operacao(xlsx_path)
    db.session.add(ControleSistema(chave=_CHAVE_SINCRONIZACAO_GO_28_08_2026))
    db.session.commit()
    app.logger.info(
        "Sincronização Gestão Operação 28/08/2026: %d linhas | %d pedidos atualizados | "
        "%d pedidos criados | %d campos atualizados | %d casamentos exatos | %d aproximados.",
        stats["linhas_lidas"],
        stats["pedidos_atualizados"],
        stats["pedidos_criados"],
        stats["campos_atualizados"],
        stats["exato"],
        stats["aproximado"],
    )


_CHAVE_SINCRONIZACAO_GO_03_09_2026 = "sincronizacao_gestao_operacao_03_09_2026"


def _sincronizar_gestao_operacao_03_09_2026(app):
    """Sincroniza Gestão Operação (PedidoOperacao) com a planilha "03_09 Gestão
    de Fluxo Produtivo 2026" enviada pelo Bruno — mesmo formato de coluna já
    usado em 28/08/2026 (conferido coluna a coluna antes de reaproveitar
    sincronizar_gestao_operacao.py), com o bloco novo "Resultados" (valor NF/
    custo produção/custo frete, mais completo — ver _coalesce_numero em
    sincronizar_gestao_operacao.py) e rastreio de criticidade/OTD fora do
    esperado. Protegido por `ControleSistema` — roda exatamente uma vez, e
    nunca lança exceção (um erro aqui não pode derrubar o boot do site
    inteiro)."""
    if ControleSistema.query.filter_by(chave=_CHAVE_SINCRONIZACAO_GO_03_09_2026).first() is not None:
        return

    xlsx_path = os.path.join(BASE_DIR, "data", "sincronizacao_gestao_operacao_03_09_2026.xlsx")
    if not os.path.exists(xlsx_path):
        return

    from sincronizar_gestao_operacao import sincronizar_gestao_operacao

    try:
        stats = sincronizar_gestao_operacao(xlsx_path)
    except Exception:
        db.session.rollback()
        app.logger.exception("Sincronização Gestão Operação 03/09/2026 falhou com erro inesperado.")
        return

    db.session.add(ControleSistema(chave=_CHAVE_SINCRONIZACAO_GO_03_09_2026))
    db.session.commit()
    app.logger.info(
        "Sincronização Gestão Operação 03/09/2026: %d linhas | %d pedidos atualizados | "
        "%d pedidos criados | %d campos atualizados | %d casamentos exatos | %d aproximados | "
        "%d sem match de pedido_venda (viraram novos).",
        stats["linhas_lidas"],
        stats["pedidos_atualizados"],
        stats["pedidos_criados"],
        stats["campos_atualizados"],
        stats["exato"],
        stats["aproximado"],
        stats["sem_match_pedido_venda"],
    )
    if stats["criticidades_nao_reconhecidas"]:
        app.logger.warning(
            "Sincronização Gestão Operação 03/09/2026 — valores de CRITICIDADE não reconhecidos "
            "(prioridade não atualizada automaticamente nessas linhas, checar manualmente): %s",
            stats["criticidades_nao_reconhecidas"],
        )
    if stats["otd_nao_reconhecidos"]:
        app.logger.warning(
            "Sincronização Gestão Operação 03/09/2026 — valores de OTD fora do padrão SIM/NÃO "
            "(gravados como texto truncado, checar manualmente): %s",
            stats["otd_nao_reconhecidos"],
        )


_CHAVE_BACKFILL_SOLICITADA_CLIENTE_RETIRA_01_09_2026 = "backfill_go_data_solicitada_cliente_retira_01_09_2026"


def _backfill_go_data_solicitada_cliente_retira(app):
    """Correção pontual pedida pelo Bruno em 01/09/2026: "Solicitada cliente/
    retira" (PCP) tem que acompanhar a mesma data que "Data solicitada
    entrega" (Comercial) — ambas vêm de Pedido.data_cliente na criação
    automática (ver _criar_pedido_operacao_a_partir_de_producao), mas essa
    ligação só foi adicionada agora; os pedidos criados pelo fluxo automático
    ANTES desta correção ficaram com "Solicitada cliente/retira" em branco.
    Preenche só isso — nunca sobrescreve um valor que já tenha sido digitado
    manualmente em PCP (só mexe onde está None)."""
    if ControleSistema.query.filter_by(chave=_CHAVE_BACKFILL_SOLICITADA_CLIENTE_RETIRA_01_09_2026).first() is not None:
        return

    pedidos = PedidoOperacao.query.filter(
        PedidoOperacao.go_data_solicitada_cliente_retira.is_(None),
        PedidoOperacao.go_data_solicitada_entrega.isnot(None),
    ).all()
    for pedido in pedidos:
        pedido.go_data_solicitada_cliente_retira = pedido.go_data_solicitada_entrega

    app.logger.info(
        "Backfill pontual: 'Solicitada cliente/retira' preenchida em %d pedido(s) de Gestão Operação.",
        len(pedidos),
    )
    db.session.add(ControleSistema(chave=_CHAVE_BACKFILL_SOLICITADA_CLIENTE_RETIRA_01_09_2026))
    db.session.commit()


_CHAVE_CORRECAO_SEMANA_MORKEN_835 = "correcao_semana_pcp_morken_835_28_08_2026"


def _corrigir_semana_pcp_morken_835(app):
    """Correção pontual pedida pelo Bruno em 28/08/2026: o pedido 835 (Morken)
    veio da planilha com Término Semanal PCP "SEMANA 04 / AGO / 2026", mas a
    Previsão de Liberação PCP dele é 04/09/2026 (entrega solicitada 24/09) —
    ele não é um pedido de agosto. O Bruno confirmou que é pra mover pra
    setembro, então troco pra "SEMANA 01 / SET / 2026" (semana que contém o
    dia 4, mesmo critério de gerar_semanas_pcp). Com isso ele some do
    Faturamento por Semana de agosto e passa a aparecer em setembro.

    Guardado por ControleSistema — correção pontual, roda só uma vez, mesmo
    padrão das outras correções/sincronizações desta leva."""
    if ControleSistema.query.filter_by(chave=_CHAVE_CORRECAO_SEMANA_MORKEN_835).first() is not None:
        return

    pedido = PedidoOperacao.query.filter_by(pedido_venda="835").first()
    if pedido is not None and pedido.cliente and "MORKEN" in pedido.cliente.upper():
        pedido.go_termino_semanal_pcp = "SEMANA 01 / SET / 2026"
        app.logger.info(
            "Correção pontual: pedido 835 (Morken) movido de Semana 04/Ago para Semana 01/Set/2026."
        )
    else:
        # Não achou o pedido esperado (ou o cliente não bate) — não mexe em
        # nada pra não arriscar corrigir o pedido errado, só registra que já
        # tentou (pra não ficar reavaliando isso a cada boot).
        app.logger.warning(
            "Correção pontual pedido 835 (Morken): pedido não encontrado ou cliente não confere — nada foi alterado."
        )
    db.session.add(ControleSistema(chave=_CHAVE_CORRECAO_SEMANA_MORKEN_835))
    db.session.commit()


def _migrar_rdim_inspecao_final(app):
    """Adiciona os 2 campos novos da Fase 2 do RDIM (pedido do Bruno,
    02/09/2026, depois de já usar a área) na tabela `inspecoes_finais` —
    que já existe em produção desde a Fase 1, então essas colunas não são
    cobertas só por `db.create_all()` (que só cria tabelas novas). Mesmo
    padrão de `_migrar_faturamento_itens`: campos 100% novos e opcionais,
    ficam em branco nas inspeções já registradas."""
    inspector = inspect(db.engine)
    if "inspecoes_finais" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("inspecoes_finais")}
    faltando = [c for c in ("subcategoria_desvio", "quantidade_com_desvio") if c not in colunas]
    if not faltando:
        return

    with db.engine.begin() as conn:
        if "subcategoria_desvio" not in colunas:
            conn.execute(text("ALTER TABLE inspecoes_finais ADD COLUMN subcategoria_desvio VARCHAR(40)"))
        if "quantidade_com_desvio" not in colunas:
            conn.execute(text("ALTER TABLE inspecoes_finais ADD COLUMN quantidade_com_desvio FLOAT"))
    app.logger.info("Migração automática: campos subcategoria_desvio e quantidade_com_desvio adicionados em inspecoes_finais.")


def _migrar_rdim_pecas_desvio(app):
    """Adiciona os 2 campos novos da Fase 4 do RDIM (pedido do Bruno,
    02/09/2026: contexto do desvio por peça — espec. mín/máx x medido) na
    tabela `rdim_pecas_desvio` — que já existe em produção desde a Fase 3
    (feature "Apontamentos por peça"), então essas colunas não são cobertas
    só por `db.create_all()`. Mesmo padrão de `_migrar_rdim_inspecao_final`:
    campos 100% novos e opcionais, ficam em branco nos apontamentos já
    registrados."""
    inspector = inspect(db.engine)
    if "rdim_pecas_desvio" not in inspector.get_table_names():
        return

    colunas = {c["name"] for c in inspector.get_columns("rdim_pecas_desvio")}
    faltando = [c for c in ("especificado_min", "especificado_max") if c not in colunas]
    if not faltando:
        return

    with db.engine.begin() as conn:
        if "especificado_min" not in colunas:
            conn.execute(text("ALTER TABLE rdim_pecas_desvio ADD COLUMN especificado_min FLOAT"))
        if "especificado_max" not in colunas:
            conn.execute(text("ALTER TABLE rdim_pecas_desvio ADD COLUMN especificado_max FLOAT"))
    app.logger.info("Migração automática: campos especificado_min e especificado_max adicionados em rdim_pecas_desvio.")


_CHAVE_SEED_RNC_QUALIDADE_31_08_2026 = "seed_rnc_qualidade_31_08_2026"


def _seed_rnc_qualidade(app):
    """Importa (uma única vez) a planilha "Controle RNC - Qualidade" que o
    Bruno enviou em 31/08/2026 pra dentro da tabela `rnc_qualidade`, nova
    (nasce vazia — `db.create_all()` já criou a tabela nesta mesma execução
    do boot, antes desta função rodar).

    Guardado por `ControleSistema` (mesmo padrão das outras importações/
    sincronizações pontuais) em vez de só checar "tabela vazia": assim, se o
    Bruno apagar algum RNC de teste depois, o próximo boot não reimporta os
    10 RNCs de agosto por cima — roda exatamente uma vez, para sempre."""
    if ControleSistema.query.filter_by(chave=_CHAVE_SEED_RNC_QUALIDADE_31_08_2026).first() is not None:
        return

    xlsx_path = os.path.join(BASE_DIR, "data", "rnc_qualidade_31_08_2026.xlsx")
    if not os.path.exists(xlsx_path):
        return

    from seed_rnc_qualidade import importar_rnc_qualidade

    total = importar_rnc_qualidade(xlsx_path)
    db.session.add(ControleSistema(chave=_CHAVE_SEED_RNC_QUALIDADE_31_08_2026))
    db.session.commit()
    app.logger.info("Importação Qualidade/RNC: %d RNCs importados da planilha 31/08/2026.", total)


def _parse_data_form(valor):
    if not valor:
        return None
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_float_form(valor, default=0.0):
    if valor in (None, ""):
        return default
    try:
        return float(str(valor).replace(",", "."))
    except ValueError:
        return default


def _registrar_alteracoes(entidade_tipo, entidade_id, pedido_id, antes, depois, campos):
    """Compara os valores de "antes"/"depois" de uma lista de campos e grava uma
    linha no histórico para cada campo que realmente mudou de valor. Gravado na
    mesma sessão do banco que a alteração em si — só vira permanente quando o
    resto da rota der commit()."""
    for campo in campos:
        v_antes = antes.get(campo)
        v_depois = depois.get(campo)
        if v_antes == v_depois:
            continue
        db.session.add(
            HistoricoAlteracao(
                entidade_tipo=entidade_tipo,
                entidade_id=entidade_id,
                pedido_id=pedido_id,
                usuario_id=current_user.id if current_user.is_authenticated else None,
                usuario_nome=current_user.nome if current_user.is_authenticated else None,
                campo=campo,
                valor_anterior=None if v_antes is None else str(v_antes),
                valor_novo=None if v_depois is None else str(v_depois),
            )
        )


def _predicado_status(status):
    """Traduz o status "calculado" do pedido (Pedido.status_producao) para uma
    condição de SQL equivalente, para poder filtrar/contar direto no banco em
    vez de carregar todo mundo em Python (como a tela inicial fazia antes).

    Precisa reproduzir exatamente as mesmas regras da property Pedido.status_producao:
      - algum item "EM TRATATIVA"   -> EM TRATATIVA (tem prioridade sobre o resto)
      - todos os itens "FINALIZADO" -> FINALIZADO
      - todos os itens "PENDENTE" (ou pedido sem item) -> PENDENTE
      - qualquer outra mistura      -> ANDAMENTO
    """
    tem_itens = Pedido.itens.any()

    if status == "EM TRATATIVA":
        return Pedido.itens.any(ItemPedido.status_producao == "EM TRATATIVA")

    if status == "FINALIZADO":
        return and_(tem_itens, ~Pedido.itens.any(ItemPedido.status_producao != "FINALIZADO"))

    if status == "PENDENTE":
        return or_(~tem_itens, and_(tem_itens, ~Pedido.itens.any(ItemPedido.status_producao != "PENDENTE")))

    if status == "ANDAMENTO":
        return and_(
            tem_itens,
            ~Pedido.itens.any(ItemPedido.status_producao == "EM TRATATIVA"),
            Pedido.itens.any(ItemPedido.status_producao != "FINALIZADO"),
            Pedido.itens.any(ItemPedido.status_producao != "PENDENTE"),
        )

    return None


def _predicado_atrasado():
    """Pedido "atrasado": tem pelo menos um item com liberação prevista já
    vencida e que ainda não foi finalizado — mesma regra do semáforo vermelho."""
    return Pedido.itens.any(
        and_(ItemPedido.liberacao_prevista < date.today(), ItemPedido.status_producao != "FINALIZADO")
    )


def _calcular_resumo():
    """Contagens gerais (cards do topo) calculadas direto no banco — antes esta
    função carregava TODOS os pedidos em Python a cada acesso à tela inicial."""
    total = Pedido.query.count()
    pendente = Pedido.query.filter(_predicado_status("PENDENTE")).count()
    em_tratativa = Pedido.query.filter(_predicado_status("EM TRATATIVA")).count()
    andamento = Pedido.query.filter(_predicado_status("ANDAMENTO")).count()
    finalizado = Pedido.query.filter(_predicado_status("FINALIZADO")).count()
    valor_total = db.session.query(func.sum(ItemPedido.quantidade * ItemPedido.custo_unitario)).scalar() or 0.0
    return {
        "total": total,
        "pendente": pendente,
        "em_tratativa": em_tratativa,
        "andamento": andamento,
        "finalizado": finalizado,
        "valor_total": round(valor_total, 2),
    }


def _calcular_resumo_filtrado(linhas):
    """Mesmos cards de resumo de _calcular_resumo (topo da Listagem Geral),
    mas recalculados em cima do que está de fato filtrado/visível na tela —
    pedido do Bruno (03/09/2026): "quero que TODO esse painel seja
    totalmente dinâmico... se eu filtrar algo como mês, região, estação,
    status, quero que automaticamente ele atualize". Antes esses 6 números
    vinham sempre de _calcular_resumo() (banco inteiro, sem filtro nenhum) —
    por isso nunca mudavam ao filtrar a tabela abaixo.

    `linhas` é a lista já achatada (1 linha por item, ver _LinhaListagemGeral)
    depois de TODOS os filtros da Listagem Geral já aplicados. Total de
    pedidos/Pendentes/Em tratativa/Em andamento/Finalizados contam PEDIDOS
    distintos entre essas linhas — usando o status "rollup" de cada pedido
    (Pedido.status_producao, o mesmo critério já usado pelo próprio filtro de
    Status em _predicado_status), não o status do item individual. Valor
    total soma só os ITENS que aparecem nas linhas (bate exatamente com a
    tabela logo abaixo — se um pedido tem 5 itens e só 2 passaram no filtro
    de estação, por exemplo, conta só esses 2, não o pedido inteiro).

    Calculado em Python sobre objetos que a rota já carregou (nenhuma query
    nova ao banco) — os mesmos `pedido`/`item` que _linhas_listagem_geral
    usou pra montar as linhas."""
    pedidos_vistos = {}
    valor_total = 0.0
    for l in linhas:
        pedidos_vistos[l.pedido_id] = l.pedido
        valor_total += l.venda_total or 0

    contagem = {"PENDENTE": 0, "EM TRATATIVA": 0, "ANDAMENTO": 0, "FINALIZADO": 0}
    for pedido in pedidos_vistos.values():
        status = pedido.status_producao
        if status in contagem:
            contagem[status] += 1

    return {
        "total": len(pedidos_vistos),
        "pendente": contagem["PENDENTE"],
        "em_tratativa": contagem["EM TRATATIVA"],
        "andamento": contagem["ANDAMENTO"],
        "finalizado": contagem["FINALIZADO"],
        "valor_total": round(valor_total, 2),
    }


def _resumo_otd(query=None):
    """Estatísticas de OTD (On-Time Delivery) da Gestão Operação — usa o campo
    go_otd_realizado (SIM/NÃO preenchido manualmente na planilha/tela), bem
    mais confiável que o _otd_percentual() antigo (que depende de datas quase
    nunca preenchidas no histórico). Só considera pedidos com OTD preenchido —
    quem ainda não tem essa informação fica de fora do percentual (não conta
    como "não cumpriu").

    `query` (opcional): uma query de PedidoOperacao já filtrada (ver
    _filtrar_pedidos_operacao) — pedido do Bruno (03/09/2026): filtrar
    Resultados/OTD por mês/segmento (planejamento semanal PCP / faturados) e
    ver o OTD só desse recorte. Sem argumento, mantém o comportamento antigo
    (todos os pedidos).

    Gestão Operação tem tabela própria (PedidoOperacao) — cada linha já é um
    pedido comercial, sem precisar agrupar nada em tempo de execução."""
    base = query if query is not None else PedidoOperacao.query
    pedidos = base.filter(PedidoOperacao.go_otd_realizado.isnot(None)).all()

    total = len(pedidos)
    no_prazo = sum(1 for p in pedidos if p.go_otd_realizado == "SIM")
    percentual = round((no_prazo / total) * 100, 1) if total else None

    def _quebra_por(atributo):
        contagem = {}
        for p in pedidos:
            chave = getattr(p, atributo)
            if not chave:
                continue
            c = contagem.setdefault(chave, {"total": 0, "no_prazo": 0})
            c["total"] += 1
            if p.go_otd_realizado == "SIM":
                c["no_prazo"] += 1
        top10 = sorted(contagem.items(), key=lambda kv: -kv[1]["total"])[:10]
        return [
            {
                "chave": chave,
                "total": v["total"],
                "no_prazo": v["no_prazo"],
                "percentual": round((v["no_prazo"] / v["total"]) * 100, 1) if v["total"] else None,
            }
            for chave, v in top10
        ]

    return {
        "total": total,
        "no_prazo": no_prazo,
        "percentual": percentual,
        "atinge_meta": (percentual is not None and percentual >= GO_OTD_META_PERCENTUAL),
        "por_vendedor": _quebra_por("vendedor"),
        "por_cliente": _quebra_por("cliente"),
    }


def _media_dias(valores):
    """Média simples de uma lista de inteiros/None, ignorando os None — mesmo
    critério do resto do resumo de Resultados/OTD: quem não tem o dado
    calculável fica de fora da média (não conta como zero). Devolve
    (média arredondada em 1 casa, quantidade de pedidos que entraram na
    conta) — a quantidade é sempre mostrada junto do número no card (pedido
    do Bruno: "bem claro, didático"), pra deixar claro quando a média vem de
    poucos pedidos."""
    validos = [v for v in valores if v is not None]
    if not validos:
        return None, 0
    return round(sum(validos) / len(validos), 1), len(validos)


def _resumo_lead_times(query):
    """Lead times médios da Gestão Operação pra tela Resultados/OTD — pedido
    do Bruno (03/09/2026): "preciso ver os resultados detalhados de cada
    mês, como otd, lead time operação, lead time chão de fábrica (produção),
    lead time operação cif, lead time operação fob". Recebe a MESMA query já
    filtrada por mês/segmento usada em _resumo_otd (ver
    _filtrar_pedidos_operacao) — cada card do resumo reflete o mesmo recorte.

      - lt_operacao: PedidoOperacao.go_lead_time_operacao_dias (inclusão do
        pedido até entrega efetiva no cliente) — geral, depois quebrado por
        modalidade de frete (CIF/FOB, ver FRETE_OPCOES).
      - lt_frete: PedidoOperacao.go_lead_time_frete_dias (expedição até
        entrega/coleta real).
      - lt_producao: "chão de fábrica" — vem de Gestão Produção (ItemPedido.
        lt_producao_dias: início produção até término inspeção), cruzado
        pelo mesmo casamento por pedido_venda (trim, sem FK, nunca
        aproximado) já usado em _itens_producao_por_pedido_venda. Pedidos
        que ainda não foram lançados em Gestão Produção, ou cujos itens
        ainda não têm as duas datas, ficam de fora da média."""
    pedidos = query.all()

    lt_operacao_media, lt_operacao_n = _media_dias(p.go_lead_time_operacao_dias for p in pedidos)
    lt_operacao_cif_media, lt_operacao_cif_n = _media_dias(
        p.go_lead_time_operacao_dias for p in pedidos if p.frete == "CIF"
    )
    lt_operacao_fob_media, lt_operacao_fob_n = _media_dias(
        p.go_lead_time_operacao_dias for p in pedidos if p.frete == "FOB"
    )
    lt_frete_media, lt_frete_n = _media_dias(p.go_lead_time_frete_dias for p in pedidos)

    itens_por_pedido = _itens_producao_por_pedido_venda([p.pedido_venda for p in pedidos])
    valores_lt_producao = [
        item.lt_producao_dias for itens in itens_por_pedido.values() for item in itens
    ]
    lt_producao_media, lt_producao_n = _media_dias(valores_lt_producao)

    return {
        "lt_operacao": {"media": lt_operacao_media, "n": lt_operacao_n},
        "lt_operacao_cif": {"media": lt_operacao_cif_media, "n": lt_operacao_cif_n},
        "lt_operacao_fob": {"media": lt_operacao_fob_media, "n": lt_operacao_fob_n},
        "lt_frete": {"media": lt_frete_media, "n": lt_frete_n},
        "lt_producao": {"media": lt_producao_media, "n": lt_producao_n},
    }


def _semana_label_curto(semana):
    """Encurta "SEMANA 03 / AGO / 2026" pra "S03 AGO/26" (rótulo do eixo do
    gráfico de projeção) — valores antigos que não seguem esse padrão (ex.:
    "DEZEMBRO/2025", digitado à mão antes desta convenção existir) aparecem
    como estão, sem tentar encurtar."""
    m = re.match(r"SEMANA (\d{2}) / (\w{3}) / (\d{4})", semana or "")
    if not m:
        return semana
    semana_num, mes, ano = m.groups()
    return f"S{semana_num} {mes}/{ano[2:]}"


def _projecao_pcp():
    """Projeção de carga por semana de PCP (Painel) — quantos PEDIDOS
    COMERCIAIS estão marcados pra cada "Término Semanal PCP" (Gestão
    Operação), separando o que já foi finalizado do que ainda está em aberto.
    Usa a mesma lista/ordem cronológica de gerar_semanas_pcp() (1 mês atrás
    até 6 meses à frente) — semanas sem nenhum pedido nas pontas são cortadas
    pra não poluir o gráfico."""
    semanas = gerar_semanas_pcp()

    pedidos = PedidoOperacao.query.filter(PedidoOperacao.go_termino_semanal_pcp.in_(semanas)).all()

    contagem = {}
    for p in pedidos:
        c = contagem.setdefault(p.go_termino_semanal_pcp, {"finalizado": 0, "em_aberto": 0})
        if p.status_producao == "FINALIZADO":
            c["finalizado"] += 1
        else:
            c["em_aberto"] += 1

    linhas = [
        {
            "semana": s,
            "semana_curta": _semana_label_curto(s),
            "finalizado": contagem.get(s, {}).get("finalizado", 0),
            "em_aberto": contagem.get(s, {}).get("em_aberto", 0),
        }
        for s in semanas
    ]
    while linhas and not linhas[0]["finalizado"] and not linhas[0]["em_aberto"]:
        linhas.pop(0)
    while linhas and not linhas[-1]["finalizado"] and not linhas[-1]["em_aberto"]:
        linhas.pop()
    return linhas


# Nomes de mês (abreviados e por extenso) que já apareceram no campo Término
# Semanal PCP — cobre tanto o formato novo ("SEMANA 04 / JUL / 2026") quanto
# valores antigos digitados à mão antes dessa convenção existir ("DEZEMBRO/2025").
_MESES_NUM_PCP = {
    "JAN": 1, "JANEIRO": 1,
    "FEV": 2, "FEVEREIRO": 2,
    "MAR": 3, "MARÇO": 3, "MARCO": 3,
    "ABR": 4, "ABRIL": 4,
    "MAI": 5, "MAIO": 5,
    "JUN": 6, "JUNHO": 6,
    "JUL": 7, "JULHO": 7,
    "AGO": 8, "AGOSTO": 8,
    "SET": 9, "SETEMBRO": 9,
    "OUT": 10, "OUTUBRO": 10,
    "NOV": 11, "NOVEMBRO": 11,
    "DEZ": 12, "DEZEMBRO": 12,
}


def _mes_ano_da_semana_pcp(semana):
    """Extrai (ano, mês) de dentro do valor de Término Semanal PCP, não
    importa se é "SEMANA 04 / JUL / 2026" (formato novo) ou "DEZEMBRO/2025"
    (formato antigo, digitado à mão) — devolve None se não conseguir entender."""
    if not semana:
        return None
    m = re.search(r"([A-ZÇÃÕ]+)\s*/\s*(\d{4})", semana.upper())
    if not m:
        return None
    nome_mes, ano = m.groups()
    mes = _MESES_NUM_PCP.get(nome_mes)
    if not mes:
        return None
    return (int(ano), mes)


def _chave_semana_pcp(semana):
    """Chave de ordenação cronológica pro rótulo de semana (ano, mês, nº da
    semana) — usada tanto pelo planejamento semanal da Listagem Geral quanto,
    se precisar no futuro, por qualquer outro campo no mesmo formato."""
    if not semana:
        return None
    mes_ano = _mes_ano_da_semana_pcp(semana)
    m = re.search(r"SEMANA\s*(\d+)", semana.upper())
    semana_num = int(m.group(1)) if m else 0
    if mes_ano is None:
        return (9999, 99, semana_num)
    ano, mes = mes_ano
    return (ano, mes, semana_num)


def _somar_meses(ano, mes, delta):
    total = (ano * 12 + (mes - 1)) + delta
    return (total // 12, total % 12 + 1)


# ---------------------------------------------------------------------------
# Calendário PCP da tela de Programação (pedido do Bruno, 31/08/2026) — semana
# de verdade (domingo a sábado), só pra esta tela. NÃO usa nem mexe no padrão
# "SEMANA NN / MÊS / ANO" (dia 1-7, 8-14...) usado em Gestão Operação,
# Faturamento por Semana e no Planejamento Semanal da Listagem Geral — esses
# três continuam exatamente como estão, comparando pelo texto já gravado nos
# pedidos importados da planilha. Misturar as duas convenções faria pedido
# sumir de semana sem ninguém perceber (testado: quase metade dos meses tem
# número de semanas diferente entre os dois padrões), por isso são
# propositalmente independentes.
def _semanas_calendario_pcp(ano, mes):
    """Devolve as semanas (domingo a sábado) cujo domingo de início cai no
    mês/ano dado, cada uma como {"numero", "inicio", "fim"}. Uma semana pode
    terminar no mês seguinte (ex.: semana 5 de agosto/2026 vai até 05/09) —
    isso é esperado, é só o card ficando "encostado" no mês seguinte."""
    primeiro_dia = date(ano, mes, 1)
    dias_ate_domingo = (6 - primeiro_dia.weekday()) % 7  # weekday(): segunda=0 ... domingo=6
    domingo = primeiro_dia + timedelta(days=dias_ate_domingo)
    semanas = []
    numero = 1
    while domingo.year == ano and domingo.month == mes:
        semanas.append({"numero": numero, "inicio": domingo, "fim": domingo + timedelta(days=6)})
        numero += 1
        domingo += timedelta(days=7)
    return semanas


def _parse_mes_ano_form(valor, default):
    """Lê um <input type=month> (formato "YYYY-MM") — devolve `default` se
    vier vazio ou num formato que não reconhece."""
    if not valor:
        return default
    try:
        ano_s, mes_s = valor.split("-")
        ano, mes = int(ano_s), int(mes_s)
        if 1 <= mes <= 12:
            return (ano, mes)
    except (ValueError, AttributeError):
        pass
    return default


def _projecao_pcp_mensal(mes_de, mes_ate):
    """Soma a projeção semanal de PCP por MÊS (soma de todas as semanas
    dentro do mês), separando o que já foi finalizado do que ainda está em
    aberto — em quantidade de PEDIDOS COMERCIAIS e em valor (R$, a partir de
    go_valor_pedido_operacao — já é o total do pedido, não soma nenhum item).
    `mes_de`/`mes_ate` são tuplas (ano, mês), intervalo fechado."""
    pedidos = PedidoOperacao.query.filter(PedidoOperacao.go_termino_semanal_pcp.isnot(None)).all()

    baldes = {}
    for p in pedidos:
        chave = _mes_ano_da_semana_pcp(p.go_termino_semanal_pcp)
        if chave is None or not (mes_de <= chave <= mes_ate):
            continue
        b = baldes.setdefault(chave, {"pedidos_fin": 0, "valor_fin": 0.0, "pedidos_aberto": 0, "valor_aberto": 0.0})
        valor = p.go_valor_pedido_operacao or 0.0
        if p.status_producao == "FINALIZADO":
            b["pedidos_fin"] += 1
            b["valor_fin"] += valor
        else:
            b["pedidos_aberto"] += 1
            b["valor_aberto"] += valor

    linhas = []
    ano, mes = mes_de
    while (ano, mes) <= mes_ate:
        b = baldes.get((ano, mes), {"pedidos_fin": 0, "valor_fin": 0.0, "pedidos_aberto": 0, "valor_aberto": 0.0})
        linhas.append(
            {
                "ano": ano,
                "mes": mes,
                "label": f"{MESES_PT[mes - 1]}/{ano}",
                "pedidos_finalizados": b["pedidos_fin"],
                "valor_finalizado": round(b["valor_fin"], 2),
                "pedidos_em_aberto": b["pedidos_aberto"],
                "valor_em_aberto": round(b["valor_aberto"], 2),
                "pedidos_total": b["pedidos_fin"] + b["pedidos_aberto"],
                "valor_total_mes": round(b["valor_fin"] + b["valor_aberto"], 2),
            }
        )
        ano, mes = _somar_meses(ano, mes, 1)
    return linhas


_PERIODO_TIPOS = ("mes", "tri", "sem", "ano")


def _parse_periodo(valor_str):
    """Decodifica o parâmetro `periodo` da tela Resultados/OTD — pedido do
    Bruno (03/09/2026): "inclua em formato de lista... além do mês, inclua
    também 1º trimestre, 2, 3, 4 trimestre, 1º e segundo semestre... e o
    filtro geral (incluindo todos os períodos/meses)". Formatos aceitos:
    "todos", "mes-AAAA-M", "tri-AAAA-T" (T 1-4), "sem-AAAA-S" (S 1-2),
    "ano-AAAA". Sem valor reconhecível, cai no mês atual (mesmo default de
    sempre). Devolve (tipo, ano, valor, label) — `valor` é None pra
    tipo "ano" e pra "todos"."""
    hoje = date.today()
    padrao_mes_atual = ("mes", hoje.year, hoje.month, f"{MESES_PT[hoje.month - 1]}/{hoje.year}")

    if valor_str == "todos":
        return "todos", None, None, "Todos os períodos"

    m = re.match(r"^(mes|tri|sem|ano)-(\d{4})(?:-(\d+))?$", valor_str or "")
    if not m:
        return padrao_mes_atual
    tipo, ano_s, valor_s = m.groups()
    ano = int(ano_s)
    valor = int(valor_s) if valor_s else None

    if tipo == "mes" and valor and 1 <= valor <= 12:
        return "mes", ano, valor, f"{MESES_PT[valor - 1]}/{ano}"
    if tipo == "tri" and valor and 1 <= valor <= 4:
        return "tri", ano, valor, f"{valor}º Trimestre/{ano}"
    if tipo == "sem" and valor and 1 <= valor <= 2:
        return "sem", ano, valor, f"{valor}º Semestre/{ano}"
    if tipo == "ano":
        return "ano", ano, None, f"Ano {ano}"
    return padrao_mes_atual


def _periodo_para_str(tipo, ano, valor):
    """Inverso de _parse_periodo (sem o label) — monta a string `periodo`
    a partir de (tipo, ano, valor), pros links/campos ocultos do
    formulário."""
    if tipo == "todos":
        return "todos"
    if tipo == "ano":
        return f"ano-{ano}"
    return f"{tipo}-{ano}-{valor}"


def _periodo_vizinho(tipo, ano, valor, direcao):
    """String `periodo` do período vizinho (direcao -1 ou +1), na MESMA
    granularidade — usado nos botões "Período anterior"/"Próximo período"
    da tela Resultados/OTD, que agora navegam mês a mês, trimestre a
    trimestre, semestre a semestre ou ano a ano, dependendo do que está
    selecionado. "todos" não tem vizinho (botões ficam escondidos)."""
    if tipo == "mes":
        m, a = valor + direcao, ano
        if m < 1:
            m, a = 12, ano - 1
        elif m > 12:
            m, a = 1, ano + 1
        return _periodo_para_str("mes", a, m)
    if tipo == "tri":
        t, a = valor + direcao, ano
        if t < 1:
            t, a = 4, ano - 1
        elif t > 4:
            t, a = 1, ano + 1
        return _periodo_para_str("tri", a, t)
    if tipo == "sem":
        s, a = valor + direcao, ano
        if s < 1:
            s, a = 2, ano - 1
        elif s > 2:
            s, a = 1, ano + 1
        return _periodo_para_str("sem", a, s)
    if tipo == "ano":
        return _periodo_para_str("ano", ano + direcao, None)
    return None


def _opcoes_periodo():
    """Lista de opções pro dropdown de período da tela Resultados/OTD,
    agrupada por ano (Ano atual - 1 até Ano atual + 1) — cada grupo com Ano
    inteiro, os 2 semestres, os 4 trimestres e os 12 meses, mais "Todos os
    períodos" no topo, fora de qualquer grupo."""
    hoje = date.today()
    grupos = []
    for ano in range(hoje.year - 1, hoje.year + 2):
        itens = [{"valor": _periodo_para_str("ano", ano, None), "label": f"Ano {ano}"}]
        itens += [
            {"valor": _periodo_para_str("sem", ano, s), "label": f"{s}º Semestre {ano}"} for s in (1, 2)
        ]
        itens += [
            {"valor": _periodo_para_str("tri", ano, t), "label": f"{t}º Trimestre {ano}"} for t in (1, 2, 3, 4)
        ]
        itens += [
            {"valor": _periodo_para_str("mes", ano, m), "label": f"{MESES_PT[m - 1]}/{ano}"} for m in range(1, 13)
        ]
        grupos.append({"ano": ano, "itens": itens})
    return grupos


def _semanas_do_periodo(tipo, ano, valor):
    """Lista de rótulos de semana PCP (formato de gerar_semanas_pcp) cobertos
    por um período mes/tri/sem/ano — não trata "todos" (ver
    _filtrar_por_periodo_pcp, que pra "todos" não filtra nada)."""
    if tipo == "mes":
        meses = [valor]
    elif tipo == "tri":
        meses = list(range((valor - 1) * 3 + 1, (valor - 1) * 3 + 4))
    elif tipo == "sem":
        meses = list(range(1, 7)) if valor == 1 else list(range(7, 13))
    else:  # "ano"
        meses = list(range(1, 13))
    semanas = []
    for m in meses:
        semanas.extend(gerar_semanas_pcp(meses_atras=0, meses_frente=0, hoje=date(ano, m, 1)))
    return semanas


def _filtrar_por_periodo_pcp(query, tipo, ano, valor):
    """Aplica (ou não) o filtro de Término Semanal PCP num query de
    PedidoOperacao, de acordo com o período — "todos" não filtra nada
    (literalmente todos os pedidos, tenham ou não Término Semanal PCP
    definido; pedido do Bruno, 03/09/2026: "o filtro geral, incluindo
    todos os períodos/meses")."""
    if tipo == "todos":
        return query
    semanas = _semanas_do_periodo(tipo, ano, valor)
    return query.filter(PedidoOperacao.go_termino_semanal_pcp.in_(semanas))


def _pedidos_operacao_do_periodo(tipo, ano, valor):
    """Query de PedidoOperacao filtrada pelo período (ver
    _filtrar_por_periodo_pcp), direta (sem passar pelos filtros de
    cliente/vendedor/busca da tela). Usada pelo resumo fixo do topo da tela
    Resultados/OTD (pedido do Bruno, 03/09/2026) — sempre mostra o período
    selecionado, independente do dropdown de segmento mais abaixo na
    página."""
    return _filtrar_por_periodo_pcp(PedidoOperacao.query, tipo, ano, valor)


def _faturamento_por_periodo(tipo, ano, valor):
    """Generaliza o antigo "Faturamento por Semana" (só um mês) pra
    qualquer período — mês, trimestre, semestre, ano ou "todos" — pedido do
    Bruno em 28/08/2026 (base) e 03/09/2026 (dropdown de período). Uma
    linha por semana de PCP dentro do período; pra "todos", usa toda semana
    que realmente tem algum pedido (em vez de gerar um calendário sem fim).

    Revisado em 28/08/2026 depois que o Bruno conferiu os números da semana
    04/Ago manualmente: a versão original desta função exigia Data Efetiva de
    Liberação PCP preenchida pra contar "Qtd/Valor liberado" (replicando a
    fórmula SUMPRODUCT/IFERROR que eu tinha lido da planilha) — mas o Bruno
    confirmou que quer TODO pedido cujo Término Semanal PCP cai naquela
    semana, tenha ele já sido liberado ou não (ver AskUserQuestion — "Somar
    todos os pedidos da semana, liberados ou não"). Ou seja, "liberado" aqui
    não significa "já com Data Efetiva de Liberação preenchida", e sim
    "planejado pro PCP encerrar naquela semana" — mesmo agrupamento por
    Término Semanal PCP (coluna AA), só que sem o filtro extra que eu tinha
    adicionado por conta própria.

      - Qtd/Valor liberado: todo pedido cujo Término Semanal PCP
        (go_termino_semanal_pcp) cai nessa semana — valor é a soma de
        go_valor_pedido_operacao (valor total do pedido).
      - Qtd/Valor faturado: pedidos cujo Término Semanal PCP cai nessa
        semana, somando go_valor_nf_emitida.

    Diferença deliberada da planilha do Bruno: lá, valores de "VALOR NF
    EMITIDA" digitados em formato brasileiro com vírgula decimal (texto, não
    número) são silenciosamente zerados pelo IFERROR(...*1, 0) da fórmula
    dele — aqui esses valores são interpretados corretamente como número
    (mesmo parser usado no resto do site, ver _parse_numero em
    importar_gestao_operacao.py), então o total pode ficar um pouco MAIOR
    que o da planilha nesses meses com célula de texto — reportado ao Bruno
    junto com a entrega.

    Cada linha também carrega a lista de `pedidos` daquela semana (pedido do
    Bruno, 03/09/2026: clicar na semana e abrir os pedidos dela, sem sair da
    página) — já vem pronta daqui (sem N+1) porque os pedidos da semana já
    são carregados de qualquer forma pra somar qtd/valor."""
    if tipo == "todos":
        pedidos_todos = (
            PedidoOperacao.query.filter(PedidoOperacao.go_termino_semanal_pcp.isnot(None))
            .order_by(PedidoOperacao.pedido_venda)
            .all()
        )
        semanas = sorted({p.go_termino_semanal_pcp for p in pedidos_todos}, key=_chave_semana_pcp)
        pedidos = pedidos_todos
    else:
        semanas = _semanas_do_periodo(tipo, ano, valor)
        pedidos = (
            PedidoOperacao.query.filter(PedidoOperacao.go_termino_semanal_pcp.in_(semanas))
            .order_by(PedidoOperacao.pedido_venda)
            .all()
        )

    baldes = {
        s: {"qtd_liberada": 0, "valor_liberado": 0.0, "qtd_faturada": 0, "valor_faturado": 0.0, "pedidos": []}
        for s in semanas
    }
    for p in pedidos:
        b = baldes[p.go_termino_semanal_pcp]
        b["qtd_liberada"] += 1
        b["valor_liberado"] += p.go_valor_pedido_operacao or 0.0
        b["pedidos"].append(p)
        if p.go_valor_nf_emitida:
            b["qtd_faturada"] += 1
            b["valor_faturado"] += p.go_valor_nf_emitida

    linhas = [
        {
            "semana": s,
            "semana_curta": _semana_label_curto(s),
            "qtd_liberada": baldes[s]["qtd_liberada"],
            "valor_liberado": round(baldes[s]["valor_liberado"], 2),
            "qtd_faturada": baldes[s]["qtd_faturada"],
            "valor_faturado": round(baldes[s]["valor_faturado"], 2),
            "pedidos": baldes[s]["pedidos"],
        }
        for s in semanas
    ]
    totais = {
        "qtd_liberada": sum(l["qtd_liberada"] for l in linhas),
        "valor_liberado": round(sum(l["valor_liberado"] for l in linhas), 2),
        "qtd_faturada": sum(l["qtd_faturada"] for l in linhas),
        "valor_faturado": round(sum(l["valor_faturado"] for l in linhas), 2),
    }
    return {"linhas": linhas, "totais": totais}


def _otd_mensal_ano(ano):
    """OTD (Gestão Operação, go_otd_realizado) mês a mês, Jan-Dez de um ano —
    pedido do Bruno (03/09/2026): gráfico de OTD mensal no Painel, mesma
    meta mínima da tela Resultados/OTD (GO_OTD_META_PERCENTUAL, 78%). Usa o
    mesmo recorte por Término Semanal PCP já usado em Resultados/OTD
    (_pedidos_operacao_do_periodo/_resumo_otd), só resumido mês a mês pro
    ano inteiro."""
    resultado = []
    for mes in range(1, 13):
        query = _pedidos_operacao_do_periodo("mes", ano, mes)
        otd = _resumo_otd(query)
        resultado.append({
            "mes": MESES_PT[mes - 1],
            "percentual": otd["percentual"],
            "total": otd["total"],
            "atinge_meta": otd["atinge_meta"],
        })
    return resultado


def _predicado_vencendo():
    """Pedido "vencendo": tem item com liberação prevista nos próximos
    PRAZO_ALERTA_DIAS dias (e ainda não atrasado nenhum item) — mesma regra do
    semáforo amarelo."""
    limite = date.today() + timedelta(days=PRAZO_ALERTA_DIAS)
    tem_item_vencendo = Pedido.itens.any(
        and_(
            ItemPedido.liberacao_prevista.isnot(None),
            ItemPedido.liberacao_prevista >= date.today(),
            ItemPedido.liberacao_prevista <= limite,
            ItemPedido.status_producao != "FINALIZADO",
        )
    )
    return and_(tem_item_vencendo, ~_predicado_atrasado())


def _faturamento_mes(ano, mes, filtros=None):
    """Soma o valor dos itens com liberação PREVISTA (usa quantidade × custo)
    e com liberação REALIZADA (usa o valor faturado de verdade quando
    preenchido, senão cai pro quantidade × custo) dentro de um mês — usado no
    previsto × realizado. `filtros` (opcional) é uma lista de condições extras
    (ex.: cliente/região/vendedor) aplicadas a ambas as somas via join com Pedido."""
    inicio = date(ano, mes, 1)
    fim = date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1)
    valor_calculado = ItemPedido.quantidade * ItemPedido.custo_unitario
    valor_realizado = func.coalesce(ItemPedido.valor_faturado, valor_calculado)
    filtros = filtros or []

    previsto = (
        db.session.query(func.sum(valor_calculado))
        .join(Pedido, ItemPedido.pedido_id == Pedido.id)
        .filter(ItemPedido.liberacao_prevista >= inicio, ItemPedido.liberacao_prevista < fim, *filtros)
        .scalar()
        or 0.0
    )
    realizado = (
        db.session.query(func.sum(valor_realizado))
        .join(Pedido, ItemPedido.pedido_id == Pedido.id)
        .filter(ItemPedido.liberacao_faturamento >= inicio, ItemPedido.liberacao_faturamento < fim, *filtros)
        .scalar()
        or 0.0
    )
    return round(previsto, 2), round(realizado, 2)


def _faturamento_tendencia(meses=6):
    """Previsto × realizado dos últimos N meses (incluindo o atual), para o gráfico de tendência."""
    hoje = date.today()
    pontos = []
    ano, mes = hoje.year, hoje.month
    for _ in range(meses):
        pontos.append((ano, mes))
        mes -= 1
        if mes == 0:
            mes, ano = 12, ano - 1
    pontos.reverse()

    resultado = []
    for ano, mes in pontos:
        previsto, realizado = _faturamento_mes(ano, mes)
        resultado.append({"mes": f"{MESES_PT[mes - 1]}/{ano}", "previsto": previsto, "realizado": realizado})
    return resultado


def _faturamento_mensal_ano(ano):
    """Faturamento REALIZADO mês a mês, Jan-Dez de um ano — pedido do Bruno
    (03/09/2026): trocar o gráfico "Faturamento previsto × realizado (6
    meses)" do Painel por uma visão anual, só do realizado (o previsto
    continua disponível no card "Faturamento previsto (mês)" acima e na tela
    de Faturamento). Reaproveita _faturamento_mes (mesma fonte/critério de
    sempre: liberação de faturamento dentro do mês, valor faturado real
    quando preenchido, senão quantidade × custo)."""
    resultado = []
    for mes in range(1, 13):
        _, realizado = _faturamento_mes(ano, mes)
        resultado.append({"mes": MESES_PT[mes - 1], "realizado": realizado})
    return resultado


def _lead_time_medio_dias():
    """Média de dias entre início de produção e término de inspeção/embalagem,
    entre os itens que já têm as duas datas preenchidas.

    Cálculo feito em Python (não em SQL) de propósito: diferença de datas tem
    sintaxe diferente entre SQLite e Postgres, e a quantidade de itens aqui é
    pequena o bastante pra isso não pesar."""
    itens = ItemPedido.query.filter(
        ItemPedido.inicio_producao.isnot(None), ItemPedido.termino_inspecao.isnot(None)
    ).all()
    if not itens:
        return None
    dias = [(i.termino_inspecao - i.inicio_producao).days for i in itens]
    return round(sum(dias) / len(dias), 1)


def _otd_percentual():
    """OTD (On-Time Delivery): % dos itens finalizados cujo término de inspeção
    aconteceu até a data de liberação prevista. Primeira proposta de cálculo —
    ajustável se o critério de "no prazo" precisar ser outro."""
    itens = ItemPedido.query.filter(
        ItemPedido.status_producao == "FINALIZADO",
        ItemPedido.liberacao_prevista.isnot(None),
        ItemPedido.termino_inspecao.isnot(None),
    ).all()
    if not itens:
        return None
    no_prazo = sum(1 for i in itens if i.termino_inspecao <= i.liberacao_prevista)
    return round(100 * no_prazo / len(itens), 1)


def _backlog_por_estacao():
    """Quantidade de itens não finalizados por estação, com uma amostra dos
    próprios pedidos parados ali — usado no gráfico "Backlog por estação" do
    Painel (única tela que chama esta função). Pedido do Bruno (03/09/2026):

      1) não contar "OUTROS"/"PROJETO ESPECIAL" nesse gráfico específico —
         são "estações" genéricas demais aqui; a tela /estacoes continua
         mostrando as duas normalmente (estacoes_lista() tem sua própria
         consulta, independente desta função, então não é afetada);
      2) ao passar o mouse, já ver quais pedidos estão parados naquela
         estação, com link direto pra abrir o Kanban dela — por isso cada
         linha já sai com uma amostra de itens (pedido/cliente/produto) e a
         própria URL da estação, prontas pro tooltip/clique do gráfico (ver
         painel.html)."""
    ESTACOES_FORA_DO_GRAFICO_PAINEL = {"OUTROS", "PROJETO ESPECIAL"}
    itens = (
        ItemPedido.query.options(selectinload(ItemPedido.pedido))
        .filter(
            ItemPedido.status_producao != "FINALIZADO",
            ItemPedido.estacao.isnot(None),
            ~ItemPedido.estacao.in_(ESTACOES_FORA_DO_GRAFICO_PAINEL),
        )
        .all()
    )

    agrupado = {}
    for item in itens:
        agrupado.setdefault(item.estacao, []).append(item)

    resultado = []
    for estacao, lista in agrupado.items():
        lista.sort(key=lambda i: i.pedido.data_inclusao_pedido if (i.pedido and i.pedido.data_inclusao_pedido) else date.min)
        amostra = [
            {
                "pedido_venda": i.pedido.pedido_venda if i.pedido else None,
                "cliente": i.pedido.cliente if i.pedido else "—",
                "produto": i.descricao_produto,
            }
            for i in lista[:8]
        ]
        resultado.append({
            "estacao": estacao,
            "quantidade": len(lista),
            "itens_amostra": amostra,
            "restante": max(0, len(lista) - len(amostra)),
            "url": url_for("estacao_kanban", nome=estacao),
        })
    resultado.sort(key=lambda r: -r["quantidade"])
    return resultado


def _tempo_relativo(quando):
    """"agora mesmo" / "há Xmin" / "há Xh" / "há X dias" a partir de um
    datetime (UTC, mesmo padrão de todo `criado_em` no app) — usado só no
    feed de apontamentos de Qualidade do Painel, pra dar uma ideia rápida de
    "quando" sem cravar hora exata. Itens com mais de 30 dias caem pra data
    cheia (dd/mm/aaaa), já que "há 47 dias" deixa de ser uma leitura útil."""
    if not quando:
        return "—"
    segundos = (datetime.utcnow() - quando).total_seconds()
    if segundos < 60:
        return "agora mesmo"
    if segundos < 3600:
        return f"há {int(segundos // 60)}min"
    if segundos < 86400:
        return f"há {int(segundos // 3600)}h"
    dias = int(segundos // 86400)
    if dias == 1:
        return "há 1 dia"
    if dias < 30:
        return f"há {dias} dias"
    return quando.strftime("%d/%m/%Y")


def _apontamentos_recentes_qualidade(desde=None, limite=15):
    """Últimos apontamentos de Qualidade (RDIM + RNC), mais recentes primeiro
    — pedido do Bruno (03/09/2026): "todo apontamento diário da qualidade...
    tenha um campo de aviso ou notificação dentro do painel", pra ele, como
    gestor, ter ciência breve de todo apontamento sem precisar entrar na
    área de Qualidade todo dia — cada linha já sai com link direto pro
    registro (RDIM -> rdim_editar, RNC -> qualidade_editar).

    `desde` (opcional): datetime da última vez que o Painel foi aberto (ver
    sessão 'ultima_visita_painel' na rota `painel()`) — marca quais entram
    como "novo desde a última visita", sem precisar de tabela/coluna nova de
    controle de leitura (guardado na sessão do próprio navegador)."""
    rdims = (
        InspecaoFinal.query
        .options(selectinload(InspecaoFinal.item).selectinload(ItemPedido.pedido))
        .order_by(InspecaoFinal.criado_em.desc())
        .limit(limite)
        .all()
    )
    rncs = RncQualidade.query.order_by(RncQualidade.criado_em.desc()).limit(limite).all()

    eventos = []
    for i in rdims:
        eventos.append({
            "tipo": "RDIM",
            "criado_em": i.criado_em,
            "titulo": f"{i.pedido_venda or 'Pedido —'} · {i.cliente or 'Sem cliente'}",
            "detalhe": RDIM_RESULTADO_LABELS.get(i.resultado, i.resultado or "Sem resultado"),
            "cor": RDIM_RESULTADO_CORES.get(i.resultado, "secondary"),
            "link": url_for("rdim_editar", inspecao_id=i.id),
        })
    for r in rncs:
        eventos.append({
            "tipo": "RNC",
            "criado_em": r.criado_em,
            "titulo": f"{r.numero_rnc or ('RNC #' + str(r.id))} · {r.cliente_projeto or 'Sem cliente/projeto'}",
            "detalhe": r.tipo_nc or r.status_geral or "Sem tipo",
            "cor": RNC_SEVERIDADE_CORES.get(r.severidade, "secondary"),
            "link": url_for("qualidade_editar", rnc_id=r.id),
        })

    eventos.sort(key=lambda e: e["criado_em"] or datetime.min, reverse=True)
    eventos = eventos[:limite]
    for e in eventos:
        e["ha_quanto_tempo"] = _tempo_relativo(e["criado_em"])
        e["novo"] = bool(desde and e["criado_em"] and e["criado_em"] > desde)
    return eventos


def _lead_time_por_estacao(desde=None):
    """Lead time médio (dias) agrupado por estação, entre os itens que já têm
    início de produção e término de inspeção preenchidos."""
    query = ItemPedido.query.filter(
        ItemPedido.inicio_producao.isnot(None), ItemPedido.termino_inspecao.isnot(None), ItemPedido.estacao.isnot(None)
    )
    if desde:
        query = query.filter(ItemPedido.termino_inspecao >= desde)
    agrupado = {}
    for item in query.all():
        agrupado.setdefault(item.estacao, []).append((item.termino_inspecao - item.inicio_producao).days)

    resultado = [
        {"estacao": estacao, "lt_medio": round(sum(dias) / len(dias), 1), "quantidade": len(dias)}
        for estacao, dias in agrupado.items()
    ]
    resultado.sort(key=lambda r: r["lt_medio"], reverse=True)
    return resultado


def _otd_por_vendedor(desde=None):
    """OTD (% no prazo) agrupado por vendedor, entre os itens finalizados que
    têm liberação prevista e término de inspeção preenchidos."""
    query = ItemPedido.query.options(selectinload(ItemPedido.pedido)).filter(
        ItemPedido.status_producao == "FINALIZADO",
        ItemPedido.liberacao_prevista.isnot(None),
        ItemPedido.termino_inspecao.isnot(None),
    )
    if desde:
        query = query.filter(ItemPedido.termino_inspecao >= desde)

    agrupado = {}
    for item in query.all():
        vendedor = (item.pedido.vendedor if item.pedido and item.pedido.vendedor else "Sem vendedor")
        grupo = agrupado.setdefault(vendedor, {"no_prazo": 0, "total": 0})
        grupo["total"] += 1
        if item.termino_inspecao <= item.liberacao_prevista:
            grupo["no_prazo"] += 1

    resultado = [
        {"vendedor": vendedor, "otd": round(100 * g["no_prazo"] / g["total"], 1), "total": g["total"]}
        for vendedor, g in agrupado.items()
    ]
    resultado.sort(key=lambda r: r["otd"])
    return resultado


def _tendencia_kpis(meses=6):
    """Finalizados / lead time médio / OTD por mês, dos últimos N meses — para
    os gráficos de tendência da tela de KPIs."""
    hoje = date.today()
    pontos = []
    ano, mes = hoje.year, hoje.month
    for _ in range(meses):
        pontos.append((ano, mes))
        mes -= 1
        if mes == 0:
            mes, ano = 12, ano - 1
    pontos.reverse()

    resultado = []
    for ano, mes in pontos:
        inicio = date(ano, mes, 1)
        fim = date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1)
        itens = ItemPedido.query.filter(
            ItemPedido.termino_inspecao.isnot(None),
            ItemPedido.termino_inspecao >= inicio,
            ItemPedido.termino_inspecao < fim,
        ).all()

        lt_medio, otd = None, None
        if itens:
            lts = [(i.termino_inspecao - i.inicio_producao).days for i in itens if i.inicio_producao]
            if lts:
                lt_medio = round(sum(lts) / len(lts), 1)
            com_prazo = [i for i in itens if i.liberacao_prevista]
            if com_prazo:
                no_prazo = sum(1 for i in com_prazo if i.termino_inspecao <= i.liberacao_prevista)
                otd = round(100 * no_prazo / len(com_prazo), 1)

        resultado.append({"mes": f"{MESES_PT[mes - 1]}/{ano}", "finalizados": len(itens), "lt_medio": lt_medio, "otd": otd})
    return resultado


def _gargalos_por_estacao():
    """Ranking de estações por "quanto está travado ali": fila, atraso, tempo
    de espera médio, lead time médio e valor parado (não finalizado).

    Consultado estação por estação (poucas estações, poucas linhas cada) em vez
    de carregar TODOS os pedidos em Python — o mesmo cuidado de performance já
    aplicado na tela inicial (fase 3) e na visão geral de Estações (fase 5)."""
    hoje = date.today()
    estacoes = Estacao.query.filter_by(ativo=True).order_by(Estacao.ordem_exibicao).all()
    resultado = []

    for e in estacoes:
        nome = e.nome
        abertos = ItemPedido.query.filter(ItemPedido.estacao == nome, ItemPedido.status_producao != "FINALIZADO")
        fila = abertos.count()
        atraso = abertos.filter(
            ItemPedido.liberacao_prevista.isnot(None), ItemPedido.liberacao_prevista < hoje
        ).count()
        valor_parado = (
            db.session.query(func.sum(ItemPedido.quantidade * ItemPedido.custo_unitario))
            .filter(ItemPedido.estacao == nome, ItemPedido.status_producao != "FINALIZADO")
            .scalar()
            or 0.0
        )

        itens_com_inicio = (
            ItemPedido.query.options(selectinload(ItemPedido.pedido))
            .filter(ItemPedido.estacao == nome, ItemPedido.inicio_producao.isnot(None))
            .all()
        )
        esperas = [
            (i.inicio_producao - i.pedido.data_inclusao_pedido).days
            for i in itens_com_inicio
            if i.pedido and i.pedido.data_inclusao_pedido
        ]
        tempo_espera_medio = round(sum(esperas) / len(esperas), 1) if esperas else None

        lts = [(i.termino_inspecao - i.inicio_producao).days for i in itens_com_inicio if i.termino_inspecao]
        lt_medio = round(sum(lts) / len(lts), 1) if lts else None

        resultado.append(
            {
                "estacao": nome,
                "fila": fila,
                "atraso": atraso,
                "tempo_espera_medio": tempo_espera_medio,
                "lt_medio": lt_medio,
                "valor_parado": round(valor_parado, 2),
            }
        )

    resultado.sort(key=lambda r: (r["fila"], r["atraso"]), reverse=True)
    return resultado


def _faturamento_detalhado(ano, mes, cliente=None, regiao=None, vendedor=None):
    """Previsto × realizado de um mês específico, com o mesmo valor já
    quebrado por cliente / região / vendedor — usado na tela de Faturamento."""
    inicio = date(ano, mes, 1)
    fim = date(ano + 1, 1, 1) if mes == 12 else date(ano, mes + 1, 1)

    base = ItemPedido.query.options(selectinload(ItemPedido.pedido)).join(Pedido, ItemPedido.pedido_id == Pedido.id)
    if cliente:
        base = base.filter(Pedido.cliente.ilike(f"%{cliente}%"))
    if vendedor:
        base = base.filter(Pedido.vendedor.ilike(f"%{vendedor}%"))
    if regiao:
        ufs_da_regiao = [uf for uf, r in REGIAO_POR_UF.items() if r == regiao]
        if ufs_da_regiao:
            base = base.filter(Pedido.estado.in_(ufs_da_regiao))

    itens_previstos = base.filter(ItemPedido.liberacao_prevista >= inicio, ItemPedido.liberacao_prevista < fim).all()
    itens_realizados = base.filter(
        ItemPedido.liberacao_faturamento >= inicio, ItemPedido.liberacao_faturamento < fim
    ).all()

    def _agrupar(itens, chave_fn):
        agrupado = {}
        for i in itens:
            chave = chave_fn(i) or "—"
            agrupado[chave] = agrupado.get(chave, 0.0) + i.valor_faturamento_realizado
        linhas = [{"chave": k, "valor": round(v, 2)} for k, v in agrupado.items()]
        linhas.sort(key=lambda l: l["valor"], reverse=True)
        return linhas

    return {
        "previsto_total": round(sum(i.valor_total for i in itens_previstos), 2),
        "realizado_total": round(sum(i.valor_faturamento_realizado for i in itens_realizados), 2),
        "itens_realizados": len(itens_realizados),
        "por_cliente": _agrupar(itens_realizados, lambda i: i.pedido.cliente if i.pedido else None),
        "por_regiao": _agrupar(itens_realizados, lambda i: REGIAO_POR_UF.get(i.pedido.estado) if i.pedido and i.pedido.estado else None),
        "por_vendedor": _agrupar(itens_realizados, lambda i: i.pedido.vendedor if i.pedido else None),
        # lista "crua" dos itens realizados (não usada na tela, só na exportação
        # de relatório — mantida separada da contagem "itens_realizados" acima
        # pra não mudar o que a tela de Faturamento já espera receber)
        "itens_realizados_lista": itens_realizados,
    }


def _faturamento_previsto_nao_realizado():
    """Itens cuja liberação PREVISTA já passou mas a liberação de FATURAMENTO
    ainda não aconteceu — ou seja, o faturamento que era esperado até agora
    ainda não se realizou (mesmo conceito de "previsto" usado na tela de
    Faturamento, só que olhando pro atraso em vez de somar por mês)."""
    hoje = date.today()
    return (
        ItemPedido.query.options(selectinload(ItemPedido.pedido))
        .join(Pedido, ItemPedido.pedido_id == Pedido.id)
        .filter(
            ItemPedido.liberacao_prevista.isnot(None),
            ItemPedido.liberacao_prevista < hoje,
            ItemPedido.liberacao_faturamento.is_(None),
        )
        .order_by(ItemPedido.liberacao_prevista)
        .all()
    )


def _construir_timeline(pedido):
    """Monta uma lista de eventos (data + descrição) a partir das datas já
    preenchidas no pedido e em cada item, para exibir como linha do tempo."""
    eventos = []
    if pedido.data_cliente:
        eventos.append({"data": pedido.data_cliente, "titulo": "Data do cliente", "item": None})
    if pedido.data_inclusao_pedido:
        eventos.append({"data": pedido.data_inclusao_pedido, "titulo": "Pedido incluído no sistema", "item": None})

    for item in pedido.itens:
        rotulo = item.descricao_produto
        if item.inicio_producao:
            eventos.append({"data": item.inicio_producao, "titulo": f"Início de produção — {rotulo}", "item": item})
        if item.inicio_inspecao:
            eventos.append({"data": item.inicio_inspecao, "titulo": f"Início de inspeção/embalagem — {rotulo}", "item": item})
        if item.termino_inspecao:
            eventos.append({"data": item.termino_inspecao, "titulo": f"Término de inspeção/embalagem — {rotulo}", "item": item})
        if item.liberacao_faturamento:
            eventos.append({"data": item.liberacao_faturamento, "titulo": f"Liberado para faturamento — {rotulo}", "item": item})

    eventos.sort(key=lambda e: e["data"])
    return eventos


def _filtrar_pedidos(args):
    """Aplica exatamente os mesmos filtros da tela de Listagem (rota "/") a
    partir de um dict tipo request.args, devolvendo a query já filtrada (sem
    paginação) e o dicionário de filtros usado.

    Extraído da rota "dashboard" pra ser compartilhado com a exportação de
    relatório (fase 12) — assim a exportação nunca corre o risco de aplicar
    uma regra de filtro diferente da que a tela usa."""
    query = Pedido.query.options(selectinload(Pedido.itens))

    cliente = args.get("cliente", "").strip()
    status = args.get("status", "").strip()
    estacao = args.get("estacao", "").strip()
    vendedor = args.get("vendedor", "").strip()
    busca = args.get("busca", "").strip()
    produto = args.get("produto", "").strip()
    regiao = args.get("regiao", "").strip()
    data_inicio = args.get("data_inicio", "").strip()
    data_fim = args.get("data_fim", "").strip()
    atrasados = args.get("atrasados", "").strip()
    planejamento_semanal = args.get("planejamento_semanal", "").strip()
    planejamento_mensal = args.get("planejamento_mensal", "").strip()
    mes_inclusao = args.get("mes_inclusao", "").strip()
    mes_entrega_cliente = args.get("mes_entrega_cliente", "").strip()

    if cliente:
        query = query.filter(Pedido.cliente.ilike(f"%{cliente}%"))
    if estacao:
        query = query.filter(Pedido.itens.any(ItemPedido.estacao == estacao))
    if vendedor:
        query = query.filter(Pedido.vendedor.ilike(f"%{vendedor}%"))
    if busca:
        like = f"%{busca}%"
        query = query.filter(
            or_(
                Pedido.pedido_venda.ilike(like),
                Pedido.cliente.ilike(like),
                Pedido.itens.any(ItemPedido.descricao_produto.ilike(like)),
            )
        )
    if produto:
        query = query.filter(Pedido.itens.any(ItemPedido.descricao_produto.ilike(f"%{produto}%")))
    if regiao:
        ufs_da_regiao = [uf for uf, r in REGIAO_POR_UF.items() if r == regiao]
        if ufs_da_regiao:
            query = query.filter(Pedido.estado.in_(ufs_da_regiao))
    if data_inicio:
        data_inicio_parsed = _parse_data_form(data_inicio)
        if data_inicio_parsed:
            query = query.filter(Pedido.data_inclusao_pedido >= data_inicio_parsed)
    if data_fim:
        data_fim_parsed = _parse_data_form(data_fim)
        if data_fim_parsed:
            query = query.filter(Pedido.data_inclusao_pedido <= data_fim_parsed)
    if atrasados:
        query = query.filter(_predicado_atrasado())
    if status:
        predicado = _predicado_status(status)
        if predicado is not None:
            query = query.filter(predicado)
    if planejamento_semanal:
        query = query.filter(Pedido.itens.any(ItemPedido.planejamento_semanal == planejamento_semanal))
    if planejamento_mensal:
        mes_ano = _parse_mes_ano_form(planejamento_mensal, None)
        if mes_ano:
            semanas_do_mes = [
                s for (s,) in db.session.query(ItemPedido.planejamento_semanal)
                .filter(ItemPedido.planejamento_semanal.isnot(None))
                .distinct()
                if _mes_ano_da_semana_pcp(s) == mes_ano
            ]
            if semanas_do_mes:
                query = query.filter(Pedido.itens.any(ItemPedido.planejamento_semanal.in_(semanas_do_mes)))
            else:
                # Mês escolhido não tem nenhum planejamento semanal preenchido
                # ainda — não deve mostrar nada (em vez de ignorar o filtro).
                query = query.filter(false())
    if mes_inclusao:
        mes_ano = _parse_mes_ano_form(mes_inclusao, None)
        if mes_ano:
            ano, mes = mes_ano
            query = query.filter(
                extract("year", Pedido.data_inclusao_pedido) == ano,
                extract("month", Pedido.data_inclusao_pedido) == mes,
            )
        else:
            query = query.filter(false())
    if mes_entrega_cliente:
        mes_ano = _parse_mes_ano_form(mes_entrega_cliente, None)
        if mes_ano:
            ano, mes = mes_ano
            query = query.filter(
                extract("year", Pedido.data_cliente) == ano,
                extract("month", Pedido.data_cliente) == mes,
            )
        else:
            query = query.filter(false())

    query = query.order_by(Pedido.data_inclusao_pedido.desc().nullslast(), Pedido.id.desc())

    filtros = dict(
        cliente=cliente,
        status=status,
        estacao=estacao,
        vendedor=vendedor,
        busca=busca,
        produto=produto,
        regiao=regiao,
        data_inicio=data_inicio,
        data_fim=data_fim,
        atrasados=atrasados,
        planejamento_semanal=planejamento_semanal,
        planejamento_mensal=planejamento_mensal,
        mes_inclusao=mes_inclusao,
        mes_entrega_cliente=mes_entrega_cliente,
    )
    return query, filtros


class _LinhaListagemGeral:
    """Uma linha da Listagem Geral = 1 pedido + 1 item (produto) dele — o
    mesmo número de pedido pode aparecer em várias linhas, uma por produto
    distinto, igual à planilha de referência. Só une os dois objetos num
    lugar só pra o template não precisar fazer `linha.pedido.x` /
    `linha.item.y` o tempo todo."""

    def __init__(self, pedido, item):
        self.pedido = pedido
        self.item = item

    # ---- identidade do pedido ----
    @property
    def pedido_id(self):
        return self.pedido.id

    @property
    def pedido_venda(self):
        return self.pedido.pedido_venda

    @property
    def cliente(self):
        return self.pedido.cliente

    @property
    def vendedor(self):
        return self.pedido.vendedor

    @property
    def data_inclusao_pedido(self):
        return self.pedido.data_inclusao_pedido

    @property
    def data_cliente(self):
        """Data solicitada pelo cliente (prazo comercial) — já existia no
        pedido (campo "Data do cliente" na tela de editar), só não aparecia
        na Listagem Geral."""
        return self.pedido.data_cliente

    @property
    def prioridade(self):
        return self.pedido.prioridade

    @property
    def frete(self):
        return self.pedido.frete

    @property
    def pais(self):
        return self.pedido.pais

    @property
    def estado(self):
        return self.pedido.estado

    @property
    def cidade(self):
        return self.pedido.cidade

    # ---- dados do item (produto) ----
    @property
    def item_id(self):
        return self.item.id

    @property
    def descricao_produto(self):
        return self.item.descricao_produto

    @property
    def quantidade(self):
        return self.item.quantidade

    @property
    def venda_unidade(self):
        """Mesmo dado de custo que já existe (custo_unitario) — só exibido
        sob o rótulo "Venda" pedido pelo Bruno."""
        return self.item.custo_unitario

    @property
    def venda_total(self):
        return self.item.valor_total

    @property
    def estacao(self):
        return self.item.estacao

    @property
    def status_producao(self):
        return self.item.status_producao

    @property
    def liberacao_prevista(self):
        return self.item.liberacao_prevista

    @property
    def liberacao_real(self):
        return self.item.liberacao_real

    @property
    def planejamento_semanal(self):
        return self.item.planejamento_semanal

    @property
    def venda_total_pedido(self):
        """Soma o custo de TODOS os itens do pedido (não só deste produto) —
        reaproveita Pedido.valor_total, que já faz exatamente essa soma."""
        return self.pedido.valor_total

    @property
    def semaforo(self):
        return self.item.semaforo


def _linhas_listagem_geral(pedidos, args):
    """Achata a lista de Pedido (com itens já carregados) em 1 linha por
    ItemPedido — a granularidade que a Listagem Geral usa agora (1 linha por
    produto, igual ao print de referência).

    `_filtrar_pedidos` já decidiu quais PEDIDOS entram (algum item bate o
    filtro); aqui, pros filtros que são naturalmente por ITEM — estação,
    produto, planejamento semanal/mensal —, mostra só os produtos que batem,
    não o pedido inteiro. Sem isso, filtrar por "semana X" mostraria também
    os outros produtos do mesmo pedido que caem em semanas diferentes."""
    estacao = args.get("estacao", "").strip()
    produto = args.get("produto", "").strip().upper()
    planejamento_semanal = args.get("planejamento_semanal", "").strip()
    planejamento_mensal = args.get("planejamento_mensal", "").strip()
    mes_ano = _parse_mes_ano_form(planejamento_mensal, None) if planejamento_mensal else None

    linhas = []
    for pedido in pedidos:
        for item in pedido.itens:
            if estacao and item.estacao != estacao:
                continue
            if produto and produto not in (item.descricao_produto or "").upper():
                continue
            if planejamento_semanal and item.planejamento_semanal != planejamento_semanal:
                continue
            if mes_ano and _mes_ano_da_semana_pcp(item.planejamento_semanal) != mes_ano:
                continue
            linhas.append(_LinhaListagemGeral(pedido, item))
    return linhas


def _ordenar_com_nulos_no_fim(linhas, chave, reverse):
    """Ordena por `chave(linha)`, deixando quem não tem valor (None) sempre
    no fim, não importa a direção — comportamento mais previsível pro
    usuário do que deixar o Python inverter os vazios junto com o resto."""
    com_valor = [l for l in linhas if chave(l) is not None]
    sem_valor = [l for l in linhas if chave(l) is None]
    com_valor.sort(key=chave, reverse=reverse)
    return com_valor + sem_valor


SORT_KEYS_LISTAGEM_GERAL = {
    "pedido_venda": lambda l: (l.pedido_venda or "").upper() or None,
    "cliente": lambda l: (l.cliente or "").upper() or None,
    "vendedor": lambda l: (l.vendedor or "").upper() or None,
    "data_inclusao": lambda l: l.data_inclusao_pedido,
    "data_cliente": lambda l: l.data_cliente,
    "prioridade": lambda l: PRIORIDADE_OPCOES.index(l.prioridade) if l.prioridade in PRIORIDADE_OPCOES else None,
    "produto": lambda l: (l.descricao_produto or "").upper() or None,
    "quantidade": lambda l: l.quantidade,
    "venda_unidade": lambda l: l.venda_unidade,
    "venda_total": lambda l: l.venda_total,
    "venda_total_pedido": lambda l: l.venda_total_pedido,
    "estacao": lambda l: (l.estacao or "").upper() or None,
    "status": lambda l: STATUS_OPCOES.index(l.status_producao) if l.status_producao in STATUS_OPCOES else None,
    "liberacao_prevista": lambda l: l.liberacao_prevista,
    "liberacao_real": lambda l: l.liberacao_real,
    "planejamento_semanal": lambda l: _chave_semana_pcp(l.planejamento_semanal),
    "frete": lambda l: (l.frete or "").upper() or None,
    "pais": lambda l: (l.pais or "").upper() or None,
    "estado": lambda l: (l.estado or "").upper() or None,
    "cidade": lambda l: (l.cidade or "").upper() or None,
}

SORT_PADRAO = "data_inclusao"
DIR_PADRAO = "desc"


def _filtrar_pedidos_operacao(args):
    """Filtros das 4 sub-abas de Gestão Operação (Comercial/PCP/Logística/
    Resultados). Independente de _filtrar_pedidos (Gestão Produção) — opera só
    em PedidoOperacao, sem nenhum join com Pedido/ItemPedido/estação.

    `segmento` + `periodo` (pedido do Bruno, 03/09/2026 — tela Resultados/
    OTD: "quero ver todos os pedidos de julho faturados ou dentro do
    planejamento semanal do pcp, com isso ver o otd"; ampliado no mesmo dia
    pra aceitar não só mês, mas trimestre/semestre/ano/"todos" — ver
    _parse_periodo) — reaproveita a MESMA definição de "período" já usada em
    _faturamento_por_periodo: agrupa pelo Término Semanal PCP
    (go_termino_semanal_pcp), não por nenhuma data de calendário.
      - "planejamento": todo pedido cujo Término Semanal PCP cai no período
        escolhido (equivalente ao "Qtd/Valor liberado" da tabela semanal).
      - "faturados": o mesmo conjunto acima, restrito a quem já tem Valor NF
        Emitida preenchido (equivalente ao "Qtd/Valor faturado").
    Sem `segmento`, nenhum filtro de período é aplicado — comportamento
    antigo, todos os pedidos."""
    query = PedidoOperacao.query

    cliente = args.get("cliente", "").strip()
    vendedor = args.get("vendedor", "").strip()
    busca = args.get("busca", "").strip()
    segmento = args.get("segmento", "").strip()
    periodo_str = args.get("periodo", "").strip()

    if cliente:
        query = query.filter(PedidoOperacao.cliente.ilike(f"%{cliente}%"))
    if vendedor:
        query = query.filter(PedidoOperacao.vendedor.ilike(f"%{vendedor}%"))
    if busca:
        like = f"%{busca}%"
        query = query.filter(
            or_(
                PedidoOperacao.pedido_venda.ilike(like),
                PedidoOperacao.cliente.ilike(like),
            )
        )

    if segmento in ("planejamento", "faturados"):
        tipo_p, ano_p, valor_p, _ = _parse_periodo(periodo_str)
        query = _filtrar_por_periodo_pcp(query, tipo_p, ano_p, valor_p)
        if segmento == "faturados":
            query = query.filter(PedidoOperacao.go_valor_nf_emitida.isnot(None))
    else:
        segmento = ""

    query = query.order_by(PedidoOperacao.data_inclusao_pedido.desc().nullslast(), PedidoOperacao.id.desc())

    filtros = dict(cliente=cliente, vendedor=vendedor, busca=busca, segmento=segmento)
    return query, filtros


def _linhas_gestao_operacao(args):
    """Usado pelas 4 sub-abas de Gestão Operação: pagina o resultado de
    _filtrar_pedidos_operacao. Cada linha já é 1 pedido comercial (tabela
    própria PedidoOperacao) — sem duplicidade legada, sem precisar agrupar
    nada em Python. Devolve também a `query` (filtrada, sem paginação) —
    usada pela tela Resultados/OTD pra calcular o resumo de OTD só sobre o
    mesmo conjunto filtrado (ver _resumo_otd)."""
    page = args.get("page", 1, type=int)
    query, filtros = _filtrar_pedidos_operacao(args)
    total_filtrado = query.count()
    total_paginas = max(1, (total_filtrado + PAGE_SIZE - 1) // PAGE_SIZE)
    pagina = query.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    return pagina, page, total_paginas, total_filtrado, filtros, query


def _itens_producao_por_pedido_venda(pedidos_venda):
    """Pra tela "Listagem Geral" de Gestão Operação (pedido do Bruno: passar o
    mouse num pedido mostra os itens/quantidades já preenchidos no PCP, em
    Gestão Produção). PedidoOperacao e Pedido/ItemPedido são tabelas
    INDEPENDENTES (sem FK) — o único jeito de ligar um ao outro é o texto do
    "nº pedido de venda" que aparece nos dois. Casa só por igualdade exata
    (já normalizado/tirado espaço), NUNCA por aproximação — é só pra exibir
    numa dica visual, e uma correspondência errada mostraria os itens do
    pedido errado. Uma única query com IN (nunca N+1) — recebe a lista de
    pedido_venda já normalizada da página atual."""
    valores = sorted({v.strip() for v in pedidos_venda if v and v.strip()})
    if not valores:
        return {}
    pedidos = (
        Pedido.query.options(selectinload(Pedido.itens))
        # func.trim() nos dois lados: o texto salvo em Pedido.pedido_venda às
        # vezes tem espaço a mais (import antigo de planilha) — sem isso, a
        # igualdade exata falharia por causa só do espaço, escondendo itens
        # que na prática são do mesmo pedido.
        .filter(func.trim(Pedido.pedido_venda).in_(valores))
        .all()
    )
    mapa = {}
    for pedido in pedidos:
        chave = (pedido.pedido_venda or "").strip()
        if not chave:
            continue
        mapa.setdefault(chave, []).extend(pedido.itens)
    return mapa


def _status_producao_por_pedido_venda(pedidos_venda):
    """Pedido do Bruno (01/09/2026): a coluna "Status produção" das telas de
    Gestão Operação (PCP e a tela de editar) tem que refletir o andamento
    REAL na fábrica — mesmo casamento por pedido_venda (trim, sem FK, nunca
    aproximado) já usado pra mostrar os itens na Listagem Geral — em vez do
    status calculado só a partir dos próprios dados da Operação (que
    continua existindo como fallback em PedidoOperacao.status_producao, pra
    quando o pedido ainda não foi lançado em Gestão Produção). Reaproveita a
    mesma regra de agregação de Pedido.status_producao (EM TRATATIVA vence
    tudo; só FINALIZADO se todos os itens estiverem; só PENDENTE se todos
    estiverem; senão ANDAMENTO)."""
    itens_por_pedido = _itens_producao_por_pedido_venda(pedidos_venda)
    status_por_pedido = {}
    for chave, itens in itens_por_pedido.items():
        if not itens:
            continue
        status_itens = {item.status_producao for item in itens}
        if "EM TRATATIVA" in status_itens:
            status_por_pedido[chave] = "EM TRATATIVA"
        elif status_itens == {"FINALIZADO"}:
            status_por_pedido[chave] = "FINALIZADO"
        elif status_itens == {"PENDENTE"}:
            status_por_pedido[chave] = "PENDENTE"
        else:
            status_por_pedido[chave] = "ANDAMENTO"
    return status_por_pedido


def _liberacao_pcp_por_pedido_venda(pedidos_venda):
    """Pedido do Bruno (01/09/2026, ampliado 03/09/2026): as colunas
    "Previsão liberação PCP", "Data efetiva liberação" e "Término semanal"
    da tela PCP de Gestão Operação têm que acompanhar automaticamente os
    dados equivalentes já preenchidos em Gestão Produção ("Liberação
    prevista"/"Liberação real"/"Planejamento semanal" de cada item) — mesmo
    casamento por pedido_venda (trim, sem FK, nunca aproximado) já usado pra
    Status produção/itens da Listagem Geral. Enquanto o pedido não tiver
    sido lançado em Gestão Produção (sem match), os campos próprios de
    PedidoOperacao (go_previsao_liberacao_pcp/go_data_efetiva_liberacao_pcp/
    go_termino_semanal_pcp) continuam valendo como fallback editável
    manualmente — mesmo espírito de status_producao/
    _status_producao_por_pedido_venda.

    Previsão/Efetiva são preenchidas em Produção por um único bloco no topo
    do formulário de edição que aplica o MESMO valor a todos os itens do
    pedido de uma vez (decisão de 31/08/2026) — então, quando por algum
    motivo os itens de um mesmo pedido têm valores diferentes entre si (dado
    legado, de antes dessa decisão existir), usamos a data mais recente
    entre eles. "Término semanal" é texto (ex. "SEMANA 03 / AGO / 2026"), não
    dado — usamos _chave_semana_pcp (mesma função que já ordena esse rótulo
    em outros lugares do sistema) pra achar o rótulo cronologicamente mais
    recente entre os itens."""
    itens_por_pedido = _itens_producao_por_pedido_venda(pedidos_venda)
    liberacao_por_pedido = {}
    for chave, itens in itens_por_pedido.items():
        previstas = [i.liberacao_prevista for i in itens if i.liberacao_prevista]
        reais = [i.liberacao_real for i in itens if i.liberacao_real]
        semanas = [i.planejamento_semanal for i in itens if i.planejamento_semanal]
        if not previstas and not reais and not semanas:
            continue
        liberacao_por_pedido[chave] = {
            "previsao": max(previstas) if previstas else None,
            "efetiva": max(reais) if reais else None,
            "termino_semanal": max(semanas, key=_chave_semana_pcp) if semanas else None,
        }
    return liberacao_por_pedido


def _data_cliente_por_pedido_venda(pedidos_venda):
    """Pedido.data_cliente por pedido_venda (trim, sem FK) — pedido do Bruno
    (03/09/2026): "quero que todos os dados dentro da gestão operação seja
    extraída automaticamente da gestão produção". "Solicitada cliente/
    retira" (PCP) e "Data solicitada entrega" (Comercial) de Gestão Operação
    são o MESMO dado que "Data do cliente" de Gestão Produção (já
    documentado em _criar_pedido_operacao_a_partir_de_producao) — até aqui
    elas só copiavam esse valor UMA VEZ na inclusão do pedido em Operação;
    agora acompanham ao vivo, mesmo espírito de _liberacao_pcp_por_pedido_
    venda. Os campos próprios de PedidoOperacao continuam valendo como
    fallback editável manualmente enquanto o pedido não tiver sido lançado
    em Gestão Produção."""
    valores = sorted({v.strip() for v in pedidos_venda if v and v.strip()})
    if not valores:
        return {}
    pedidos = (
        Pedido.query
        .filter(func.trim(Pedido.pedido_venda).in_(valores), Pedido.data_cliente.isnot(None))
        .all()
    )
    mapa = {}
    for pedido in pedidos:
        chave = (pedido.pedido_venda or "").strip()
        if chave:
            mapa[chave] = pedido.data_cliente
    return mapa


def _buscar_pedidos_para_status(termo, limite=12):
    """Pedido do Bruno (01/09/2026): canal único de busca no topo do Painel —
    "sou o PCP, comercial me cobrou de um pedido" — digita o nº do pedido de
    venda OU o cliente e recebe sugestões pra abrir o status completo
    (produção + operação) daquele pedido específico.

    Casa em Pedido (Gestão Produção) e em PedidoOperacao (Gestão Operação)
    separadamente — cada tabela pode ter pedidos que a outra não tem — e
    devolve no máximo 1 sugestão por pedido_venda (nunca duplicada), usando
    o próprio pedido_venda como identificador da busca de detalhe. Por isso
    só considera pedidos com pedido_venda preenchido: sem esse número não dá
    pra cruzar as duas tabelas de forma confiável — mesma regra de "match
    exato, nunca aproximado" já usada em toda outra cross-referência do
    sistema (ver _itens_producao_por_pedido_venda)."""
    termo = (termo or "").strip()
    if not termo:
        return []
    padrao = f"%{termo}%"

    candidatos = {}  # pedido_venda (trim) -> dict de exibição

    pedidos = (
        Pedido.query.filter(
            Pedido.pedido_venda.isnot(None),
            func.trim(Pedido.pedido_venda) != "",
            or_(Pedido.pedido_venda.ilike(padrao), Pedido.cliente.ilike(padrao)),
        )
        .order_by(Pedido.data_inclusao_pedido.desc().nullslast())
        .limit(limite * 3)
        .all()
    )
    for p in pedidos:
        chave = (p.pedido_venda or "").strip()
        if not chave or chave in candidatos:
            continue
        candidatos[chave] = {"pedido_venda": chave, "cliente": p.cliente, "tem_producao": True}

    if len(candidatos) < limite:
        pedidos_operacao = (
            PedidoOperacao.query.filter(
                PedidoOperacao.pedido_venda.isnot(None),
                func.trim(PedidoOperacao.pedido_venda) != "",
                or_(PedidoOperacao.pedido_venda.ilike(padrao), PedidoOperacao.cliente.ilike(padrao)),
            )
            .order_by(PedidoOperacao.data_inclusao_pedido.desc().nullslast())
            .limit(limite * 3)
            .all()
        )
        for go in pedidos_operacao:
            chave = (go.pedido_venda or "").strip()
            if not chave or chave in candidatos:
                continue
            candidatos[chave] = {"pedido_venda": chave, "cliente": go.cliente, "tem_producao": False}

    termo_upper = termo.upper()

    def _relevancia(c):
        pv = c["pedido_venda"].upper()
        cli = (c["cliente"] or "").upper()
        if pv == termo_upper:
            return (0, pv)
        if pv.startswith(termo_upper):
            return (1, pv)
        if termo_upper in pv:
            return (2, pv)
        if cli.startswith(termo_upper):
            return (3, cli)
        return (4, cli)

    return sorted(candidatos.values(), key=_relevancia)[:limite]


def _situacao_entrega_go(go, pedido=None):
    """Resumo em UMA frase só do que mais se pergunta pro PCP/Comercial: "cadê
    esse pedido? já foi expedido? já chegou no cliente?" — a informação
    "principal" que o Bruno pediu (pedido de 01/09/2026), consultando a aba
    Expedição/Logística de Gestão Operação. Prioriza a data de entrega mais
    confiável entre as duas que a Operação guarda (Logística x
    Resultados/OTD — historicamente nem sempre as duas são preenchidas
    juntas)."""
    if go is None:
        if pedido is not None and pedido.status_producao == "FINALIZADO":
            return {
                "texto": "Produção finalizada, mas o pedido ainda não foi lançado em Gestão Operação — sem dado de expedição.",
                "cor": "warning",
                "icone": "bi-exclamation-triangle",
            }
        return {
            "texto": "Pedido ainda não lançado em Gestão Operação — sem dados de expedição/logística.",
            "cor": "secondary",
            "icone": "bi-question-circle",
        }

    data_entrega = go.go_data_entregue_cliente or go.go_data_real_entrega
    if data_entrega:
        return {
            "texto": f"Entregue ao cliente em {data_entrega.strftime('%d/%m/%Y')}",
            "cor": "success",
            "icone": "bi-check-circle-fill",
        }
    if go.go_data_pedido_expedido:
        return {
            "texto": f"Expedido em {go.go_data_pedido_expedido.strftime('%d/%m/%Y')} — aguardando confirmação de entrega/coleta",
            "cor": "info",
            "icone": "bi-truck",
        }
    if go.go_data_efetiva_liberacao_pcp:
        return {
            "texto": f"Liberado pelo PCP em {go.go_data_efetiva_liberacao_pcp.strftime('%d/%m/%Y')} — ainda não expedido",
            "cor": "warning",
            "icone": "bi-hourglass-split",
        }
    return {"texto": "Ainda não expedido.", "cor": "secondary", "icone": "bi-hourglass"}


def _otd_do_pedido(go):
    """OTD (On-Time Delivery) de UM pedido específico — pedido do Bruno
    (03/09/2026, tela Consulta Pedido): "incluir o OTD do pedido, se atendeu
    ou não, bem didático e informativo", ao lado da "Situação de entrega".
    Mesma fonte de verdade da tela Resultados/OTD (go.go_otd_realizado,
    preenchido manualmente na aba Resultados/OTD de Gestão Operação) — nunca
    recalculado por conta própria a partir de datas, pra não divergir do
    número que já aparece agregado em Resultados/OTD.

    go_dias_atraso_antecipacao (solicitado x entregue de verdade) entra só
    como detalhe complementar, quando disponível — o "atendeu ou não" em si
    sempre vem do campo manual."""
    if go is None:
        return {
            "texto": "OTD não disponível — pedido ainda não lançado em Gestão Operação.",
            "cor": "secondary",
            "icone": "bi-question-circle",
        }
    if not go.go_otd_realizado:
        return {
            "texto": "OTD ainda não registrado para este pedido.",
            "cor": "secondary",
            "icone": "bi-hourglass",
        }

    atraso = go.go_dias_atraso_antecipacao
    if atraso is None:
        detalhe = ""
    elif atraso > 0:
        detalhe = f" — entregue {atraso} dia(s) após o prazo solicitado."
    elif atraso < 0:
        detalhe = f" — entregue {abs(atraso)} dia(s) antes do prazo solicitado."
    else:
        detalhe = " — entregue exatamente no dia solicitado."

    if go.go_otd_realizado == "SIM":
        return {
            "texto": f"Atendeu o OTD (dentro do prazo solicitado){detalhe}",
            "cor": "success",
            "icone": "bi-check-circle-fill",
        }
    return {
        "texto": f"Não atendeu o OTD (fora do prazo solicitado){detalhe}",
        "cor": "danger",
        "icone": "bi-x-circle-fill",
    }


# "Acompanhamento do pedido" — trilha de 5 etapas (recebido -> produção ->
# inspeção/expedição -> transporte -> entrega) no painel de status de um
# pedido. Pedido do Bruno (01/09/2026), a partir de uma imagem de referência
# que ele anexou (infográfico com o mesmo formato/rótulos), além do resumo
# em uma frase que já existia (_situacao_entrega_go) — este aqui é o
# complemento visual "em que pé exatamente está".
_ETAPAS_ACOMPANHAMENTO_PEDIDO = [
    {
        "label": "Pedido recebido",
        "descricao": "Pedido de venda recebido e registrado no sistema.",
        "icone": "bi-receipt",
        "cor": "#0d6efd",
        "cor_fraca": "rgba(13, 110, 253, .18)",
    },
    {
        "label": "Produção",
        "descricao": "Pedido em produção. Materiais separados e processos em execução.",
        "icone": "bi-gear-wide-connected",
        "cor": "#12b886",
        "cor_fraca": "rgba(18, 184, 134, .18)",
    },
    {
        "label": "Inspeção / Expedição",
        "descricao": "Pedido finalizado. Inspeção realizada e liberado para expedição.",
        "icone": "bi-clipboard2-check",
        "cor": "#f59f00",
        "cor_fraca": "rgba(245, 159, 0, .18)",
    },
    {
        "label": "Em transporte",
        "descricao": "Pedido coletado e em transporte até o destino final.",
        "icone": "bi-truck",
        "cor": "#7048e8",
        "cor_fraca": "rgba(112, 72, 232, .18)",
    },
    {
        "label": "Entrega realizada",
        "descricao": "Pedido entregue ao cliente com sucesso.",
        "icone": "bi-box-seam",
        "cor": "#198754",
        "cor_fraca": "rgba(25, 135, 84, .18)",
    },
]


def _indice_etapa_pedido(pedido, go):
    """Em que das 5 etapas do "Acompanhamento do pedido" ele está agora.
    Mesmos sinais já usados em _situacao_entrega_go, só que granulares em 5
    passos em vez de só "expedido/entregue" — chamada só quando pedido ou go
    existem (nunca os dois None), por isso sempre devolve pelo menos 1
    ("Pedido recebido")."""
    if go is not None and (go.go_data_entregue_cliente or go.go_data_real_entrega):
        return 5
    if go is not None and go.go_data_pedido_expedido:
        return 4
    producao_finalizada = pedido is not None and pedido.status_producao == "FINALIZADO"
    if producao_finalizada or (go is not None and go.go_data_efetiva_liberacao_pcp):
        return 3
    em_producao = pedido is not None and any(item.inicio_producao for item in pedido.itens)
    if em_producao:
        return 2
    return 1


def _etapas_acompanhamento_pedido(pedido, go):
    etapa_atual = _indice_etapa_pedido(pedido, go)
    etapas = []
    for i, base in enumerate(_ETAPAS_ACOMPANHAMENTO_PEDIDO, start=1):
        if i < etapa_atual:
            estado = "concluida"
        elif i == etapa_atual:
            estado = "atual"
        else:
            estado = "pendente"
        etapas.append({**base, "numero": i, "estado": estado, "linha_concluida": i <= etapa_atual})
    return etapas


# Campos "comercial" que só existem em PedidoOperacao (não têm equivalente em
# Pedido) — pedido do Bruno, 01/09/2026: quem inclui um pedido novo em Gestão
# Produção (PCP) passa a preencher também essas informações do pedido como um
# todo (não só do produto), e ao salvar isso alimenta automaticamente a
# Listagem Geral de Gestão Operação.
CAMPOS_GO_COMERCIAL_NOVO_PEDIDO = [
    "go_tipo_pedido", "go_contrato", "go_proposta", "go_pedido_compra_cliente",
    "go_status_pedido_info", "go_valor_pedido_operacao",
]


def _criar_pedido_operacao_a_partir_de_producao(pedido, f):
    """Cria automaticamente 1 PedidoOperacao a partir de um Pedido recém-
    incluído em Gestão Produção (pedido do Bruno, 01/09/2026) — SÓ na
    inclusão (não em edição futura), e SÓ isso: um cópia inicial dos campos,
    não um vínculo permanente. PedidoOperacao continua sendo uma tabela
    INDEPENDENTE (sem FK) — dali em diante os dois são editados cada um na
    sua própria tela, sem sincronização automática nenhuma (é o próprio
    Bruno quem vai "manusear manualmente entre PCP/Logística/Resultados").

    "Data solicitada entrega" (Comercial) E "Solicitada cliente/retira" (PCP)
    de Gestão Operação são o mesmo dado que "Data do cliente" de Gestão
    Produção (já documentado em _LinhaListagemGeral) — pedido do Bruno
    (01/09/2026): "Solicitada cliente/retira" também tem que acompanhar essa
    data desde a inclusão do pedido, não é um campo novo — por isso não
    existe um campo novo pra ela no formulário de Produção, só reaproveita
    pedido.data_cliente pros dois."""
    novo = PedidoOperacao(
        cliente=pedido.cliente,
        vendedor=pedido.vendedor,
        pedido_venda=pedido.pedido_venda,
        data_inclusao_pedido=pedido.data_inclusao_pedido,
        prioridade=pedido.prioridade,
        frete=pedido.frete,
        pais=pedido.pais,
        estado=pedido.estado,
        cidade=pedido.cidade,
        go_data_solicitada_entrega=pedido.data_cliente,
        go_data_solicitada_cliente_retira=pedido.data_cliente,
        go_tipo_pedido=f.get("go_tipo_pedido", "").strip() or None,
        go_contrato=f.get("go_contrato", "").strip() or None,
        go_proposta=f.get("go_proposta", "").strip() or None,
        go_pedido_compra_cliente=f.get("go_pedido_compra_cliente", "").strip() or None,
        go_status_pedido_info=f.get("go_status_pedido_info", "").strip() or None,
        go_valor_pedido_operacao=(
            _parse_float_form(f.get("go_valor_pedido_operacao"), default=None)
            if f.get("go_valor_pedido_operacao", "").strip()
            else None
        ),
    )
    db.session.add(novo)
    return novo


# ----------------------------------------------------------------------
# Qualidade — RNC (Relatório de Não Conformidade). Área nova, independente
# de Gestão Produção/Operação (RncQualidade não tem FK com nada). Mesmo
# padrão de filtro/paginação de _filtrar_pedidos_operacao/_linhas_gestao_operacao.
# ----------------------------------------------------------------------
CAMPOS_HISTORICO_RNC = [
    "descricao_nc", "tipo_nc", "severidade", "causa_raiz", "disposicao_produto",
    "acao_corretiva_descricao", "responsavel_acao_corretiva", "prazo_acao_corretiva",
    "status_acao_corretiva", "eficacia_acao", "status_geral", "data_fechamento",
]


def _filtrar_rnc_qualidade(args):
    """`args` é sempre `request.args` (MultiDict) — os 5 filtros de "múltiplas
    opções" (status geral, severidade, origem, tipo de NC, setor) usam
    `getlist`, porque um <select multiple> manda um par nome=valor repetido
    pra cada opção marcada (pedido do Bruno: poder marcar mais de uma opção
    no mesmo filtro, ex.: Espumagem + PU, ou Dureza + Dimensional)."""
    query = RncQualidade.query

    busca = args.get("busca", "").strip()
    status_geral = [v for v in args.getlist("status_geral") if v]
    severidade = [v for v in args.getlist("severidade") if v]
    origem = [v for v in args.getlist("origem") if v]
    tipo_nc = [v for v in args.getlist("tipo_nc") if v]
    setor = [v for v in args.getlist("setor") if v]
    mes_emissao = args.get("mes_emissao", "").strip()  # "AAAA-MM", do <input type="month">
    apenas_abertas = args.get("apenas_abertas", "").strip()

    if busca:
        like = f"%{busca}%"
        query = query.filter(
            or_(
                RncQualidade.numero_rnc.ilike(like),
                RncQualidade.cliente_projeto.ilike(like),
                RncQualidade.produto_equipamento.ilike(like),
                RncQualidade.numero_pedido_contrato.ilike(like),
            )
        )
    if status_geral:
        query = query.filter(RncQualidade.status_geral.in_(status_geral))
    if severidade:
        query = query.filter(RncQualidade.severidade.in_(severidade))
    if origem:
        query = query.filter(RncQualidade.origem.in_(origem))
    if tipo_nc:
        query = query.filter(RncQualidade.tipo_nc.in_(tipo_nc))
    if setor:
        query = query.filter(RncQualidade.setor.in_(setor))
    if mes_emissao:
        try:
            ano_m, mes_m = (int(p) for p in mes_emissao.split("-"))
            inicio = date(ano_m, mes_m, 1)
            fim = date(ano_m, mes_m, monthrange(ano_m, mes_m)[1])
            query = query.filter(RncQualidade.data_emissao.between(inicio, fim))
        except (ValueError, TypeError):
            mes_emissao = ""  # valor incompreensível — ignora o filtro em vez de quebrar a busca
    if apenas_abertas == "1":
        query = query.filter(RncQualidade.status_geral.in_(RNC_STATUS_GERAL_ABERTOS))

    query = query.order_by(RncQualidade.data_emissao.desc().nullslast(), RncQualidade.id.desc())

    filtros = dict(
        busca=busca, status_geral=status_geral, severidade=severidade,
        origem=origem, tipo_nc=tipo_nc, setor=setor, mes_emissao=mes_emissao,
        apenas_abertas=apenas_abertas,
    )
    return query, filtros


def _rnc_opcoes_filtro(campo, opcoes_curadas):
    """Opções de um filtro de RNC (multi-seleção) — combina os valores REAIS
    já cadastrados nesse campo (que podem não bater com a lista de sugestão,
    já que todo campo de "lista" do RNC é texto livre — ver RNC_*_OPCOES em
    models.py) com a lista de sugestão, pra sempre dar pra filtrar por
    qualquer valor que já apareça em algum RNC, mesmo com grafia diferente
    da lista padrão (ex.: "Dureza" na planilha do Bruno vs. "Dureza /
    Material" na lista de sugestão)."""
    coluna = getattr(RncQualidade, campo)
    existentes = [v for (v,) in db.session.query(coluna).filter(coluna.isnot(None)).distinct()]
    existentes_lower = {v.lower() for v in existentes}
    extras = [op for op in opcoes_curadas if op.lower() not in existentes_lower]
    return sorted(existentes + extras, key=lambda s: s.lower())


def _linhas_rnc_qualidade(args):
    page = args.get("page", 1, type=int)
    query, filtros = _filtrar_rnc_qualidade(args)
    total_filtrado = query.count()
    total_paginas = max(1, (total_filtrado + PAGE_SIZE - 1) // PAGE_SIZE)
    pagina = query.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    return pagina, page, total_paginas, total_filtrado, filtros


def _dashboard_rnc_qualidade():
    """Recalcula ao vivo os mesmos KPIs/quebras da aba "Dashboard" da planilha
    do Bruno, direto de RncQualidade — nada é guardado pré-calculado, então
    fica sempre em dia com o que estiver cadastrado (diferente da planilha
    original, que só atualizava quando alguém reabria o arquivo)."""
    rncs = RncQualidade.query.all()
    total = len(rncs)
    abertas = [r for r in rncs if r.esta_aberto]
    fechadas = [r for r in rncs if not r.esta_aberto]

    com_reincidencia = [r for r in rncs if (r.reincidencia or "").strip().lower() == "sim"]
    pct_reincidencia = round((len(com_reincidencia) / total) * 100, 1) if total else 0

    acoes_nao_eficazes = sum(1 for r in rncs if r.eficacia_acao == "Não Eficaz")
    criticas_abertas = sum(1 for r in rncs if r.severidade == "Crítica" and r.esta_aberto)

    dias_abertos = [r.dias_em_aberto for r in rncs if r.dias_em_aberto is not None]
    tempo_medio_aberto = round(sum(dias_abertos) / len(dias_abertos), 1) if dias_abertos else 0

    custo_total = sum(r.custo_estimado or 0 for r in rncs)

    def _quebra_por(atributo, opcoes_ordem=None):
        contagem = {}
        for r in rncs:
            chave = getattr(r, atributo) or "—"
            contagem[chave] = contagem.get(chave, 0) + 1
        if opcoes_ordem:
            chaves = list(opcoes_ordem) + sorted(k for k in contagem if k not in opcoes_ordem and k != "—")
            if "—" in contagem:
                chaves.append("—")
            return [{"chave": k, "total": contagem.get(k, 0)} for k in chaves if k in contagem or k in opcoes_ordem]
        return sorted(({"chave": k, "total": v} for k, v in contagem.items()), key=lambda d: -d["total"])

    return {
        "total": total,
        "abertas": len(abertas),
        "fechadas": len(fechadas),
        "pct_reincidencia": pct_reincidencia,
        "acoes_nao_eficazes": acoes_nao_eficazes,
        "criticas_abertas": criticas_abertas,
        "tempo_medio_aberto": tempo_medio_aberto,
        "custo_total": custo_total,
        "por_origem": _quebra_por("origem", RNC_ORIGEM_OPCOES),
        "por_tipo_nc": _quebra_por("tipo_nc", RNC_TIPO_NC_OPCOES),
        "por_severidade": _quebra_por("severidade", RNC_SEVERIDADE_OPCOES),
        "por_status_geral": _quebra_por("status_geral", RNC_STATUS_GERAL_OPCOES),
        "por_eficacia": _quebra_por("eficacia_acao", RNC_EFICACIA_OPCOES),
        "por_status_acao": _quebra_por("status_acao_corretiva", RNC_STATUS_ACAO_OPCOES),
    }


def _campos_form_rnc(f):
    """Lê e converte todos os campos do formulário de RNC (novo/editar) — uma
    função só, reaproveitada pelas duas rotas, pra não duplicar a conversão
    de cada um dos ~35 campos editáveis."""
    def _txt(nome):
        return f.get(nome, "").strip() or None

    def _num_int(nome):
        valor = f.get(nome, "").strip()
        try:
            return int(valor) if valor else None
        except ValueError:
            return None

    def _num_float(nome):
        valor = f.get(nome, "").strip().replace(",", ".")
        try:
            return float(valor) if valor else None
        except ValueError:
            return None

    return dict(
        numero_rnc=_txt("numero_rnc"),
        revisao=_num_int("revisao") or 0,
        data_emissao=_parse_data_form(f.get("data_emissao")),
        emitente=_txt("emitente"),
        setor=_txt("setor"),
        origem=_txt("origem"),
        cliente_projeto=_txt("cliente_projeto"),
        numero_pedido_contrato=_txt("numero_pedido_contrato"),
        produto_equipamento=_txt("produto_equipamento"),
        numero_op=_txt("numero_op"),
        local_setor=_txt("local_setor"),
        data_identificacao=_parse_data_form(f.get("data_identificacao")),
        responsavel_identificacao=_txt("responsavel_identificacao"),
        descricao_nc=_txt("descricao_nc"),
        qtd_nao_conforme=_num_int("qtd_nao_conforme"),
        requisito_nao_atendido=_txt("requisito_nao_atendido"),
        tipo_nc=_txt("tipo_nc"),
        severidade=_txt("severidade"),
        acao_contencao_imediata=_txt("acao_contencao_imediata"),
        porque_1=_txt("porque_1"),
        porque_2=_txt("porque_2"),
        porque_3=_txt("porque_3"),
        porque_4=_txt("porque_4"),
        porque_5=_txt("porque_5"),
        causa_raiz=_txt("causa_raiz"),
        ferramenta_analise=_txt("ferramenta_analise"),
        disposicao_produto=_txt("disposicao_produto"),
        acao_corretiva_descricao=_txt("acao_corretiva_descricao"),
        responsavel_acao_corretiva=_txt("responsavel_acao_corretiva"),
        prazo_acao_corretiva=_parse_data_form(f.get("prazo_acao_corretiva")),
        data_realizacao=_parse_data_form(f.get("data_realizacao")),
        status_acao_corretiva=_txt("status_acao_corretiva"),
        data_verificacao_eficacia=_parse_data_form(f.get("data_verificacao_eficacia")),
        eficacia_acao=_txt("eficacia_acao"),
        obs_verificacao=_txt("obs_verificacao"),
        reincidencia=_txt("reincidencia"),
        numero_rnc_relacionada=_txt("numero_rnc_relacionada"),
        custo_estimado=_num_float("custo_estimado"),
        status_geral=_txt("status_geral") or "Aberto",
        responsavel_qualidade=_txt("responsavel_qualidade"),
        data_fechamento=_parse_data_form(f.get("data_fechamento")),
        evidencias_anexos=_txt("evidencias_anexos"),
        observacoes_gerais=_txt("observacoes_gerais"),
    )


_CAMPOS_DATA_RNC = [
    "data_emissao", "data_identificacao", "prazo_acao_corretiva", "data_realizacao",
    "data_verificacao_eficacia", "data_fechamento",
]


def _rnc_para_form_dict(rnc):
    """Converte um RncQualidade em dict de strings prontas pra repopular o
    formulário HTML (mesmo formato que os <input> mandam de volta) — usado
    na tela de edição (GET). Datas em ISO (yyyy-mm-dd, o que <input type=date>
    espera); os demais campos, string direta ou vazio."""
    campos = [
        "numero_rnc", "revisao", "data_emissao", "emitente", "setor", "origem",
        "cliente_projeto", "numero_pedido_contrato", "produto_equipamento", "numero_op",
        "local_setor", "data_identificacao", "responsavel_identificacao", "descricao_nc",
        "qtd_nao_conforme", "requisito_nao_atendido", "tipo_nc", "severidade",
        "acao_contencao_imediata", "porque_1", "porque_2", "porque_3", "porque_4", "porque_5",
        "causa_raiz", "ferramenta_analise", "disposicao_produto", "acao_corretiva_descricao",
        "responsavel_acao_corretiva", "prazo_acao_corretiva", "data_realizacao",
        "status_acao_corretiva", "data_verificacao_eficacia", "eficacia_acao", "obs_verificacao",
        "reincidencia", "numero_rnc_relacionada", "custo_estimado", "status_geral",
        "responsavel_qualidade", "data_fechamento", "evidencias_anexos", "observacoes_gerais",
    ]
    valores = {}
    for campo in campos:
        v = getattr(rnc, campo)
        if v is None:
            valores[campo] = ""
        elif campo in _CAMPOS_DATA_RNC:
            valores[campo] = v.isoformat()
        else:
            valores[campo] = str(v)
    return valores


# ----------------------------------------------------------------------
# Qualidade — Inspeção Final / RDIM (pedido do Bruno, 02/09/2026). Ao
# contrário da RNC, usa FK real pra ItemPedido — "OP" nesta base é o próprio
# ItemPedido. Mesmo padrão de filtro/paginação de _filtrar_rnc_qualidade.
# ----------------------------------------------------------------------
CAMPOS_HISTORICO_INSPECAO_FINAL = [
    "resultado", "categoria_desvio", "subcategoria_desvio", "desvio_encontrado",
    "observacao", "inspecao_visual", "quantidade_com_desvio",
]


def _itens_rdim_disponiveis(termo, limite=20):
    """Itens de pedido inspecionáveis pelo RDIM — busca em TODO o banco de
    Gestão Produção, sem restrição de estação (pedido do Bruno, 03/09/2026:
    "quero que os campos qualidade puxe todo o banco de dados da produção" —
    ele foi apontar um desvio no RDIM e a busca da OP não encontrou o
    pedido, porque a busca antiga só olhava itens das estações MANDRIL/PU/
    SILICONE. RDIM_ESTACOES_OPCOES continua existindo só como ordem de
    exibição no dashboard/filtro — não é mais um filtro de elegibilidade).
    Busca por descrição do produto, cliente ou nº do pedido de venda, igual
    ao padrão já usado em _buscar_pedidos_para_status."""
    termo = (termo or "").strip()
    query = ItemPedido.query.join(Pedido)
    if termo:
        like = f"%{termo}%"
        query = query.filter(
            or_(
                ItemPedido.descricao_produto.ilike(like),
                Pedido.cliente.ilike(like),
                Pedido.pedido_venda.ilike(like),
            )
        )
    itens = query.order_by(ItemPedido.atualizado_em.desc()).limit(limite).all()
    return [
        {
            "item_id": item.id,
            "cliente": item.pedido.cliente if item.pedido else "",
            "pedido_venda": item.pedido.pedido_venda if item.pedido else "",
            "produto": item.descricao_produto,
            "quantidade": item.quantidade,
            "estacao": item.estacao,
        }
        for item in itens
    ]


def _filtrar_inspecoes_finais(args):
    """`args` é sempre `request.args`. Mesmo espírito de _filtrar_rnc_qualidade:
    multi-seleção (Ctrl+clique) pra resultado/categoria de desvio/estação,
    texto livre pra cliente/produto/pedido/OP, período por data de inspeção."""
    query = InspecaoFinal.query.join(ItemPedido).join(Pedido)

    busca = args.get("busca", "").strip()
    dn_polegada = args.get("dn_polegada", "").strip()
    resultado = [v for v in args.getlist("resultado") if v]
    categoria_desvio = [v for v in args.getlist("categoria_desvio") if v]
    subcategoria_desvio = [v for v in args.getlist("subcategoria_desvio") if v]
    estacao = [v for v in args.getlist("estacao") if v]
    responsavel_id = args.get("responsavel_id", "").strip()
    data_de = args.get("data_de", "").strip()
    data_ate = args.get("data_ate", "").strip()
    mes = args.get("mes", "").strip()  # "AAAA-MM", do <input type="month"> — mesmo padrão de mes_emissao (RNC)

    if busca:
        like = f"%{busca}%"
        query = query.filter(
            or_(
                Pedido.cliente.ilike(like),
                Pedido.pedido_venda.ilike(like),
                ItemPedido.descricao_produto.ilike(like),
                InspecaoFinal.numero_rif.ilike(like),
            )
        )
    if dn_polegada:
        # Filtro separado do "busca" geral (pedido do Bruno, 02/09/2026,
        # RDIM Fase 4) — não existe campo estruturado de DN/polegada no
        # sistema, a medida do produto vem embutida em texto livre dentro de
        # descricao_produto (ex.: "DISCO SELO 18''"). Casamento só contra a
        # descrição do produto, de propósito: se entrasse no OR do "busca"
        # geral, um número de pedido ou RIF que por coincidência contivesse
        # os mesmos dígitos do DN daria falso positivo.
        query = query.filter(ItemPedido.descricao_produto.ilike(f"%{dn_polegada}%"))
    if resultado:
        query = query.filter(InspecaoFinal.resultado.in_(resultado))
    if categoria_desvio:
        query = query.filter(InspecaoFinal.categoria_desvio.in_(categoria_desvio))
    if subcategoria_desvio:
        query = query.filter(InspecaoFinal.subcategoria_desvio.in_(subcategoria_desvio))
    if estacao:
        query = query.filter(InspecaoFinal.estacao.in_(estacao))
    if responsavel_id:
        try:
            query = query.filter(InspecaoFinal.responsavel_id == int(responsavel_id))
        except ValueError:
            responsavel_id = ""
    if data_de:
        try:
            query = query.filter(InspecaoFinal.data_inspecao >= date.fromisoformat(data_de))
        except ValueError:
            data_de = ""
    if data_ate:
        try:
            query = query.filter(InspecaoFinal.data_inspecao <= date.fromisoformat(data_ate))
        except ValueError:
            data_ate = ""
    if mes:
        try:
            ano_m, mes_m = (int(p) for p in mes.split("-"))
            inicio = date(ano_m, mes_m, 1)
            fim = date(ano_m, mes_m, monthrange(ano_m, mes_m)[1])
            query = query.filter(InspecaoFinal.data_inspecao.between(inicio, fim))
        except (ValueError, TypeError):
            mes = ""  # valor incompreensível — ignora o filtro em vez de quebrar a busca

    query = query.order_by(InspecaoFinal.data_inspecao.desc().nullslast(), InspecaoFinal.id.desc())

    filtros = dict(
        busca=busca, dn_polegada=dn_polegada, resultado=resultado, categoria_desvio=categoria_desvio,
        subcategoria_desvio=subcategoria_desvio, estacao=estacao,
        responsavel_id=responsavel_id, data_de=data_de, data_ate=data_ate, mes=mes,
    )
    return query, filtros


def _linhas_inspecoes_finais(args):
    page = args.get("page", 1, type=int)
    query, filtros = _filtrar_inspecoes_finais(args)
    total_filtrado = query.count()
    total_paginas = max(1, (total_filtrado + PAGE_SIZE - 1) // PAGE_SIZE)
    pagina = query.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    return pagina, page, total_paginas, total_filtrado, filtros


def _dashboard_rdim():
    """Recalcula ao vivo os indicadores da Inspeção Final — mesmo espírito de
    _dashboard_rnc_qualidade, nada pré-calculado/guardado."""
    inspecoes = InspecaoFinal.query.options(selectinload(InspecaoFinal.item).selectinload(ItemPedido.pedido)).all()
    total = len(inspecoes)

    aprovadas = sum(1 for i in inspecoes if i.resultado == "APROVADO")
    reprovadas = sum(1 for i in inspecoes if i.resultado == "REPROVADO")
    aprovadas_desvio = sum(1 for i in inspecoes if i.resultado == "APROVADO_COM_DESVIO")
    pct_aprovacao = round((aprovadas / total) * 100, 1) if total else 0
    pct_reprovacao = round((reprovadas / total) * 100, 1) if total else 0

    # Volume de peças (não só de inspeções) — pedido do Bruno (02/09/2026):
    # "lote total contém 5 peças, mas dessas 2 unidades ficou com desvio".
    # Sempre em relação ao lote inteiro (item.quantidade), decisão já
    # confirmada com ele.
    total_quantidade_inspecionada = sum((i.item.quantidade or 0) for i in inspecoes if i.item)
    total_quantidade_com_desvio = sum(i.quantidade_com_desvio or 0 for i in inspecoes)
    pct_pecas_desvio = (
        round((total_quantidade_com_desvio / total_quantidade_inspecionada) * 100, 1)
        if total_quantidade_inspecionada else 0
    )

    def _quebra_por(chave_fn, opcoes_ordem=None):
        contagem = {}
        for i in inspecoes:
            chave = chave_fn(i) or "—"
            contagem[chave] = contagem.get(chave, 0) + 1
        if opcoes_ordem:
            chaves = list(opcoes_ordem) + sorted(k for k in contagem if k not in opcoes_ordem and k != "—")
            if "—" in contagem:
                chaves.append("—")
            return [{"chave": k, "total": contagem.get(k, 0)} for k in chaves if k in contagem]
        return sorted(({"chave": k, "total": v} for k, v in contagem.items()), key=lambda d: -d["total"])[:10]

    # Produtos/clientes com maior índice de REPROVAÇÃO (não só volume) — só
    # entram no ranking quem já teve ao menos 1 inspeção reprovada.
    def _ranking_reprovacao(chave_fn):
        por_chave = {}
        for i in inspecoes:
            chave = chave_fn(i)
            if not chave:
                continue
            d = por_chave.setdefault(chave, {"total": 0, "reprovadas": 0})
            d["total"] += 1
            if i.resultado == "REPROVADO":
                d["reprovadas"] += 1
        linhas = [
            {"chave": k, "total": v["total"], "reprovadas": v["reprovadas"],
             "pct": round((v["reprovadas"] / v["total"]) * 100, 1)}
            for k, v in por_chave.items() if v["reprovadas"] > 0
        ]
        return sorted(linhas, key=lambda d: (-d["reprovadas"], -d["pct"]))[:10]

    # Evolução mensal de reprovações — últimos 12 meses com pelo menos uma
    # inspeção, em ordem cronológica (pro gráfico de linha).
    por_mes = {}
    for i in inspecoes:
        if not i.data_inspecao:
            continue
        chave = i.data_inspecao.strftime("%m/%Y")
        d = por_mes.setdefault(chave, {"ord": i.data_inspecao.replace(day=1), "total": 0, "reprovadas": 0})
        d["total"] += 1
        if i.resultado == "REPROVADO":
            d["reprovadas"] += 1
    evolucao = [
        {"mes": k, "total": v["total"], "reprovadas": v["reprovadas"]}
        for k, v in sorted(por_mes.items(), key=lambda kv: kv[1]["ord"])
    ][-12:]

    return {
        "total": total,
        "aprovadas": aprovadas,
        "reprovadas": reprovadas,
        "aprovadas_desvio": aprovadas_desvio,
        "pct_aprovacao": pct_aprovacao,
        "pct_reprovacao": pct_reprovacao,
        "por_categoria_desvio": _quebra_por(lambda i: i.categoria_desvio, RDIM_CATEGORIA_DESVIO_OPCOES),
        "por_subcategoria_desvio": _quebra_por(lambda i: i.subcategoria_desvio, RDIM_SUBCATEGORIA_DESVIO_OPCOES),
        "por_estacao": _quebra_por(lambda i: i.estacao, RDIM_ESTACOES_OPCOES),
        "ranking_produtos": _ranking_reprovacao(lambda i: i.produto),
        "ranking_clientes": _ranking_reprovacao(lambda i: i.cliente),
        "evolucao": evolucao,
        "total_quantidade_inspecionada": total_quantidade_inspecionada,
        "total_quantidade_com_desvio": total_quantidade_com_desvio,
        "pct_pecas_desvio": pct_pecas_desvio,
    }


def _validar_quantidade_com_desvio(valor_form, quantidade_item):
    """Lê o campo "Quantidade com desvio" do formulário de inspeção RDIM e
    valida contra a quantidade do lote (item.quantidade) — pedido do Bruno
    (02/09/2026): "lote total contém 5 peças, mas dessas 2 unidades ficou
    com desvio", sempre em relação ao lote inteiro. Retorna (valor, erro) —
    `erro` é None quando válido; quando não, a rota deve mostrar o flash e
    NÃO salvar (mesmo padrão de bloqueio já usado pra resultado/estação).

    `default=None` explícito no _parse_float_form: o default 0.0 da função
    faria todo campo em branco virar "0 peças com desvio" em vez de "não
    informado" — mesma armadilha já corrigida uma vez nas medições."""
    quantidade_com_desvio = _parse_float_form(valor_form, default=None)
    if quantidade_com_desvio is None:
        return None, None
    if quantidade_com_desvio < 0:
        return None, "Quantidade com desvio não pode ser negativa."
    if quantidade_item is not None and quantidade_com_desvio > quantidade_item:
        return None, (
            f"Quantidade com desvio ({quantidade_com_desvio:g}) não pode ser maior que "
            f"a quantidade do lote ({quantidade_item:g})."
        )
    return quantidade_com_desvio, None


def _salvar_medicoes_rdim(inspecao, f, substituir=False):
    """Grava as linhas de medição (grandeza + especificação + faixa medida)
    a partir dos arrays paralelos do formulário — mesmo padrão de arrays
    paralelos (zip) já usado pros itens de um pedido em editar_pedido. Em
    edição, `substituir=True` apaga as medições antigas e recria do zero: é
    mais simples que casar por id e cobre bem o caso comum (poucas linhas,
    reescritas inteiras a cada salvamento) sem precisar de histórico por
    medição — só a InspecaoFinal como um todo entra no histórico de alteração."""
    if substituir:
        for m in list(inspecao.medicoes):
            db.session.delete(m)

    grandezas = f.getlist("grandeza[]")
    esp_mins = f.getlist("especificado_min[]")
    esp_maxs = f.getlist("especificado_max[]")
    med_mins = f.getlist("medido_min[]")
    med_maxs = f.getlist("medido_max[]")

    ordem = 0
    for grandeza, esp_min, esp_max, med_min, med_max in zip(grandezas, esp_mins, esp_maxs, med_mins, med_maxs):
        grandeza = grandeza.strip()
        if not grandeza:
            continue
        db.session.add(
            RdimMedicao(
                inspecao=inspecao,
                grandeza=grandeza,
                especificado_min=_parse_float_form(esp_min, default=None),
                especificado_max=_parse_float_form(esp_max, default=None),
                medido_min=_parse_float_form(med_min, default=None),
                medido_max=_parse_float_form(med_max, default=None),
                ordem=ordem,
            )
        )
        ordem += 1


def _salvar_pecas_desvio_rdim(inspecao, f, substituir=False):
    """Grava o detalhamento peça a peça do desvio (nº da peça + característica
    + valor medido + especificação mín/máx daquela peça) — pedido do Bruno
    (02/09/2026, RDIM Fase 3 e, pros campos especificado_min/max, Fase 4:
    "tolerância era de 0,5mm, peça inspecionada com 0,7mm, peça ficou 0,2mm
    acima da tolerância"), mesmo padrão de arrays paralelos (zip) já usado em
    _salvar_medicoes_rdim. Linhas sem característica selecionada são
    ignoradas (o nº da peça sozinho não basta pra fazer sentido). Não mexe em
    quantidade_com_desvio — os dois campos são independentes, por decisão do
    Bruno."""
    if substituir:
        for p in list(inspecao.pecas_desvio):
            db.session.delete(p)

    pecas_numero = f.getlist("peca_numero[]")
    pecas_caracteristica = f.getlist("peca_caracteristica[]")
    pecas_valor = f.getlist("peca_valor_medido[]")
    pecas_espec_min = f.getlist("peca_especificado_min[]")
    pecas_espec_max = f.getlist("peca_especificado_max[]")

    ordem = 0
    for peca_numero, caracteristica, valor_medido, espec_min, espec_max in zip_longest(
        pecas_numero, pecas_caracteristica, pecas_valor, pecas_espec_min, pecas_espec_max, fillvalue=""
    ):
        caracteristica = (caracteristica or "").strip()
        if not caracteristica:
            continue
        db.session.add(
            RdimPecaDesvio(
                inspecao=inspecao,
                peca_numero=(peca_numero or "").strip() or None,
                caracteristica=caracteristica,
                valor_medido=_parse_float_form(valor_medido, default=None),
                especificado_min=_parse_float_form(espec_min, default=None),
                especificado_max=_parse_float_form(espec_max, default=None),
                ordem=ordem,
            )
        )
        ordem += 1


def _inspecoes_rdim_por_item(item_ids):
    """dict item_pedido_id -> InspecaoFinal mais recente (data_inspecao
    desc, depois id desc) — 1 query com IN, sem N+1. Pedido do Bruno
    (02/09/2026): conectar o RDIM com Produção/Operação, mostrando o status
    de qualidade de cada item nas telas que já existem (Consulta Pedido,
    Detalhe do Pedido, Listagem Geral) sem duplicar a lógica em cada uma.
    Se um item tiver mais de uma inspeção, mostra só a mais recente — a
    tabela InspecaoFinal.query já vem ordenada, `setdefault` fica só com a
    primeira ocorrência de cada item_pedido_id."""
    ids = sorted({i for i in item_ids if i})
    if not ids:
        return {}
    inspecoes = (
        InspecaoFinal.query
        .filter(InspecaoFinal.item_pedido_id.in_(ids))
        .order_by(InspecaoFinal.item_pedido_id, InspecaoFinal.data_inspecao.desc().nullslast(), InspecaoFinal.id.desc())
        .all()
    )
    mapa = {}
    for insp in inspecoes:
        mapa.setdefault(insp.item_pedido_id, insp)
    return mapa


def _resumo_rdim_pedido(inspecoes):
    """Agrega as InspecaoFinal de UM pedido (normalmente vindas de
    _inspecoes_rdim_por_item(...).values()) em contagens por resultado + o
    quantitativo de peças com desvio/total — usado no bloco de resumo de
    Qualidade da Consulta Pedido e do Detalhe do Pedido. Retorna None se a
    lista vier vazia (pedido sem nenhuma inspeção RDIM ainda)."""
    inspecoes = list(inspecoes)
    if not inspecoes:
        return None
    return {
        "total": len(inspecoes),
        "aprovadas": sum(1 for i in inspecoes if i.resultado == "APROVADO"),
        "reprovadas": sum(1 for i in inspecoes if i.resultado == "REPROVADO"),
        "aprovadas_desvio": sum(1 for i in inspecoes if i.resultado == "APROVADO_COM_DESVIO"),
        "quantidade_com_desvio": sum(i.quantidade_com_desvio or 0 for i in inspecoes),
        "quantidade_total": sum((i.item.quantidade or 0) for i in inspecoes if i.item),
    }


def _rdim_resumo_por_pedido_venda(pedidos_venda):
    """dict pedido_venda (trim) -> {total, reprovadas, com_desvio,
    quantidade_com_desvio} — pro indicador (mais simples, pedido-level) de
    Qualidade na Listagem Geral de Gestão Operação. Mesmo casamento por
    texto (sem FK, trim() nos dois lados) já usado por
    _itens_producao_por_pedido_venda — só pra exibir uma dica visual, nunca
    aproximado."""
    valores = sorted({v.strip() for v in pedidos_venda if v and v.strip()})
    if not valores:
        return {}
    inspecoes = (
        InspecaoFinal.query.join(ItemPedido).join(Pedido)
        .options(selectinload(InspecaoFinal.item).selectinload(ItemPedido.pedido))
        .filter(func.trim(Pedido.pedido_venda).in_(valores))
        .all()
    )
    mapa = {}
    for insp in inspecoes:
        chave = (insp.pedido_venda or "").strip()
        if not chave:
            continue
        d = mapa.setdefault(chave, {"total": 0, "reprovadas": 0, "com_desvio": 0, "quantidade_com_desvio": 0})
        d["total"] += 1
        if insp.resultado == "REPROVADO":
            d["reprovadas"] += 1
        if insp.resultado == "APROVADO_COM_DESVIO":
            d["com_desvio"] += 1
        d["quantidade_com_desvio"] += insp.quantidade_com_desvio or 0
    return mapa


# ----------------------------------------------------------------------
# P&D — Pesquisa e Desenvolvimento (Fase 14). Segue o mesmo padrão de
# Qualidade/RNC acima: tabela própria, controle manual, funções auxiliares
# separadas de filtro/listagem/dashboard/form pra não misturar com nenhuma
# outra área do sistema.
# ----------------------------------------------------------------------
def _campos_form_pd(f):
    """Lê e converte todos os campos do formulário de Projeto de P&D (novo/
    editar) — mesma estrutura de _campos_form_rnc."""
    def _txt(nome):
        return f.get(nome, "").strip() or None

    def _num_int(nome):
        valor = f.get(nome, "").strip()
        try:
            return int(valor) if valor else None
        except ValueError:
            return None

    def _num_float(nome):
        valor = f.get(nome, "").strip().replace(",", ".")
        try:
            return float(valor) if valor else None
        except ValueError:
            return None

    resultado_esperado = "; ".join(f.getlist("resultado_esperado")) or None

    return dict(
        codigo=_txt("codigo"),
        nome=_txt("nome"),
        descricao=_txt("descricao"),
        objetivo=_txt("objetivo"),
        justificativa=_txt("justificativa"),
        categoria=_txt("categoria"),
        prioridade=_txt("prioridade"),
        responsavel=_txt("responsavel"),
        participantes=_txt("participantes"),
        cliente=_txt("cliente"),
        produto=_txt("produto"),
        fornecedor=_txt("fornecedor"),
        area_envolvida=_txt("area_envolvida"),
        etapa_atual=_txt("etapa_atual") or "Ideia",
        percentual_conclusao=_num_int("percentual_conclusao") or 0,
        data_inicio=_parse_data_form(f.get("data_inicio")),
        data_prevista_conclusao=_parse_data_form(f.get("data_prevista_conclusao")),
        data_real_conclusao=_parse_data_form(f.get("data_real_conclusao")),
        proxima_entrega=_txt("proxima_entrega"),
        data_proxima_entrega=_parse_data_form(f.get("data_proxima_entrega")),
        responsavel_proxima_entrega=_txt("responsavel_proxima_entrega"),
        custo_previsto=_num_float("custo_previsto"),
        custo_realizado=_num_float("custo_realizado"),
        investimento_previsto=_num_float("investimento_previsto"),
        investimento_realizado=_num_float("investimento_realizado"),
        economia_prevista=_num_float("economia_prevista"),
        economia_realizada=_num_float("economia_realizada"),
        resultado_esperado=resultado_esperado,
        resultado_obtido=_txt("resultado_obtido"),
        problema=_txt("problema"),
        solucao=_txt("solucao"),
        licoes_aprendidas=_txt("licoes_aprendidas"),
        observacoes_gerais=_txt("observacoes_gerais"),
    )


_CAMPOS_DATA_PD = [
    "data_inicio", "data_prevista_conclusao", "data_real_conclusao", "data_proxima_entrega",
]


def _pd_para_form_dict(projeto):
    """Converte um ProjetoPD em dict de strings prontas pra repopular o
    formulário HTML — usado na tela de edição (GET), igual a _rnc_para_form_dict."""
    campos = [
        "codigo", "nome", "descricao", "objetivo", "justificativa", "categoria", "prioridade",
        "responsavel", "participantes", "cliente", "produto", "fornecedor", "area_envolvida",
        "etapa_atual", "percentual_conclusao", "data_inicio", "data_prevista_conclusao",
        "data_real_conclusao", "proxima_entrega", "data_proxima_entrega", "responsavel_proxima_entrega",
        "custo_previsto", "custo_realizado", "investimento_previsto", "investimento_realizado",
        "economia_prevista", "economia_realizada", "resultado_obtido", "problema", "solucao",
        "licoes_aprendidas", "observacoes_gerais",
    ]
    valores = {}
    for campo in campos:
        v = getattr(projeto, campo)
        if v is None:
            valores[campo] = ""
        elif campo in _CAMPOS_DATA_PD:
            valores[campo] = v.isoformat()
        else:
            valores[campo] = str(v)
    valores["resultado_esperado"] = projeto.resultado_esperado_lista
    return valores


def _filtrar_projetos_pd(args):
    busca = args.get("busca", "").strip()
    etapa = args.getlist("etapa")
    categoria = args.getlist("categoria")
    responsavel = args.get("responsavel", "").strip()
    prioridade = args.getlist("prioridade")
    cliente = args.get("cliente", "").strip()
    produto = args.get("produto", "").strip()
    fornecedor = args.get("fornecedor", "").strip()
    area_envolvida = args.get("area_envolvida", "").strip()
    apenas_atrasados = args.get("apenas_atrasados", "").strip()
    apenas_criticos = args.get("apenas_criticos", "").strip()

    query = ProjetoPD.query
    if busca:
        termo = f"%{busca}%"
        query = query.filter(
            or_(
                ProjetoPD.nome.ilike(termo),
                ProjetoPD.codigo.ilike(termo),
                ProjetoPD.cliente.ilike(termo),
                ProjetoPD.produto.ilike(termo),
                ProjetoPD.fornecedor.ilike(termo),
                ProjetoPD.descricao.ilike(termo),
            )
        )
    if etapa:
        query = query.filter(ProjetoPD.etapa_atual.in_(etapa))
    if categoria:
        query = query.filter(ProjetoPD.categoria.in_(categoria))
    if responsavel:
        query = query.filter(ProjetoPD.responsavel.ilike(f"%{responsavel}%"))
    if prioridade:
        query = query.filter(ProjetoPD.prioridade.in_(prioridade))
    if cliente:
        query = query.filter(ProjetoPD.cliente.ilike(f"%{cliente}%"))
    if produto:
        query = query.filter(ProjetoPD.produto.ilike(f"%{produto}%"))
    if fornecedor:
        query = query.filter(ProjetoPD.fornecedor.ilike(f"%{fornecedor}%"))
    if area_envolvida:
        query = query.filter(ProjetoPD.area_envolvida.ilike(f"%{area_envolvida}%"))
    if apenas_atrasados == "1":
        query = query.filter(
            ProjetoPD.data_prevista_conclusao.isnot(None),
            ProjetoPD.data_prevista_conclusao < date.today(),
            ProjetoPD.etapa_atual != "Concluído",
        )
    if apenas_criticos == "1":
        query = query.filter(
            ProjetoPD.prioridade == "ALTA",
            ProjetoPD.data_prevista_conclusao.isnot(None),
            ProjetoPD.data_prevista_conclusao < date.today(),
            ProjetoPD.etapa_atual != "Concluído",
        )

    query = query.order_by(ProjetoPD.data_prevista_conclusao.asc().nullslast(), ProjetoPD.id.desc())

    filtros = dict(
        busca=busca, etapa=etapa, categoria=categoria, responsavel=responsavel,
        prioridade=prioridade, cliente=cliente, produto=produto, fornecedor=fornecedor,
        area_envolvida=area_envolvida, apenas_atrasados=apenas_atrasados, apenas_criticos=apenas_criticos,
    )
    return query, filtros


def _pd_opcoes_filtro(campo, opcoes_curadas):
    """Mesmo espírito de _rnc_opcoes_filtro: combina valores já cadastrados
    (texto livre) com a lista de sugestão curada, pra nenhum valor real ficar
    de fora do filtro."""
    coluna = getattr(ProjetoPD, campo)
    existentes = [v for (v,) in db.session.query(coluna).filter(coluna.isnot(None)).distinct()]
    existentes_lower = {v.lower() for v in existentes}
    extras = [op for op in opcoes_curadas if op.lower() not in existentes_lower]
    return sorted(existentes + extras, key=lambda s: s.lower())


def _linhas_projetos_pd(args):
    page = args.get("page", 1, type=int)
    query, filtros = _filtrar_projetos_pd(args)
    total_filtrado = query.count()
    total_paginas = max(1, (total_filtrado + PAGE_SIZE - 1) // PAGE_SIZE)
    pagina = query.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    return pagina, page, total_paginas, total_filtrado, filtros


def _dashboard_pd():
    """Recalcula ao vivo os indicadores do Dashboard de P&D — nada fica
    pré-calculado/guardado, sempre em dia com o que estiver cadastrado
    (mesmo espírito de _dashboard_rnc_qualidade)."""
    projetos = ProjetoPD.query.all()
    total = len(projetos)
    ativos = [p for p in projetos if not p.concluido]
    atrasados = [p for p in projetos if p.atrasado]
    criticos = [p for p in projetos if p.critico]
    concluidos = [p for p in projetos if p.concluido]

    no_prazo = [p for p in ativos if not p.atrasado]
    pct_no_prazo = round((len(no_prazo) / len(ativos)) * 100, 1) if ativos else 0
    pct_atrasados = round((len(atrasados) / len(ativos)) * 100, 1) if ativos else 0

    duracoes = [
        (p.data_real_conclusao - p.data_inicio).days
        for p in concluidos if p.data_inicio and p.data_real_conclusao
    ]
    lead_time_medio = round(sum(duracoes) / len(duracoes), 1) if duracoes else 0

    todos_testes = TesteProjetoPD.query.all()
    testes_com_resultado = [t for t in todos_testes if t.resultado in ("Aprovado", "Reprovado")]
    testes_aprovados = [t for t in todos_testes if t.resultado == "Aprovado"]
    taxa_aprovacao_testes = round((len(testes_aprovados) / len(testes_com_resultado)) * 100, 1) if testes_com_resultado else 0

    projetos_homologados_ou_alem = [
        p for p in projetos if PD_ETAPA_OPCOES.index(p.etapa_atual) >= PD_ETAPA_OPCOES.index("Homologação")
    ]
    taxa_homologacao = round((len(projetos_homologados_ou_alem) / total) * 100, 1) if total else 0

    investimento_previsto = sum(p.investimento_previsto or 0 for p in projetos)
    investimento_realizado = sum(p.investimento_realizado or 0 for p in projetos)
    economia_prevista = sum(p.economia_prevista or 0 for p in projetos)
    economia_realizada = sum(p.economia_realizada or 0 for p in projetos)
    roi_geral = (
        round(((economia_realizada - investimento_realizado) / investimento_realizado) * 100, 1)
        if investimento_realizado else None
    )

    def _quebra_por(atributo, opcoes_ordem=None):
        contagem = {}
        for p in projetos:
            chave = getattr(p, atributo) or "—"
            contagem[chave] = contagem.get(chave, 0) + 1
        if opcoes_ordem:
            chaves = list(opcoes_ordem) + sorted(k for k in contagem if k not in opcoes_ordem and k != "—")
            if "—" in contagem:
                chaves.append("—")
            return [{"chave": k, "total": contagem.get(k, 0)} for k in chaves if k in contagem or k in opcoes_ordem]
        return sorted(({"chave": k, "total": v} for k, v in contagem.items()), key=lambda d: -d["total"])

    return {
        "total": total,
        "ativos": len(ativos),
        "em_desenvolvimento": sum(1 for p in projetos if p.etapa_atual == "Desenvolvimento"),
        "em_teste": sum(1 for p in projetos if p.etapa_atual == "Teste"),
        "em_validacao": sum(1 for p in projetos if p.etapa_atual == "Validação"),
        "em_homologacao": sum(1 for p in projetos if p.etapa_atual == "Homologação"),
        "concluidos": len(concluidos),
        "atrasados": len(atrasados),
        "criticos": len(criticos),
        "pct_no_prazo": pct_no_prazo,
        "pct_atrasados": pct_atrasados,
        "lead_time_medio": lead_time_medio,
        "taxa_aprovacao_testes": taxa_aprovacao_testes,
        "taxa_homologacao": taxa_homologacao,
        "investimento_previsto": investimento_previsto,
        "investimento_realizado": investimento_realizado,
        "economia_prevista": economia_prevista,
        "economia_realizada": economia_realizada,
        "roi_geral": roi_geral,
        "por_etapa": _quebra_por("etapa_atual", PD_ETAPA_OPCOES),
        "por_categoria": _quebra_por("categoria", PD_CATEGORIA_OPCOES),
        "por_responsavel": _quebra_por("responsavel"),
        "por_prioridade": _quebra_por("prioridade", PRIORIDADE_OPCOES),
        "projetos_atrasados": sorted(atrasados, key=lambda p: p.data_prevista_conclusao)[:10],
        "projetos_criticos": criticos[:10],
    }


def _cronograma_pd():
    """Monta as linhas do Cronograma Geral (Gantt simplificado, sem
    biblioteca externa — barras posicionadas por porcentagem dentro de uma
    faixa de datas comum). Só entram projetos com data de início preenchida;
    os demais são contados à parte (sem_data) pra não sumirem silenciosamente."""
    projetos = ProjetoPD.query.filter(ProjetoPD.data_inicio.isnot(None)).order_by(ProjetoPD.data_inicio.asc()).all()
    sem_data = ProjetoPD.query.filter(ProjetoPD.data_inicio.is_(None)).count()
    hoje = date.today()

    if not projetos:
        return {"linhas": [], "sem_data": sem_data, "marcadores_mes": [], "hoje_pct": None}

    def _fim_projeto(p):
        fim = p.data_real_conclusao or p.data_prevista_conclusao or p.data_inicio
        return fim if fim >= p.data_inicio else p.data_inicio

    inicio_min = min(p.data_inicio for p in projetos)
    fim_max = max([_fim_projeto(p) for p in projetos] + [hoje])
    total_dias = max((fim_max - inicio_min).days, 1)

    def _pct(d):
        return max(0.0, min(100.0, ((d - inicio_min).days / total_dias) * 100))

    linhas = []
    for p in projetos:
        fim = _fim_projeto(p)
        marcos = []
        for t in p.testes:
            data_evento = t.data_realizada or t.data_planejada
            if data_evento and inicio_min <= data_evento <= fim_max:
                info = PD_TESTE_RESULTADO_INFO.get(t.resultado, {})
                marcos.append({
                    "pos_pct": _pct(data_evento),
                    "label": f"Teste {t.numero}".strip() if t.numero else "Teste",
                    "cor": info.get("cor", "secondary"),
                    "emoji": info.get("emoji", ""),
                })
        linhas.append({
            "projeto": p,
            "offset_pct": _pct(p.data_inicio),
            "largura_pct": max(_pct(fim) - _pct(p.data_inicio), 0.6),
            "marcos": marcos,
        })

    marcadores_mes = []
    cursor = date(inicio_min.year, inicio_min.month, 1)
    while cursor <= fim_max:
        marcadores_mes.append({"pos_pct": _pct(cursor), "label": f"{MESES_PT[cursor.month - 1]}/{cursor.year}"})
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)

    return {
        "linhas": linhas, "sem_data": sem_data, "marcadores_mes": marcadores_mes,
        "hoje_pct": _pct(hoje), "inicio_min": inicio_min, "fim_max": fim_max,
    }


def _filtrar_testes_pd(args):
    """Visão global de Testes & Validações — mesmos testes já vistos dentro
    de cada projeto (aba Testes & Validações), aqui juntados de todos os
    projetos numa lista só, com filtro por projeto/resultado/responsável."""
    projeto_id = args.get("projeto_id", type=int)
    resultado = args.getlist("resultado")
    responsavel = args.get("responsavel", "").strip()
    apenas_atrasados = args.get("apenas_atrasados", "").strip()

    query = TesteProjetoPD.query.join(ProjetoPD, TesteProjetoPD.projeto_id == ProjetoPD.id)
    if projeto_id:
        query = query.filter(TesteProjetoPD.projeto_id == projeto_id)
    if resultado:
        query = query.filter(TesteProjetoPD.resultado.in_(resultado))
    if responsavel:
        query = query.filter(TesteProjetoPD.responsavel.ilike(f"%{responsavel}%"))
    if apenas_atrasados == "1":
        query = query.filter(
            TesteProjetoPD.data_planejada.isnot(None),
            TesteProjetoPD.data_planejada < date.today(),
            TesteProjetoPD.data_realizada.is_(None),
        )
    query = query.order_by(TesteProjetoPD.data_planejada.desc().nullslast(), TesteProjetoPD.id.desc())

    filtros = dict(projeto_id=projeto_id, resultado=resultado, responsavel=responsavel, apenas_atrasados=apenas_atrasados)
    return query, filtros


def _custos_pd():
    """Totais previsto x realizado / economia / ROI de todos os projetos —
    página própria de analytics financeiro de P&D (Bruno pediu "Custos &
    Resultados" separado do Dashboard geral)."""
    projetos = ProjetoPD.query.order_by(ProjetoPD.data_prevista_conclusao.asc().nullslast(), ProjetoPD.id.desc()).all()

    totais = {
        "custo_previsto": sum(p.custo_previsto or 0 for p in projetos),
        "custo_realizado": sum(p.custo_realizado or 0 for p in projetos),
        "investimento_previsto": sum(p.investimento_previsto or 0 for p in projetos),
        "investimento_realizado": sum(p.investimento_realizado or 0 for p in projetos),
        "economia_prevista": sum(p.economia_prevista or 0 for p in projetos),
        "economia_realizada": sum(p.economia_realizada or 0 for p in projetos),
    }
    totais["roi_geral"] = (
        round(((totais["economia_realizada"] - totais["investimento_realizado"]) / totais["investimento_realizado"]) * 100, 1)
        if totais["investimento_realizado"] else None
    )

    acima_do_previsto = [
        p for p in projetos
        if (p.custo_realizado is not None and p.custo_previsto is not None and p.custo_realizado > p.custo_previsto)
        or (p.investimento_realizado is not None and p.investimento_previsto is not None and p.investimento_realizado > p.investimento_previsto)
    ]
    return {"projetos": projetos, "totais": totais, "acima_do_previsto": acima_do_previsto}


# Limiares (em dias) usados só pelos alertas de "sem atualização"/"parado" —
# ajustáveis aqui sem mexer no resto da lógica, caso o Bruno peça outro valor.
PD_DIAS_SEM_ATUALIZACAO = 15
PD_DIAS_PARADO = 30


def _alertas_pd():
    """Consolida os alertas de P&D pedidos pelo Bruno (atrasado, prazo
    próximo, teste atrasado, projeto sem atualização, projeto parado, custo
    acima do previsto, homologação em andamento) — só considera projetos
    ainda não concluídos, igual ao resto do sistema já faz pra Pedido."""
    projetos = ProjetoPD.query.filter(ProjetoPD.etapa_atual != "Concluído").all()
    hoje = date.today()
    agora = datetime.utcnow()

    sem_atualizacao, parados = [], []
    for p in projetos:
        referencia = p.atualizado_em or p.criado_em
        if not referencia:
            continue
        dias = (agora - referencia).days
        if dias >= PD_DIAS_PARADO:
            parados.append(p)
        elif dias >= PD_DIAS_SEM_ATUALIZACAO:
            sem_atualizacao.append(p)

    custo_acima = [
        p for p in projetos
        if (p.custo_realizado is not None and p.custo_previsto is not None and p.custo_realizado > p.custo_previsto)
        or (p.investimento_realizado is not None and p.investimento_previsto is not None and p.investimento_realizado > p.investimento_previsto)
    ]

    testes_atrasados = (
        TesteProjetoPD.query.join(ProjetoPD, TesteProjetoPD.projeto_id == ProjetoPD.id)
        .filter(
            TesteProjetoPD.data_planejada.isnot(None),
            TesteProjetoPD.data_planejada < hoje,
            TesteProjetoPD.data_realizada.is_(None),
        )
        .all()
    )

    return {
        "atrasados": [p for p in projetos if p.atrasado],
        "prazo_proximo": [p for p in projetos if p.prazo_proximo],
        "homologacao": [p for p in projetos if p.etapa_atual == "Homologação"],
        "sem_atualizacao": sem_atualizacao,
        "parados": parados,
        "custo_acima": custo_acima,
        "testes_atrasados": testes_atrasados,
    }


# ----------------------------------------------------------------------
# Relatórios (fase 12) — exportações CSV/Excel sob demanda, sem agendamento
# automático (geradas na hora, a partir dos mesmos dados já calculados
# pelas telas de Listagem, Faturamento e Gargalos).
# ----------------------------------------------------------------------
def _linhas_export_listagem(pedidos):
    cabecalho = [
        "Pedido", "Cliente", "Cidade", "UF", "Vendedor", "Produto(s)",
        "Valor total (R$)", "Prioridade", "Estação", "Status",
        "Semáforo", "Dias (negativo = atrasado)", "Data de inclusão",
    ]
    linhas = []
    for p in pedidos:
        cor, dias = p.semaforo
        linhas.append([
            p.pedido_venda or "",
            p.cliente,
            p.cidade or "",
            p.estado or "",
            p.vendedor or "",
            p.descricao_resumo,
            round(p.valor_total, 2),
            p.prioridade or "",
            p.estacao_resumo,
            p.status_producao,
            SEMAFORO_LABELS.get(cor, cor),
            dias if dias is not None else "",
            p.data_inclusao_pedido.isoformat() if p.data_inclusao_pedido else "",
        ])
    return cabecalho, linhas


def _linhas_export_faturamento(itens):
    cabecalho = ["Pedido", "Cliente", "UF", "Vendedor", "Produto", "Valor faturado (R$)", "Nº nota fiscal", "Data de faturamento"]
    linhas = []
    for i in itens:
        p = i.pedido
        linhas.append([
            p.pedido_venda if p else "",
            p.cliente if p else "—",
            p.estado if p else "",
            p.vendedor if p else "",
            i.descricao_produto,
            round(i.valor_faturamento_realizado, 2),
            i.numero_nota_fiscal or "",
            i.liberacao_faturamento.isoformat() if i.liberacao_faturamento else "",
        ])
    return cabecalho, linhas


def _linhas_export_gargalos(linhas_gargalo):
    cabecalho = ["Estação", "Fila", "Atrasados", "Tempo de espera médio (dias)", "Lead time médio (dias)", "Valor parado (R$)"]
    linhas = [
        [
            g["estacao"],
            g["fila"],
            g["atraso"],
            g["tempo_espera_medio"] if g["tempo_espera_medio"] is not None else "",
            g["lt_medio"] if g["lt_medio"] is not None else "",
            g["valor_parado"],
        ]
        for g in linhas_gargalo
    ]
    return cabecalho, linhas


def _responder_csv(nome_arquivo, cabecalho, linhas):
    """Gera um CSV com ';' como separador (padrão do Excel em pt-BR) e um BOM
    UTF-8 no início, pra acentos abrirem certo direto no Excel."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(cabecalho)
    writer.writerows(linhas)
    conteudo = "﻿" + buffer.getvalue()
    resposta = Response(conteudo, mimetype="text/csv; charset=utf-8")
    resposta.headers["Content-Disposition"] = f"attachment; filename={nome_arquivo}"
    return resposta


def _responder_xlsx(nome_arquivo, cabecalho, linhas, titulo="Relatório"):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31] or "Relatório"
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = Font(bold=True)
    for linha in linhas:
        ws.append(linha)
    for coluna in ws.columns:
        valores = [len(str(c.value)) for c in coluna if c.value is not None]
        largura = max(valores) if valores else 10
        ws.column_dimensions[coluna[0].column_letter].width = min(largura + 2, 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resposta = Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resposta.headers["Content-Disposition"] = f"attachment; filename={nome_arquivo}"
    return resposta


def _construir_backup_pedidos_wb():
    """Monta um Workbook com TODAS as colunas de cada tabela apagada por
    "/admin/zerar-dados" (uma aba cada), lendo as colunas direto do
    mapeamento do SQLAlchemy (não uma lista escrita à mão) — assim não corre
    o risco de esquecer um campo novo que apareça no futuro. Usado como rede
    de segurança antes da exclusão permanente.

    Pedido do Bruno (03/09/2026): "zerar dados" passou a apagar também
    Qualidade (RNC + RDIM) e P&D, então o backup precisa cobrir as mesmas
    tabelas — senão a "rede de segurança" fica incompleta bem na hora que
    mais importa."""
    wb = Workbook()
    wb.remove(wb.active)

    def _add_sheet(nome, modelo, linhas):
        ws = wb.create_sheet(nome[:31])
        colunas = [c.name for c in modelo.__table__.columns]
        ws.append(colunas)
        for celula in ws[1]:
            celula.font = Font(bold=True)
        for obj in linhas:
            linha = []
            for nome_col in colunas:
                valor = getattr(obj, nome_col)
                if isinstance(valor, (datetime, date)):
                    valor = valor.isoformat()
                linha.append(valor)
            ws.append(linha)
        for coluna in ws.columns:
            valores = [len(str(c.value)) for c in coluna if c.value is not None]
            largura = max(valores) if valores else 10
            ws.column_dimensions[coluna[0].column_letter].width = min(largura + 2, 40)

    # ---- Produção + Operação (já existia) ----
    _add_sheet("Pedidos", Pedido, Pedido.query.order_by(Pedido.id).all())
    _add_sheet("Itens", ItemPedido, ItemPedido.query.order_by(ItemPedido.id).all())
    _add_sheet("Gestao Operacao", PedidoOperacao, PedidoOperacao.query.order_by(PedidoOperacao.id).all())
    _add_sheet("Programacao", Programacao, Programacao.query.order_by(Programacao.id).all())
    _add_sheet("Historico Alteracoes", HistoricoAlteracao, HistoricoAlteracao.query.order_by(HistoricoAlteracao.id).all())

    # ---- Qualidade: RNC + RDIM (novo, 03/09/2026) ----
    _add_sheet("RNC Qualidade", RncQualidade, RncQualidade.query.order_by(RncQualidade.id).all())
    _add_sheet("Inspecoes RDIM", InspecaoFinal, InspecaoFinal.query.order_by(InspecaoFinal.id).all())
    _add_sheet("RDIM Medicoes", RdimMedicao, RdimMedicao.query.order_by(RdimMedicao.id).all())
    _add_sheet("RDIM Pecas Desvio", RdimPecaDesvio, RdimPecaDesvio.query.order_by(RdimPecaDesvio.id).all())

    # ---- P&D (novo, 03/09/2026) ----
    _add_sheet("Projetos PD", ProjetoPD, ProjetoPD.query.order_by(ProjetoPD.id).all())
    _add_sheet("Testes PD", TesteProjetoPD, TesteProjetoPD.query.order_by(TesteProjetoPD.id).all())
    _add_sheet("Visitas Reunioes PD", VisitaReuniaoPD, VisitaReuniaoPD.query.order_by(VisitaReuniaoPD.id).all())

    return wb


def register_routes(app):
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            senha = request.form.get("senha", "")
            usuario = Usuario.query.filter_by(username=username).first()
            if usuario and usuario.checar_senha(senha):
                login_user(usuario)
                destino = request.args.get("next") or url_for("dashboard")
                return redirect(destino)
            flash("Usuário ou senha inválidos.", "danger")

        return render_template("login.html")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    @app.route("/consulta-pedido/busca-sugestoes")
    @login_required
    def consulta_pedido_busca_sugestoes():
        """Pedido do Bruno (01/09/2026, movido pra aba própria em 02/09/2026):
        sugestões dinâmicas (digitando) pro canal de busca de status da aba
        Consulta Pedido — devolve um fragmento HTML pronto (mesmo padrão do
        resto do sistema, que é 100% server-rendered) pra ser injetado direto
        na página via fetch()."""
        termo = request.args.get("q", "")
        sugestoes = _buscar_pedidos_para_status(termo)
        return render_template("_consulta_pedido_sugestoes.html", sugestoes=sugestoes, termo=termo.strip())

    @app.route("/consulta-pedido/busca-detalhe")
    @login_required
    def consulta_pedido_busca_detalhe():
        """Status completo de UM pedido (Produção + Operação, cruzados por
        pedido_venda) — devolve o fragmento HTML do painel de detalhe,
        aberto ao clicar numa sugestão/atalho da aba Consulta Pedido."""
        chave = (request.args.get("pedido_venda", "") or "").strip()
        if not chave:
            return "", 204

        pedido = (
            Pedido.query.options(selectinload(Pedido.itens))
            .filter(func.trim(Pedido.pedido_venda) == chave)
            .first()
        )
        go = PedidoOperacao.query.filter(func.trim(PedidoOperacao.pedido_venda) == chave).first()

        if pedido is None and go is None:
            return render_template("_consulta_pedido_detalhe.html", nao_encontrado=True, pedido_venda=chave)

        liberacao_pcp = _liberacao_pcp_por_pedido_venda([chave]).get(chave, {})
        data_cliente_producao = _data_cliente_por_pedido_venda([chave]).get(chave)
        situacao_entrega = _situacao_entrega_go(go, pedido)
        otd_pedido = _otd_do_pedido(go)
        etapas = _etapas_acompanhamento_pedido(pedido, go)

        # Qualidade (RDIM) — pedido do Bruno (02/09/2026): ver o processo de
        # qualidade item a item dentro do pedido, na mesma tela que já une
        # Produção + Operação.
        inspecoes_rdim = _inspecoes_rdim_por_item([i.id for i in pedido.itens]) if pedido else {}
        resumo_rdim = _resumo_rdim_pedido(inspecoes_rdim.values())

        return render_template(
            "_consulta_pedido_detalhe.html",
            pedido=pedido, go=go, pedido_venda=chave,
            liberacao_pcp=liberacao_pcp, data_cliente_producao=data_cliente_producao,
            situacao_entrega=situacao_entrega, otd_pedido=otd_pedido,
            etapas=etapas, nao_encontrado=False,
            inspecoes_rdim=inspecoes_rdim, resumo_rdim=resumo_rdim,
        )

    @app.route("/consulta-pedido")
    @login_required
    def consulta_pedido():
        """Aba própria "🔍 Consulta Pedido" (pedido do Bruno, 02/09/2026): o
        canal de busca de status que antes vivia dentro do Painel virou uma
        tela dedicada — só a busca por nº do pedido/cliente. Sem atalhos de
        Atrasados/Vencendo aqui de propósito (pedido do Bruno, 02/09/2026):
        esta área é de acesso comercial, sem esse tipo de informação interna."""
        pedido_venda_inicial = (request.args.get("pedido_venda", "") or "").strip()
        return render_template("consulta_pedido.html", pedido_venda_inicial=pedido_venda_inicial)

    @app.route("/painel")
    @login_required
    def painel():
        resumo = _calcular_resumo()
        atrasados = Pedido.query.filter(_predicado_atrasado()).count()
        vencendo = Pedido.query.filter(_predicado_vencendo()).count()
        backlog = resumo["total"] - resumo["finalizado"]

        hoje = date.today()
        previsto_mes, realizado_mes = _faturamento_mes(hoje.year, hoje.month)

        pedidos_atrasados = (
            Pedido.query.options(selectinload(Pedido.itens))
            .filter(_predicado_atrasado())
            .order_by(Pedido.data_inclusao_pedido.desc().nullslast())
            .limit(10)
            .all()
        )

        # Quadro mensal de projeção PCP (soma das semanas dentro de cada mês) —
        # filtro de período em <input type=month>, com um intervalo padrão de
        # 3 meses atrás até 6 meses à frente (dá pra ver "quanto foi entregue"
        # nos meses passados e "quanto já está projetado" nos meses seguintes).
        pcp_de_padrao = _somar_meses(hoje.year, hoje.month, -3)
        pcp_ate_padrao = _somar_meses(hoje.year, hoje.month, 6)
        pcp_de = _parse_mes_ano_form(request.args.get("pcp_de"), pcp_de_padrao)
        pcp_ate = _parse_mes_ano_form(request.args.get("pcp_ate"), pcp_ate_padrao)
        if pcp_de > pcp_ate:
            pcp_de, pcp_ate = pcp_ate, pcp_de

        # OTD/Lead Time da Operação, recorte "mês atual" — pedido do Bruno
        # (03/09/2026): "os principais KPIs da operação já sejam visuais logo
        # no painel", trazendo pro Painel os mesmos indicadores da tela
        # Resultados/OTD (mesmas funções, mesmo critério/meta).
        query_operacao_mes = _pedidos_operacao_do_periodo("mes", hoje.year, hoje.month)
        otd_operacao_mes = _resumo_otd(query_operacao_mes)
        lead_times_operacao_mes = _resumo_lead_times(query_operacao_mes)
        mes_atual_label = f"{MESES_PT[hoje.month - 1]}/{hoje.year}"

        # Gráficos anuais de faturamento realizado e OTD (pedido do Bruno,
        # 03/09/2026, no lugar do antigo "previsto × realizado (6 meses)") —
        # 1 filtro de ano simples, compartilhado pelos dois gráficos.
        anos_disponiveis_graficos = list(range(hoje.year - 2, hoje.year + 1))
        ano_graficos = request.args.get("ano_graficos", hoje.year, type=int)
        if ano_graficos not in anos_disponiveis_graficos:
            ano_graficos = hoje.year

        # Apontamentos recentes de Qualidade (RDIM + RNC) — pedido do Bruno
        # (03/09/2026): "notificação breve" de todo apontamento novo, pra ele
        # ter ciência sem precisar entrar na área de Qualidade. "Novo desde a
        # última visita" é guardado na sessão do navegador (sem tabela nova),
        # e é sempre atualizado DEPOIS de calcular a lista, pra este mesmo
        # carregamento ainda mostrar o que entrou desde a visita anterior.
        ultima_visita_str = session.get("ultima_visita_painel")
        ultima_visita = datetime.fromisoformat(ultima_visita_str) if ultima_visita_str else None
        apontamentos_qualidade = _apontamentos_recentes_qualidade(desde=ultima_visita)
        novos_apontamentos = sum(1 for a in apontamentos_qualidade if a["novo"])
        session["ultima_visita_painel"] = datetime.utcnow().isoformat()

        return render_template(
            "painel.html",
            resumo=resumo,
            atrasados=atrasados,
            vencendo=vencendo,
            backlog=backlog,
            previsto_mes=previsto_mes,
            realizado_mes=realizado_mes,
            lead_time_medio=_lead_time_medio_dias(),
            otd=_otd_percentual(),
            backlog_estacao=_backlog_por_estacao(),
            pedidos_atrasados=pedidos_atrasados,
            projecao_pcp=_projecao_pcp(),
            projecao_pcp_mensal=_projecao_pcp_mensal(pcp_de, pcp_ate),
            pcp_de_str=f"{pcp_de[0]:04d}-{pcp_de[1]:02d}",
            pcp_ate_str=f"{pcp_ate[0]:04d}-{pcp_ate[1]:02d}",
            otd_operacao_mes=otd_operacao_mes,
            lead_times_operacao_mes=lead_times_operacao_mes,
            mes_atual_label=mes_atual_label,
            faturamento_mensal_ano=_faturamento_mensal_ano(ano_graficos),
            otd_mensal_ano=_otd_mensal_ano(ano_graficos),
            ano_graficos=ano_graficos,
            anos_disponiveis_graficos=anos_disponiveis_graficos,
            apontamentos_qualidade=apontamentos_qualidade,
            novos_apontamentos=novos_apontamentos,
        )

    @app.route("/kpis")
    @login_required
    def kpis():
        meses = request.args.get("meses", 6, type=int)
        if meses not in (3, 6, 12):
            meses = 6
        desde = date.today() - timedelta(days=30 * meses)

        return render_template(
            "kpis.html",
            meses=meses,
            lead_time_estacao=_lead_time_por_estacao(desde=desde),
            otd_vendedor=_otd_por_vendedor(desde=desde),
            tendencia=_tendencia_kpis(meses=meses),
        )

    @app.route("/gargalos")
    @login_required
    def gargalos():
        return render_template("gargalos.html", linhas=_gargalos_por_estacao())

    @app.route("/faturamento")
    @login_required
    def faturamento():
        hoje = date.today()
        ano = request.args.get("ano", hoje.year, type=int)
        mes = request.args.get("mes", hoje.month, type=int)
        if not (1 <= mes <= 12):
            mes = hoje.month

        cliente = request.args.get("cliente", "").strip()
        regiao = request.args.get("regiao", "").strip()
        vendedor = request.args.get("vendedor", "").strip()

        dados = _faturamento_detalhado(ano, mes, cliente=cliente or None, regiao=regiao or None, vendedor=vendedor or None)

        mes_anterior_ano, mes_anterior_mes = (ano, mes - 1) if mes > 1 else (ano - 1, 12)
        mes_seguinte_ano, mes_seguinte_mes = (ano, mes + 1) if mes < 12 else (ano + 1, 1)

        return render_template(
            "faturamento.html",
            ano=ano,
            mes=mes,
            mes_label=f"{MESES_PT[mes - 1]}/{ano}",
            mes_anterior=dict(ano=mes_anterior_ano, mes=mes_anterior_mes),
            mes_seguinte=dict(ano=mes_seguinte_ano, mes=mes_seguinte_mes),
            filtros=dict(cliente=cliente, regiao=regiao, vendedor=vendedor),
            **dados,
        )

    @app.route("/logistica")
    @login_required
    def logistica():
        query = ItemPedido.query.join(Pedido).options(selectinload(ItemPedido.pedido))

        regiao = request.args.get("regiao", "").strip()
        estado = request.args.get("estado", "").strip()
        cidade = request.args.get("cidade", "").strip()
        frete = request.args.get("frete", "").strip()
        transportadora_id = request.args.get("transportadora_id", "").strip()
        em_risco = request.args.get("em_risco", "").strip()
        mostrar_finalizados = request.args.get("mostrar_finalizados", "").strip()

        if regiao:
            ufs_da_regiao = [uf for uf, r in REGIAO_POR_UF.items() if r == regiao]
            if ufs_da_regiao:
                query = query.filter(Pedido.estado.in_(ufs_da_regiao))
        if estado:
            query = query.filter(Pedido.estado == estado)
        if cidade:
            query = query.filter(Pedido.cidade.ilike(f"%{cidade}%"))
        if frete:
            query = query.filter(Pedido.frete == frete)
        if transportadora_id.isdigit():
            query = query.filter(ItemPedido.transportadora_id == int(transportadora_id))
        if not mostrar_finalizados:
            query = query.filter(ItemPedido.status_producao != "FINALIZADO")

        query = query.order_by(Pedido.data_inclusao_pedido.desc().nullslast(), ItemPedido.id.desc())
        itens = query.all()

        # semáforo é calculado em Python (depende de "hoje"), então o filtro
        # "em risco" é aplicado depois da consulta — mas como os filtros acima
        # já reduzem bastante a lista (nunca é a tabela inteira), não há o
        # mesmo problema de performance que o dashboard antigo tinha.
        if em_risco:
            itens = [i for i in itens if i.semaforo[0] in ("amarelo", "vermelho")]

        transportadoras = Transportadora.query.filter_by(ativo=True).order_by(Transportadora.nome).all()
        estados_disponiveis = [
            r[0]
            for r in db.session.query(Pedido.estado)
            .filter(Pedido.estado.isnot(None), Pedido.estado != "")
            .distinct()
            .order_by(Pedido.estado)
            .all()
        ]

        return render_template(
            "logistica.html",
            itens=itens,
            transportadoras=transportadoras,
            estados_disponiveis=estados_disponiveis,
            filtros=dict(
                regiao=regiao,
                estado=estado,
                cidade=cidade,
                frete=frete,
                transportadora_id=transportadora_id,
                em_risco=em_risco,
                mostrar_finalizados=mostrar_finalizados,
            ),
        )

    @app.route("/cadastros/transportadoras")
    @requer_role("ADMIN", "PCP")
    def cadastros_transportadoras():
        transportadoras = Transportadora.query.order_by(Transportadora.nome).all()
        return render_template("cadastros_transportadoras.html", transportadoras=transportadoras)

    @app.route("/cadastros/transportadoras/novo", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP")
    def cadastros_transportadoras_novo():
        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip()
            if not nome:
                flash("Informe o nome da transportadora.", "danger")
                return render_template("cadastros_transportadoras_form.html", transportadora=None, form=f)
            if Transportadora.query.filter_by(nome=nome).first():
                flash("Já existe uma transportadora com esse nome.", "danger")
                return render_template("cadastros_transportadoras_form.html", transportadora=None, form=f)

            nova = Transportadora(nome=nome, ativo=True)
            db.session.add(nova)
            db.session.commit()
            flash(f"Transportadora {nome} cadastrada com sucesso.", "success")
            return redirect(url_for("cadastros_transportadoras"))

        return render_template("cadastros_transportadoras_form.html", transportadora=None, form={})

    @app.route("/cadastros/transportadoras/<int:transportadora_id>/editar", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP")
    def cadastros_transportadoras_editar(transportadora_id):
        transportadora = db.session.get(Transportadora, transportadora_id)
        if transportadora is None:
            flash("Transportadora não encontrada.", "danger")
            return redirect(url_for("cadastros_transportadoras"))

        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip()
            if not nome:
                flash("Informe o nome da transportadora.", "danger")
                return render_template("cadastros_transportadoras_form.html", transportadora=transportadora, form=f)
            outra = Transportadora.query.filter(Transportadora.nome == nome, Transportadora.id != transportadora.id).first()
            if outra:
                flash("Já existe outra transportadora com esse nome.", "danger")
                return render_template("cadastros_transportadoras_form.html", transportadora=transportadora, form=f)

            transportadora.nome = nome
            transportadora.ativo = bool(f.get("ativo"))
            db.session.commit()
            flash(f"Transportadora {transportadora.nome} atualizada com sucesso.", "success")
            return redirect(url_for("cadastros_transportadoras"))

        return render_template("cadastros_transportadoras_form.html", transportadora=transportadora, form={})

    @app.route("/alertas")
    @login_required
    def alertas():
        pedidos_atrasados = (
            Pedido.query.options(selectinload(Pedido.itens))
            .filter(_predicado_atrasado())
            .order_by(Pedido.data_inclusao_pedido.desc().nullslast())
            .all()
        )
        pedidos_vencendo = (
            Pedido.query.options(selectinload(Pedido.itens))
            .filter(_predicado_vencendo())
            .order_by(Pedido.data_inclusao_pedido.desc().nullslast())
            .all()
        )
        gargalos_criticos = [g for g in _gargalos_por_estacao() if g["fila"] > 0 or g["atraso"] > 0][:5]
        faturamento_pendente = _faturamento_previsto_nao_realizado()
        pd_alertas = _alertas_pd()

        return render_template(
            "alertas.html",
            pedidos_atrasados=pedidos_atrasados,
            pedidos_vencendo=pedidos_vencendo,
            gargalos_criticos=gargalos_criticos,
            faturamento_pendente=faturamento_pendente,
            pd_alertas=pd_alertas,
            PD_DIAS_SEM_ATUALIZACAO=PD_DIAS_SEM_ATUALIZACAO,
            PD_DIAS_PARADO=PD_DIAS_PARADO,
        )

    # ------------------------------------------------------------------
    # Relatórios (fase 12) — exportações sob demanda, sem agendamento
    # automático. Cada relatório tem uma versão .csv e uma .xlsx.
    # ------------------------------------------------------------------
    @app.route("/relatorios")
    @login_required
    def relatorios():
        hoje = date.today()
        return render_template("relatorios.html", hoje=hoje)

    @app.route("/relatorios/listagem.csv")
    @login_required
    def relatorio_listagem_csv():
        query, _ = _filtrar_pedidos(request.args)
        cabecalho, linhas = _linhas_export_listagem(query.all())
        return _responder_csv("listagem_pedidos.csv", cabecalho, linhas)

    @app.route("/relatorios/listagem.xlsx")
    @login_required
    def relatorio_listagem_xlsx():
        query, _ = _filtrar_pedidos(request.args)
        cabecalho, linhas = _linhas_export_listagem(query.all())
        return _responder_xlsx("listagem_pedidos.xlsx", cabecalho, linhas, titulo="Listagem")

    @app.route("/relatorios/faturamento.csv")
    @login_required
    def relatorio_faturamento_csv():
        hoje = date.today()
        ano = request.args.get("ano", hoje.year, type=int)
        mes = request.args.get("mes", hoje.month, type=int)
        if not (1 <= mes <= 12):
            mes = hoje.month
        cliente = request.args.get("cliente", "").strip()
        regiao = request.args.get("regiao", "").strip()
        vendedor = request.args.get("vendedor", "").strip()
        dados = _faturamento_detalhado(ano, mes, cliente=cliente or None, regiao=regiao or None, vendedor=vendedor or None)
        cabecalho, linhas = _linhas_export_faturamento(dados["itens_realizados_lista"])
        return _responder_csv(f"faturamento_{ano}_{mes:02d}.csv", cabecalho, linhas)

    @app.route("/relatorios/faturamento.xlsx")
    @login_required
    def relatorio_faturamento_xlsx():
        hoje = date.today()
        ano = request.args.get("ano", hoje.year, type=int)
        mes = request.args.get("mes", hoje.month, type=int)
        if not (1 <= mes <= 12):
            mes = hoje.month
        cliente = request.args.get("cliente", "").strip()
        regiao = request.args.get("regiao", "").strip()
        vendedor = request.args.get("vendedor", "").strip()
        dados = _faturamento_detalhado(ano, mes, cliente=cliente or None, regiao=regiao or None, vendedor=vendedor or None)
        cabecalho, linhas = _linhas_export_faturamento(dados["itens_realizados_lista"])
        return _responder_xlsx(f"faturamento_{ano}_{mes:02d}.xlsx", cabecalho, linhas, titulo="Faturamento")

    @app.route("/relatorios/gargalos.csv")
    @login_required
    def relatorio_gargalos_csv():
        cabecalho, linhas = _linhas_export_gargalos(_gargalos_por_estacao())
        return _responder_csv("gargalos.csv", cabecalho, linhas)

    @app.route("/relatorios/gargalos.xlsx")
    @login_required
    def relatorio_gargalos_xlsx():
        cabecalho, linhas = _linhas_export_gargalos(_gargalos_por_estacao())
        return _responder_xlsx("gargalos.xlsx", cabecalho, linhas, titulo="Gargalos")

    @app.route("/")
    @login_required
    def dashboard():
        page = request.args.get("page", 1, type=int)
        sort = request.args.get("sort", SORT_PADRAO)
        dir_ordenacao = request.args.get("dir", DIR_PADRAO)
        if sort not in SORT_KEYS_LISTAGEM_GERAL:
            sort = SORT_PADRAO
        if dir_ordenacao not in ("asc", "desc"):
            dir_ordenacao = DIR_PADRAO

        # A Listagem Geral mostra 1 linha por PRODUTO (item), não por pedido —
        # então a ordenação/paginação não dá mais pra fazer em SQL direto (várias
        # colunas, como prioridade/status/prazo, têm ordem própria calculada em
        # Python). Busca tudo que passou pelo filtro, achata em linhas por item,
        # ordena e pagina em Python — mesmo padrão já usado em Gestão Operação e
        # no Painel pra esse tipo de coluna.
        query, filtros = _filtrar_pedidos(request.args)
        pedidos = query.all()
        linhas = _linhas_listagem_geral(pedidos, request.args)
        linhas = _ordenar_com_nulos_no_fim(linhas, SORT_KEYS_LISTAGEM_GERAL[sort], reverse=(dir_ordenacao == "desc"))

        # A pedido de Bruno: a Listagem Geral mostra todos os itens filtrados
        # numa página só, sem paginação (nada de "página 2, 3, 4...").
        total_filtrado = len(linhas)
        total_paginas = 1
        linhas_pagina = linhas

        # Pedido do Bruno (03/09/2026): os cards do topo (Total/Pendentes/Em
        # tratativa/Em andamento/Finalizados/Valor total) agora recalculam em
        # cima do resultado já filtrado, em vez do banco inteiro sempre —
        # ver _calcular_resumo_filtrado.
        resumo = _calcular_resumo_filtrado(linhas)

        filtros_paginacao = dict(filtros, sort=sort, dir=dir_ordenacao)

        # Link de cada card de status: mantém os OUTROS filtros ativos
        # (mês, região, estação, busca...) e só troca o status — clicar em
        # "Pendentes" com um filtro de região já aplicado continua só
        # naquela região, em vez de resetar tudo (mesmo pedido do Bruno, já
        # que os números do card agora refletem esse recorte).
        filtros_status_cards = {
            "total": dict(filtros, status=""),
            "pendente": dict(filtros, status="PENDENTE"),
            "em_tratativa": dict(filtros, status="EM TRATATIVA"),
            "andamento": dict(filtros, status="ANDAMENTO"),
            "finalizado": dict(filtros, status="FINALIZADO"),
        }

        # Qualidade (RDIM) — pedido do Bruno (02/09/2026): coluna de status
        # de qualidade por item, direto na Listagem Geral.
        inspecoes_rdim = _inspecoes_rdim_por_item([l.item_id for l in linhas_pagina])

        return render_template(
            "dashboard.html",
            linhas=linhas_pagina,
            resumo=resumo,
            page=page,
            total_paginas=total_paginas,
            total_filtrado=total_filtrado,
            filtros=filtros,
            filtros_paginacao=filtros_paginacao,
            filtros_status_cards=filtros_status_cards,
            sort=sort,
            dir_ordenacao=dir_ordenacao,
            inspecoes_rdim=inspecoes_rdim,
        )

    @app.route("/pedidos/novo", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP")
    def novo_pedido():
        if request.method == "POST":
            f = request.form
            descricoes = f.getlist("item_descricao[]")
            quantidades = f.getlist("item_quantidade[]")
            custos = f.getlist("item_custo[]")
            estacoes = f.getlist("item_estacao[]")

            itens = []
            itens_recarregados = []
            for desc, qtd, custo, estacao_item in zip(descricoes, quantidades, custos, estacoes):
                itens_recarregados.append({"descricao": desc, "quantidade": qtd, "custo": custo, "estacao": estacao_item})
                desc = desc.strip()
                if not desc:
                    continue
                itens.append(
                    ItemPedido(
                        descricao_produto=desc,
                        quantidade=_parse_float_form(qtd),
                        custo_unitario=_parse_float_form(custo),
                        estacao=estacao_item or None,
                    )
                )

            pedido = Pedido(
                data_cliente=_parse_data_form(f.get("data_cliente")),
                data_inclusao_pedido=_parse_data_form(f.get("data_inclusao_pedido")),
                cliente=f.get("cliente", "").strip(),
                cnpj=f.get("cnpj", "").strip() or None,
                cidade=f.get("cidade", "").strip() or None,
                estado=f.get("estado", "").strip() or None,
                pais=f.get("pais", "").strip() or "Brasil",
                frete=f.get("frete") or None,
                vendedor=f.get("vendedor", "").strip() or None,
                pedido_venda=f.get("pedido_venda", "").strip() or None,
                prioridade=f.get("prioridade") or "MÉDIA",
            )
            pedido.itens = itens

            if not pedido.cliente or not itens:
                flash("Cliente e ao menos um produto (com descrição) são obrigatórios.", "danger")
                return render_template("novo_pedido.html", form=f, itens=itens_recarregados)

            for item in itens:
                item.atualizar_status_automatico()
            db.session.add(pedido)
            # Alimenta automaticamente a Listagem Geral de Gestão Operação
            # (pedido do Bruno, 01/09/2026) — 1 PedidoOperacao criado junto,
            # cópia inicial dos dados; dali em diante cada um é editado
            # independente na sua própria tela (ver
            # _criar_pedido_operacao_a_partir_de_producao).
            _criar_pedido_operacao_a_partir_de_producao(pedido, f)
            db.session.commit()
            flash(
                f"Pedido de {pedido.cliente} incluído com sucesso ({len(itens)} item(ns)) "
                "— também já apareceu na Listagem Geral de Gestão Operação.",
                "success",
            )
            return redirect(url_for("editar_pedido", pedido_id=pedido.id))

        return render_template("novo_pedido.html", form={}, itens=[])

    @app.route("/pedidos/<int:pedido_id>/editar", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP", "LIDER")
    def editar_pedido(pedido_id):
        pedido = db.session.get(Pedido, pedido_id)
        if pedido is None:
            flash("Pedido não encontrado.", "danger")
            return redirect(url_for("dashboard"))

        transportadoras = Transportadora.query.filter_by(ativo=True).order_by(Transportadora.nome).all()

        if request.method == "POST":
            f = request.form

            pedido_antes = {c: getattr(pedido, c) for c in CAMPOS_HISTORICO_PEDIDO}

            pedido.data_cliente = _parse_data_form(f.get("data_cliente"))
            pedido.data_inclusao_pedido = _parse_data_form(f.get("data_inclusao_pedido"))
            pedido.cliente = f.get("cliente", "").strip() or pedido.cliente
            pedido.cnpj = f.get("cnpj", "").strip() or None
            pedido.cidade = f.get("cidade", "").strip() or None
            pedido.estado = f.get("estado", "").strip() or None
            pedido.pais = f.get("pais", "").strip() or "Brasil"
            pedido.frete = f.get("frete") or None
            pedido.vendedor = f.get("vendedor", "").strip() or None
            pedido.pedido_venda = f.get("pedido_venda", "").strip() or None

            pedido.prioridade = f.get("prioridade") or pedido.prioridade
            pedido.obs = f.get("obs", "").strip() or None

            pedido_depois = {c: getattr(pedido, c) for c in CAMPOS_HISTORICO_PEDIDO}
            _registrar_alteracoes("pedido", pedido.id, pedido.id, pedido_antes, pedido_depois, CAMPOS_HISTORICO_PEDIDO)

            # Planejamento semanal / liberação prevista / liberação real (pedido do
            # Bruno, 31/08/2026): deixaram de ser editáveis por item — agora só
            # existe um campo de cada no topo do pedido, e o valor é replicado pra
            # TODOS os itens ao salvar (por isso lidos uma vez só, fora do loop de
            # itens, em vez de via getlist("item_...[]") como os campos por item).
            lib_prevista_pedido = _parse_data_form(f.get("liberacao_prevista"))
            lib_real_pedido = _parse_data_form(f.get("liberacao_real"))
            planejamento_semanal_pedido = f.get("planejamento_semanal", "").strip() or None

            # ---- itens do pedido (vários produtos, cada um com sua própria estação/status/produção) ----
            item_ids = f.getlist("item_id[]")
            descricoes = f.getlist("item_descricao[]")
            quantidades = f.getlist("item_quantidade[]")
            custos = f.getlist("item_custo[]")
            estacoes = f.getlist("item_estacao[]")
            status_itens = f.getlist("item_status[]")
            inicios_producao = f.getlist("item_inicio_producao[]")
            # Pedido do Bruno (01/09/2026): a tela de edição não mostra mais
            # início/término de inspeção-embalagem e liberação de faturamento
            # como campos separados — só "Conclusão produção", um único campo
            # que grava nos dois (termino_inspecao E liberacao_faturamento),
            # exatamente como o botão "Avançar" do Kanban já fazia (ver
            # estacao_kanban_mover) — os dois sempre andam juntos na prática,
            # então isso não muda em nada o cálculo de FINALIZADO nem os
            # relatórios de lead time/gargalos, que continuam lendo
            # termino_inspecao normalmente.
            conclusoes_producao = f.getlist("item_conclusao_producao[]")

            itens_originais = {item.id: item for item in pedido.itens}
            # snapshot ANTES de qualquer alteração, só dos itens que já existiam
            # (itens novos não têm "antes" pra comparar — são uma inclusão, não uma mudança)
            itens_antes = {iid: {c: getattr(item, c) for c in CAMPOS_HISTORICO_ITEM} for iid, item in itens_originais.items()}
            ids_mantidos = set()

            linhas = zip(
                item_ids, descricoes, quantidades, custos, estacoes, status_itens,
                inicios_producao, conclusoes_producao,
            )
            for (item_id, desc, qtd, custo, estacao_item, status_item,
                 ini_prod, concl_prod) in linhas:
                desc = desc.strip()

                if item_id:
                    iid = int(item_id)
                    item = itens_originais.get(iid)
                    if item is None:
                        continue
                    if not desc:
                        continue  # descrição apagada -> item será removido abaixo
                else:
                    if not desc:
                        continue
                    item = ItemPedido()
                    pedido.itens.append(item)

                item.descricao_produto = desc
                item.quantidade = _parse_float_form(qtd)
                item.custo_unitario = _parse_float_form(custo)
                item.estacao = estacao_item or None
                item.inicio_producao = _parse_data_form(ini_prod)
                data_conclusao = _parse_data_form(concl_prod)
                item.termino_inspecao = data_conclusao
                item.liberacao_faturamento = data_conclusao
                item.liberacao_prevista = lib_prevista_pedido
                item.liberacao_real = lib_real_pedido
                item.planejamento_semanal = planejamento_semanal_pedido

                if status_item == "EM TRATATIVA":
                    item.status_manual = True
                    item.status_producao = "EM TRATATIVA"
                else:
                    item.status_manual = False
                item.atualizar_status_automatico()

                if item_id:
                    ids_mantidos.add(int(item_id))
                    item_depois = {c: getattr(item, c) for c in CAMPOS_HISTORICO_ITEM}
                    _registrar_alteracoes(
                        "item_pedido", int(item_id), pedido.id, itens_antes[int(item_id)], item_depois, CAMPOS_HISTORICO_ITEM
                    )

            for iid, item in itens_originais.items():
                if iid not in ids_mantidos:
                    pedido.itens.remove(item)

            if not pedido.itens:
                flash("O pedido precisa ter ao menos um produto.", "danger")
                return render_template("editar_pedido.html", pedido=pedido, transportadoras=transportadoras)

            db.session.commit()
            flash("Pedido atualizado com sucesso.", "success")
            return redirect(url_for("editar_pedido", pedido_id=pedido.id))

        return render_template("editar_pedido.html", pedido=pedido, transportadoras=transportadoras)

    @app.route("/pedidos/<int:pedido_id>")
    @login_required
    def detalhe_pedido(pedido_id):
        pedido = db.session.get(Pedido, pedido_id)
        if pedido is None:
            flash("Pedido não encontrado.", "danger")
            return redirect(url_for("dashboard"))
        historico = (
            HistoricoAlteracao.query.filter_by(pedido_id=pedido.id)
            .order_by(HistoricoAlteracao.criado_em.desc())
            .all()
        )
        timeline = _construir_timeline(pedido)
        inspecoes_rdim = _inspecoes_rdim_por_item([i.id for i in pedido.itens])
        resumo_rdim = _resumo_rdim_pedido(inspecoes_rdim.values())
        return render_template(
            "detalhe_pedido.html", pedido=pedido, historico=historico, timeline=timeline,
            inspecoes_rdim=inspecoes_rdim, resumo_rdim=resumo_rdim,
        )

    @app.route("/pedidos/<int:pedido_id>/excluir", methods=["POST"])
    @requer_role("ADMIN", "PCP")
    def excluir_pedido(pedido_id):
        pedido = db.session.get(Pedido, pedido_id)
        if pedido is not None:
            db.session.delete(pedido)
            try:
                db.session.commit()
            except IntegrityError:
                # Defesa extra (bug real corrigido 03/09/2026, pedido do
                # Bruno "não consigo apagar nenhum pedido"): faltavam
                # cascades de exclusão pra Programacao/InspecaoFinal/
                # HistoricoAlteracao, o que fazia qualquer pedido já editado
                # ou inspecionado estourar um 500 sem aviso nenhum. Os
                # relacionamentos já foram corrigidos com cascade="all,
                # delete-orphan" em models.py — isto aqui é só uma rede de
                # segurança pra nunca mais devolver um 500 cru se algum
                # cadastro novo no futuro esquecer o mesmo cuidado.
                db.session.rollback()
                flash(
                    "Não foi possível excluir este pedido: ainda há registros vinculados a ele "
                    "(histórico, inspeção ou programação) que não puderam ser removidos. Avise o suporte.",
                    "danger",
                )
                return redirect(url_for("dashboard"))
            flash("Pedido excluído.", "info")
        return redirect(url_for("dashboard"))

    @app.route("/api/pedidos/<int:pedido_id>/resumo")
    @login_required
    def api_resumo_pedido(pedido_id):
        pedido = db.session.get(Pedido, pedido_id)
        if pedido is None:
            return jsonify({"erro": "não encontrado"}), 404
        return jsonify(
            {
                "valor_total": pedido.valor_total,
                "lt_comercial_dias": pedido.lt_comercial_dias,
                "prazo_total_dias": pedido.prazo_total_dias,
                "status_producao": pedido.status_producao,
                "itens": [
                    {
                        "descricao_produto": item.descricao_produto,
                        "estacao": item.estacao,
                        "status_producao": item.status_producao,
                        "tempo_espera_dias": item.tempo_espera_dias,
                        "lt_producao_dias": item.lt_producao_dias,
                        "prazo_total_dias": item.prazo_total_dias,
                        "valor_total": item.valor_total,
                    }
                    for item in pedido.itens
                ],
            }
        )

    # ------------------------------------------------------------------
    # Gestão Operação (Fase 13) — sub-abas coloridas (PCP/Logística/
    # Resultados), uma linha por PEDIDO, reaproveitando os mesmos filtros e
    # paginação da Listagem Geral (_filtrar_pedidos) — só muda o conjunto de
    # colunas mostrado em cada template. A sub-aba "Comercial" (lista) foi
    # removida (pedido do Bruno, 01/09/2026): a Listagem Geral já mostra as
    # mesmas informações comerciais, então a lista separada virou
    # redundante — o FORMULÁRIO de edição da seção Comercial continua
    # existindo normalmente em gestao_operacao_editar (ver GO_SECAO_ENDPOINT).
    # ------------------------------------------------------------------
    @app.route("/gestao-operacao/pcp")
    @login_required
    def gestao_operacao_pcp():
        pedidos, page, total_paginas, total_filtrado, filtros, _query_operacao = _linhas_gestao_operacao(request.args)
        status_real_por_pedido_venda = _status_producao_por_pedido_venda([p.pedido_venda for p in pedidos])
        liberacao_pcp_por_pedido_venda = _liberacao_pcp_por_pedido_venda([p.pedido_venda for p in pedidos])
        # "Solicitada cliente/retira" também acompanha ao vivo a "Data do
        # cliente" de Gestão Produção — pedido do Bruno (03/09/2026).
        data_cliente_por_pedido_venda = _data_cliente_por_pedido_venda([p.pedido_venda for p in pedidos])
        return render_template(
            "gestao_operacao_pcp.html",
            pedidos=pedidos, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros,
            status_real_por_pedido_venda=status_real_por_pedido_venda,
            liberacao_pcp_por_pedido_venda=liberacao_pcp_por_pedido_venda,
            data_cliente_por_pedido_venda=data_cliente_por_pedido_venda,
        )

    @app.route("/gestao-operacao/logistica")
    @login_required
    def gestao_operacao_logistica():
        pedidos, page, total_paginas, total_filtrado, filtros, _query_operacao = _linhas_gestao_operacao(request.args)
        return render_template(
            "gestao_operacao_logistica.html",
            pedidos=pedidos, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros,
        )

    @app.route("/gestao-operacao/resultados")
    @login_required
    def gestao_operacao_resultados():
        # Período (pedido do Bruno, 28/08/2026: mês selecionável; ampliado
        # 03/09/2026 — "inclua em formato de lista... além do mês, inclua
        # também trimestre, semestre... e o filtro geral" — ver
        # _parse_periodo/_opcoes_periodo). Sem `periodo` na URL, cai no mês
        # atual (mesmo default de sempre). Calculado ANTES de
        # _linhas_gestao_operacao pra poder usar `periodo_str` (já com o
        # default aplicado) no campo oculto do formulário de segmento (ver
        # abaixo) — assim o filtro de segmento sempre navega com um período
        # explícito, mesmo na primeira visita à página.
        tipo_periodo, ano_periodo, valor_periodo, periodo_label = _parse_periodo(request.args.get("periodo", ""))
        periodo_str = _periodo_para_str(tipo_periodo, ano_periodo, valor_periodo)
        faturamento_semanal = _faturamento_por_periodo(tipo_periodo, ano_periodo, valor_periodo)
        periodo_anterior = _periodo_vizinho(tipo_periodo, ano_periodo, valor_periodo, -1)
        periodo_seguinte = _periodo_vizinho(tipo_periodo, ano_periodo, valor_periodo, 1)

        pedidos, page, total_paginas, total_filtrado, filtros, query_operacao = _linhas_gestao_operacao(request.args)
        # Pedido do Bruno (03/09/2026): "quero ver todos os pedidos de julho
        # faturados ou dentro do planejamento semanal do pcp, com isso ver o
        # otd" — filtros["segmento"] (junto com o período) já veio aplicado
        # em query_operacao (ver _filtrar_pedidos_operacao); o resumo de OTD
        # abaixo usa o MESMO recorte, em vez de sempre olhar pra todos os
        # pedidos.
        otd = _resumo_otd(query_operacao)

        # Resumo fixo do período (pedido do Bruno, 03/09/2026: "preciso ver
        # os resultados detalhados de cada mês, como otd, lead time
        # operação, lead time chão de fábrica (produção), lead time operação
        # cif, lead time operação fob... de preferência no topo da página,
        # do lado do faturamento") — SEMPRE o período selecionado em cima,
        # no mesmo recorte "planejamento" (Término Semanal PCP no período)
        # usado em "Qtd/Valor liberado", independente do dropdown de
        # segmento mais abaixo (que continua controlando só os cards de OTD
        # "geral" e a lista de pedidos).
        query_periodo = _pedidos_operacao_do_periodo(tipo_periodo, ano_periodo, valor_periodo)
        otd_mes = _resumo_otd(query_periodo)
        lead_times_mes = _resumo_lead_times(query_periodo)

        return render_template(
            "gestao_operacao_resultados.html",
            pedidos=pedidos, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros, otd=otd,
            faturamento_semanal=faturamento_semanal,
            otd_mes=otd_mes, lead_times_mes=lead_times_mes,
            periodo=periodo_str, tipo_periodo=tipo_periodo, mes_label=periodo_label,
            periodo_anterior=periodo_anterior, periodo_seguinte=periodo_seguinte,
            opcoes_periodo=_opcoes_periodo(),
        )

    @app.route("/gestao-operacao/listagem-geral")
    @login_required
    def gestao_operacao_listagem_geral():
        """"Listagem Geral" de Gestão Operação (pedido do Bruno) — 1 linha por
        PEDIDO (não por produto, diferente da Listagem Geral de Gestão
        Produção), com as colunas comerciais principais; passar o mouse (ou
        clicar, no touch) sobre "Itens" mostra os produtos/quantidades já
        preenchidos em Gestão Produção pelo PCP, casando pelo nº de pedido de
        venda — sem criar nenhum vínculo real entre as duas tabelas."""
        pedidos, page, total_paginas, total_filtrado, filtros, _query_operacao = _linhas_gestao_operacao(request.args)
        itens_por_pedido_venda = _itens_producao_por_pedido_venda([p.pedido_venda for p in pedidos])
        # Qualidade (RDIM) — pedido do Bruno (02/09/2026): indicador simples
        # (pedido-level, sem granularidade de item/estação) de "contém
        # desvio" nesta listagem.
        rdim_por_pedido_venda = _rdim_resumo_por_pedido_venda([p.pedido_venda for p in pedidos])
        # "Data solic. cliente" também acompanha ao vivo a "Data do cliente"
        # de Gestão Produção — pedido do Bruno (03/09/2026).
        data_cliente_por_pedido_venda = _data_cliente_por_pedido_venda([p.pedido_venda for p in pedidos])
        return render_template(
            "gestao_operacao_listagem_geral.html",
            pedidos=pedidos, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros,
            itens_por_pedido_venda=itens_por_pedido_venda,
            rdim_por_pedido_venda=rdim_por_pedido_venda,
            data_cliente_por_pedido_venda=data_cliente_por_pedido_venda,
        )

    @app.route("/gestao-operacao/<int:pedido_id>/editar", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP")
    def gestao_operacao_editar(pedido_id):
        pedido = db.session.get(PedidoOperacao, pedido_id)
        if pedido is None:
            flash("Pedido não encontrado.", "danger")
            return redirect(url_for("gestao_operacao_listagem_geral"))

        # Pedido do Bruno (31/08/2026): cada aba só edita os campos da
        # própria área (PCP só mostra/edita campos de PCP, Logística só os
        # de Logística...) — nunca lê nem sobrescreve campos de outra seção,
        # mesmo que o form de outra aba tivesse ficado aberto em outra guia.
        secao = request.args.get("secao", "comercial")
        if secao not in GO_SECOES:
            secao = "comercial"
        campos_secao = GO_CAMPOS_POR_SECAO[secao]

        transportadoras = Transportadora.query.filter_by(ativo=True).order_by(Transportadora.nome).all()

        if request.method == "POST":
            f = request.form
            if secao == "pcp":
                # Pedido do Bruno (01/09/2026, ampliado 03/09/2026): quando já
                # existe o dado automático vindo de Gestão Produção pra
                # "Previsão liberação PCP"/"Data efetiva liberação"/
                # "Solicitada cliente/retira"/"Término semanal", essas colunas
                # viram só leitura aqui — não sobrescreve (nem zera) o campo
                # antigo de PedidoOperacao ao salvar o resto da aba PCP, senão
                # o valor manual histórico se perderia à toa mesmo sem o
                # usuário ter mexido nele (o formulário nem mostra mais um
                # <input> pra esses campos nesse caso — ver
                # gestao_operacao_editar.html).
                chave_pv = (pedido.pedido_venda or "").strip()
                liberacao_real_pcp = _liberacao_pcp_por_pedido_venda([pedido.pedido_venda]).get(chave_pv) or {}
                data_cliente_real = _data_cliente_por_pedido_venda([pedido.pedido_venda]).get(chave_pv)
                campos_secao = [
                    c for c in campos_secao
                    if not (c == "go_previsao_liberacao_pcp" and liberacao_real_pcp.get("previsao"))
                    and not (c == "go_data_efetiva_liberacao_pcp" and liberacao_real_pcp.get("efetiva"))
                    and not (c == "go_data_solicitada_cliente_retira" and data_cliente_real)
                    and not (c == "go_termino_semanal_pcp" and liberacao_real_pcp.get("termino_semanal"))
                ]
            elif secao == "comercial":
                # Mesmo espírito acima, pro campo equivalente da aba Comercial
                # (pedido do Bruno, 03/09/2026: "quero que todos os dados
                # dentro da gestão operação seja extraída automaticamente da
                # gestão produção").
                chave_pv = (pedido.pedido_venda or "").strip()
                data_cliente_real = _data_cliente_por_pedido_venda([pedido.pedido_venda]).get(chave_pv)
                campos_secao = [
                    c for c in campos_secao
                    if not (c == "go_data_solicitada_entrega" and data_cliente_real)
                ]
            campos_historico = [c for c in campos_secao if c in CAMPOS_HISTORICO_GESTAO_OPERACAO]
            antes = {c: getattr(pedido, c) for c in campos_historico}

            for campo in campos_secao:
                setattr(pedido, campo, _parse_campo_go(campo, f))

            depois = {c: getattr(pedido, c) for c in campos_historico}
            # pedido_id fica None de propósito: essa coluna tem FK de verdade pra
            # `pedidos` (Gestão Produção) — PedidoOperacao é uma tabela totalmente
            # separada, então gravar o id dela ali violaria a FK. entidade_id já
            # guarda o id certo.
            _registrar_alteracoes("pedido_operacao", pedido.id, None, antes, depois, campos_historico)

            db.session.commit()
            flash(f"Gestão Operação ({GO_SECAO_LABEL[secao]}) do pedido atualizada com sucesso.", "success")
            return redirect(url_for("gestao_operacao_editar", pedido_id=pedido.id, secao=secao))

        chave_pv = (pedido.pedido_venda or "").strip()
        status_real = _status_producao_por_pedido_venda([pedido.pedido_venda]).get(chave_pv)
        liberacao_real = _liberacao_pcp_por_pedido_venda([pedido.pedido_venda]).get(chave_pv) or {}
        data_cliente_real = _data_cliente_por_pedido_venda([pedido.pedido_venda]).get(chave_pv)

        return render_template(
            "gestao_operacao_editar.html", pedido=pedido, transportadoras=transportadoras,
            secao=secao, GO_SECOES=GO_SECOES, GO_SECAO_ENDPOINT=GO_SECAO_ENDPOINT, GO_SECAO_LABEL=GO_SECAO_LABEL,
            status_real=status_real, liberacao_real=liberacao_real, data_cliente_real=data_cliente_real,
        )

    # ------------------------------------------------------------------
    # Qualidade — RNC (Relatório de Não Conformidade). Área nova (31/08/2026),
    # independente de Gestão Produção/Operação. Pedido do Bruno: controle
    # totalmente manual e intuitivo, aberto a todos os usuários autenticados
    # (view + edição) — sem role dedicado, confirmado com ele antes de
    # implementar (ver decisão em AskUserQuestion).
    # ------------------------------------------------------------------
    @app.route("/qualidade/dashboard")
    @login_required
    def qualidade_dashboard():
        dados = _dashboard_rnc_qualidade()
        return render_template("qualidade_dashboard.html", dados=dados)

    @app.route("/qualidade")
    @login_required
    def qualidade_lista():
        rncs, page, total_paginas, total_filtrado, filtros = _linhas_rnc_qualidade(request.args)
        opcoes_filtro = dict(
            status_geral=_rnc_opcoes_filtro("status_geral", RNC_STATUS_GERAL_OPCOES),
            severidade=_rnc_opcoes_filtro("severidade", RNC_SEVERIDADE_OPCOES),
            origem=_rnc_opcoes_filtro("origem", RNC_ORIGEM_OPCOES),
            tipo_nc=_rnc_opcoes_filtro("tipo_nc", RNC_TIPO_NC_OPCOES),
            setor=_rnc_opcoes_filtro("setor", RNC_SETOR_OPCOES),
        )
        return render_template(
            "qualidade_lista.html",
            rncs=rncs, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros,
            opcoes_filtro=opcoes_filtro,
        )

    @app.route("/qualidade/novo", methods=["GET", "POST"])
    @login_required
    def qualidade_novo():
        if request.method == "POST":
            f = request.form
            cliente_projeto = f.get("cliente_projeto", "").strip()
            descricao_nc = f.get("descricao_nc", "").strip()
            if not cliente_projeto or not descricao_nc:
                flash("Cliente/Projeto e Descrição da Não Conformidade são obrigatórios.", "danger")
                return render_template("qualidade_novo.html", valores=f)

            valores = _campos_form_rnc(f)
            novo = RncQualidade(criado_por_id=current_user.id, **valores)
            db.session.add(novo)
            db.session.commit()
            flash(f"RNC {novo.numero_rnc or ('#' + str(novo.id))} cadastrado com sucesso.", "success")
            return redirect(url_for("qualidade_editar", rnc_id=novo.id))

        return render_template("qualidade_novo.html", valores={})

    @app.route("/qualidade/<int:rnc_id>/editar", methods=["GET", "POST"])
    @login_required
    def qualidade_editar(rnc_id):
        rnc = db.session.get(RncQualidade, rnc_id)
        if rnc is None:
            flash("RNC não encontrado.", "danger")
            return redirect(url_for("qualidade_lista"))

        if request.method == "POST":
            f = request.form
            cliente_projeto = f.get("cliente_projeto", "").strip()
            descricao_nc = f.get("descricao_nc", "").strip()
            if not cliente_projeto or not descricao_nc:
                flash("Cliente/Projeto e Descrição da Não Conformidade são obrigatórios.", "danger")
                return render_template("qualidade_editar.html", rnc=rnc, valores=f, historico=[])

            antes = {c: getattr(rnc, c) for c in CAMPOS_HISTORICO_RNC}
            valores = _campos_form_rnc(f)
            for campo, valor in valores.items():
                setattr(rnc, campo, valor)
            depois = {c: getattr(rnc, c) for c in CAMPOS_HISTORICO_RNC}
            _registrar_alteracoes("rnc_qualidade", rnc.id, None, antes, depois, CAMPOS_HISTORICO_RNC)

            db.session.commit()
            flash("RNC atualizado com sucesso.", "success")
            return redirect(url_for("qualidade_editar", rnc_id=rnc.id))

        historico = (
            HistoricoAlteracao.query
            .filter_by(entidade_tipo="rnc_qualidade", entidade_id=rnc.id)
            .order_by(HistoricoAlteracao.criado_em.desc())
            .all()
        )
        return render_template("qualidade_editar.html", rnc=rnc, valores=_rnc_para_form_dict(rnc), historico=historico)

    # ------------------------------------------------------------------
    # Qualidade — Inspeção Final / RDIM (pedido do Bruno, 02/09/2026).
    # ------------------------------------------------------------------
    @app.route("/qualidade/rdim/buscar-itens")
    @login_required
    def rdim_buscar_itens():
        """Sugestões (JSON) pro campo de busca de item/OP da tela de Nova
        Inspeção — só itens das estações MANDRIL/PU/SILICONE."""
        termo = request.args.get("q", "")
        return jsonify(_itens_rdim_disponiveis(termo))

    @app.route("/qualidade/rdim/dashboard")
    @login_required
    def rdim_dashboard():
        dados = _dashboard_rdim()
        return render_template("qualidade_rdim_dashboard.html", dados=dados)

    @app.route("/qualidade/rdim")
    @login_required
    def rdim_lista():
        inspecoes, page, total_paginas, total_filtrado, filtros = _linhas_inspecoes_finais(request.args)
        responsaveis = (
            Usuario.query.join(InspecaoFinal, InspecaoFinal.responsavel_id == Usuario.id)
            .distinct().order_by(Usuario.nome).all()
        )
        return render_template(
            "qualidade_rdim_lista.html",
            inspecoes=inspecoes, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros, responsaveis=responsaveis,
        )

    @app.route("/qualidade/rdim/novo", methods=["GET", "POST"])
    @login_required
    def rdim_novo():
        item_id_inicial = request.args.get("item_id", "").strip()

        if request.method == "POST":
            f = request.form
            item_id = f.get("item_pedido_id", "").strip()
            resultado = f.get("resultado", "").strip()
            item = db.session.get(ItemPedido, int(item_id)) if item_id.isdigit() else None

            if item is None:
                flash("Selecione uma OP (item de pedido) válida antes de salvar.", "danger")
                return render_template("qualidade_rdim_novo.html", valores=f, item_selecionado=None)
            if resultado not in RDIM_RESULTADO_OPCOES:
                flash("Selecione o resultado da inspeção (Aprovado / Reprovado / Aprovado com desvio).", "danger")
                return render_template("qualidade_rdim_novo.html", valores=f, item_selecionado=item)

            quantidade_com_desvio, erro_qtd = _validar_quantidade_com_desvio(f.get("quantidade_com_desvio"), item.quantidade)
            if erro_qtd:
                flash(erro_qtd, "danger")
                return render_template("qualidade_rdim_novo.html", valores=f, item_selecionado=item)

            nova = InspecaoFinal(
                item_pedido_id=item.id,
                estacao=item.estacao,
                data_inspecao=_parse_data_form(f.get("data_inspecao")) or date.today(),
                responsavel_id=current_user.id,
                numero_rif=f.get("numero_rif", "").strip() or None,
                procedimento=f.get("procedimento", "").strip() or None,
                norma=f.get("norma", "").strip() or None,
                instrucao_trabalho=f.get("instrucao_trabalho", "").strip() or None,
                inspecao_visual=f.get("inspecao_visual", "").strip() or None,
                desvio_encontrado=f.get("desvio_encontrado", "").strip() or None,
                categoria_desvio=f.get("categoria_desvio", "").strip() or None,
                subcategoria_desvio=f.get("subcategoria_desvio", "").strip() or None,
                observacao=f.get("observacao", "").strip() or None,
                resultado=resultado,
                quantidade_com_desvio=quantidade_com_desvio,
                criado_por_id=current_user.id,
            )
            db.session.add(nova)
            _salvar_medicoes_rdim(nova, f)
            _salvar_pecas_desvio_rdim(nova, f)
            db.session.commit()
            flash("Inspeção final registrada com sucesso.", "success")
            return redirect(url_for("rdim_editar", inspecao_id=nova.id))

        item_selecionado = None
        if item_id_inicial.isdigit():
            item_selecionado = db.session.get(ItemPedido, int(item_id_inicial))
        return render_template("qualidade_rdim_novo.html", valores={}, item_selecionado=item_selecionado)

    @app.route("/qualidade/rdim/<int:inspecao_id>/editar", methods=["GET", "POST"])
    @login_required
    def rdim_editar(inspecao_id):
        inspecao = db.session.get(InspecaoFinal, inspecao_id)
        if inspecao is None:
            flash("Inspeção não encontrada.", "danger")
            return redirect(url_for("rdim_lista"))

        if request.method == "POST":
            f = request.form
            resultado = f.get("resultado", "").strip()
            if resultado not in RDIM_RESULTADO_OPCOES:
                flash("Selecione o resultado da inspeção (Aprovado / Reprovado / Aprovado com desvio).", "danger")
                historico = (
                    HistoricoAlteracao.query.filter_by(entidade_tipo="inspecao_final", entidade_id=inspecao.id)
                    .order_by(HistoricoAlteracao.criado_em.desc()).all()
                )
                return render_template("qualidade_rdim_editar.html", inspecao=inspecao, historico=historico)

            quantidade_item = inspecao.item.quantidade if inspecao.item else None
            quantidade_com_desvio, erro_qtd = _validar_quantidade_com_desvio(f.get("quantidade_com_desvio"), quantidade_item)
            if erro_qtd:
                flash(erro_qtd, "danger")
                historico = (
                    HistoricoAlteracao.query.filter_by(entidade_tipo="inspecao_final", entidade_id=inspecao.id)
                    .order_by(HistoricoAlteracao.criado_em.desc()).all()
                )
                return render_template("qualidade_rdim_editar.html", inspecao=inspecao, historico=historico)

            antes = {c: getattr(inspecao, c) for c in CAMPOS_HISTORICO_INSPECAO_FINAL}
            inspecao.data_inspecao = _parse_data_form(f.get("data_inspecao")) or inspecao.data_inspecao
            inspecao.numero_rif = f.get("numero_rif", "").strip() or None
            inspecao.procedimento = f.get("procedimento", "").strip() or None
            inspecao.norma = f.get("norma", "").strip() or None
            inspecao.instrucao_trabalho = f.get("instrucao_trabalho", "").strip() or None
            inspecao.inspecao_visual = f.get("inspecao_visual", "").strip() or None
            inspecao.desvio_encontrado = f.get("desvio_encontrado", "").strip() or None
            inspecao.categoria_desvio = f.get("categoria_desvio", "").strip() or None
            inspecao.subcategoria_desvio = f.get("subcategoria_desvio", "").strip() or None
            inspecao.observacao = f.get("observacao", "").strip() or None
            inspecao.resultado = resultado
            inspecao.quantidade_com_desvio = quantidade_com_desvio
            depois = {c: getattr(inspecao, c) for c in CAMPOS_HISTORICO_INSPECAO_FINAL}
            _registrar_alteracoes("inspecao_final", inspecao.id, inspecao.item.pedido_id if inspecao.item else None,
                                   antes, depois, CAMPOS_HISTORICO_INSPECAO_FINAL)

            _salvar_medicoes_rdim(inspecao, f, substituir=True)
            _salvar_pecas_desvio_rdim(inspecao, f, substituir=True)
            db.session.commit()
            flash("Inspeção atualizada com sucesso.", "success")
            return redirect(url_for("rdim_editar", inspecao_id=inspecao.id))

        historico = (
            HistoricoAlteracao.query.filter_by(entidade_tipo="inspecao_final", entidade_id=inspecao.id)
            .order_by(HistoricoAlteracao.criado_em.desc()).all()
        )
        return render_template("qualidade_rdim_editar.html", inspecao=inspecao, historico=historico)

    @app.route("/qualidade/rdim/<int:inspecao_id>/excluir", methods=["POST"])
    @requer_role("ADMIN", "PCP")
    def rdim_excluir(inspecao_id):
        """Apagar uma inspeção RDIM definitivamente — pedido do Bruno
        (02/09/2026): só tinha editar, faltava excluir. Restrito a ADMIN/PCP,
        mesmo critério já usado em excluir_pedido — é um registro formal de
        inspeção (RIF/norma/procedimento), então a exclusão fica mais
        controlada que criar/editar (aberto a todo usuário autenticado).
        RdimMedicao e RdimPecaDesvio somem juntos (cascade="all,
        delete-orphan" já configurado no relacionamento)."""
        inspecao = db.session.get(InspecaoFinal, inspecao_id)
        if inspecao is not None:
            db.session.delete(inspecao)
            db.session.commit()
            flash("Inspeção RDIM excluída.", "info")
        return redirect(url_for("rdim_lista"))

    # ------------------------------------------------------------------
    # P&D — Pesquisa e Desenvolvimento (Fase 14, 01/09/2026). Área nova,
    # independente de PCP/Produção/Operação/Qualidade. Mesmo critério de
    # acesso já adotado em Qualidade (aberto a todo usuário autenticado,
    # sem role dedicado) — usada principalmente por Bruno e Gustavo Fugita,
    # mas sem restringir os outros papéis já existentes no sistema.
    # ------------------------------------------------------------------
    @app.route("/pd/dashboard")
    @login_required
    def pd_dashboard():
        dados = _dashboard_pd()
        return render_template("pd_dashboard.html", dados=dados)

    @app.route("/pd")
    @login_required
    def pd_lista():
        projetos, page, total_paginas, total_filtrado, filtros = _linhas_projetos_pd(request.args)
        opcoes_filtro = dict(
            etapa=PD_ETAPA_OPCOES,
            categoria=_pd_opcoes_filtro("categoria", PD_CATEGORIA_OPCOES),
            prioridade=PRIORIDADE_OPCOES,
        )
        return render_template(
            "pd_lista.html",
            projetos=projetos, page=page, total_paginas=total_paginas,
            total_filtrado=total_filtrado, filtros=filtros, opcoes_filtro=opcoes_filtro,
        )

    @app.route("/pd/novo", methods=["GET", "POST"])
    @login_required
    def pd_novo():
        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip()
            if not nome:
                flash("Nome do projeto é obrigatório.", "danger")
                valores_repopular = {k: f.get(k, "") for k in f.keys()}
                valores_repopular["resultado_esperado"] = f.getlist("resultado_esperado")
                return render_template("pd_novo.html", valores=valores_repopular)

            valores = _campos_form_pd(f)
            novo = ProjetoPD(criado_por_id=current_user.id, **valores)
            db.session.add(novo)
            db.session.commit()
            flash(f"Projeto {novo.codigo or ('#' + str(novo.id))} cadastrado com sucesso.", "success")
            return redirect(url_for("pd_editar", projeto_id=novo.id))

        return render_template("pd_novo.html", valores={})

    @app.route("/pd/<int:projeto_id>/editar", methods=["GET", "POST"])
    @login_required
    def pd_editar(projeto_id):
        projeto = db.session.get(ProjetoPD, projeto_id)
        if projeto is None:
            flash("Projeto de P&D não encontrado.", "danger")
            return redirect(url_for("pd_lista"))

        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip()
            if not nome:
                flash("Nome do projeto é obrigatório.", "danger")
                return redirect(url_for("pd_editar", projeto_id=projeto.id))

            antes = {c: getattr(projeto, c) for c in CAMPOS_HISTORICO_PD}
            valores = _campos_form_pd(f)
            for campo, valor in valores.items():
                setattr(projeto, campo, valor)
            depois = {c: getattr(projeto, c) for c in CAMPOS_HISTORICO_PD}
            _registrar_alteracoes("projeto_pd", projeto.id, None, antes, depois, CAMPOS_HISTORICO_PD)

            db.session.commit()
            flash("Projeto atualizado com sucesso.", "success")
            return redirect(url_for("pd_editar", projeto_id=projeto.id))

        historico = (
            HistoricoAlteracao.query
            .filter_by(entidade_tipo="projeto_pd", entidade_id=projeto.id)
            .order_by(HistoricoAlteracao.criado_em.desc())
            .all()
        )
        return render_template(
            "pd_editar.html", projeto=projeto, valores=_pd_para_form_dict(projeto), historico=historico,
        )

    @app.route("/pd/<int:projeto_id>/testes/novo", methods=["POST"])
    @login_required
    def pd_teste_novo(projeto_id):
        projeto = db.session.get(ProjetoPD, projeto_id)
        if projeto is None:
            flash("Projeto de P&D não encontrado.", "danger")
            return redirect(url_for("pd_lista"))

        f = request.form
        teste = TesteProjetoPD(
            projeto_id=projeto.id,
            numero=f.get("numero", "").strip() or None,
            data_planejada=_parse_data_form(f.get("data_planejada")),
            data_realizada=_parse_data_form(f.get("data_realizada")),
            responsavel=f.get("responsavel", "").strip() or None,
            material_utilizado=f.get("material_utilizado", "").strip() or None,
            lote=f.get("lote", "").strip() or None,
            fornecedor=f.get("fornecedor", "").strip() or None,
            condicoes=f.get("condicoes", "").strip() or None,
            resultado=f.get("resultado", "").strip() or "Planejado",
            observacoes=f.get("observacoes", "").strip() or None,
            anexos=f.get("anexos", "").strip() or None,
        )
        db.session.add(teste)
        db.session.commit()
        flash("Teste registrado com sucesso.", "success")
        return redirect(url_for("pd_editar", projeto_id=projeto.id) + "#testes")

    @app.route("/pd/<int:projeto_id>/eventos/novo", methods=["POST"])
    @login_required
    def pd_evento_novo(projeto_id):
        projeto = db.session.get(ProjetoPD, projeto_id)
        if projeto is None:
            flash("Projeto de P&D não encontrado.", "danger")
            return redirect(url_for("pd_lista"))

        f = request.form
        evento = VisitaReuniaoPD(
            projeto_id=projeto.id,
            data=_parse_data_form(f.get("data")),
            tipo=f.get("tipo", "").strip() or None,
            participantes=f.get("participantes", "").strip() or None,
            local=f.get("local", "").strip() or None,
            objetivo=f.get("objetivo", "").strip() or None,
            resultado=f.get("resultado", "").strip() or None,
            proximas_acoes=f.get("proximas_acoes", "").strip() or None,
            responsavel=f.get("responsavel", "").strip() or None,
            anexos=f.get("anexos", "").strip() or None,
        )
        db.session.add(evento)
        db.session.commit()
        flash("Visita/reunião registrada com sucesso.", "success")
        return redirect(url_for("pd_editar", projeto_id=projeto.id) + "#eventos")

    # Kanban de P&D — pedido do Bruno: "visão alternada" dentro de Projetos,
    # colunas = as 8 etapas do ciclo de vida, arrastável. Mover uma coluna
    # pra outra usa a mesma rota que o <select> de fallback (sem JS/toque) —
    # os dois caminhos (drag-and-drop e o <select>) chamam pd_mover_etapa.
    @app.route("/pd/kanban")
    @login_required
    def pd_kanban():
        projetos = ProjetoPD.query.order_by(ProjetoPD.data_prevista_conclusao.asc().nullslast(), ProjetoPD.id.desc()).all()
        colunas = {etapa: [] for etapa in PD_ETAPA_OPCOES}
        for p in projetos:
            colunas.setdefault(p.etapa_atual, []).append(p)
        return render_template("pd_kanban.html", colunas=colunas)

    @app.route("/pd/<int:projeto_id>/mover-etapa", methods=["POST"])
    @login_required
    def pd_mover_etapa(projeto_id):
        projeto = db.session.get(ProjetoPD, projeto_id)
        if projeto is None:
            flash("Projeto de P&D não encontrado.", "danger")
            return redirect(url_for("pd_kanban"))

        nova_etapa = request.form.get("etapa", "").strip()
        if nova_etapa not in PD_ETAPA_OPCOES:
            flash("Etapa inválida.", "danger")
            return redirect(url_for("pd_kanban"))

        if nova_etapa != projeto.etapa_atual:
            antes = {c: getattr(projeto, c) for c in CAMPOS_HISTORICO_PD}
            projeto.etapa_atual = nova_etapa
            # Conveniência: ao mover pra "Concluído" sem data real de
            # conclusão ainda preenchida, carimba hoje sozinho (o Bruno pode
            # sempre corrigir depois na aba "Dados do projeto").
            if nova_etapa == "Concluído" and not projeto.data_real_conclusao:
                projeto.data_real_conclusao = date.today()
            depois = {c: getattr(projeto, c) for c in CAMPOS_HISTORICO_PD}
            _registrar_alteracoes("projeto_pd", projeto.id, None, antes, depois, CAMPOS_HISTORICO_PD)
            db.session.commit()
            flash(f"\"{projeto.nome}\" movido para {nova_etapa}.", "success")

        return redirect(url_for("pd_kanban"))

    @app.route("/pd/cronograma")
    @login_required
    def pd_cronograma():
        dados = _cronograma_pd()
        return render_template("pd_cronograma.html", **dados)

    @app.route("/pd/testes")
    @login_required
    def pd_testes_lista():
        query, filtros = _filtrar_testes_pd(request.args)
        testes = query.all()
        projetos_opcoes = ProjetoPD.query.order_by(ProjetoPD.nome).all()
        return render_template("pd_testes.html", testes=testes, projetos_opcoes=projetos_opcoes, filtros=filtros)

    @app.route("/pd/custos")
    @login_required
    def pd_custos():
        dados = _custos_pd()
        return render_template("pd_custos.html", **dados)

    @app.route("/pd/conhecimento")
    @login_required
    def pd_conhecimento():
        busca = request.args.get("busca", "").strip()
        query = ProjetoPD.query.filter(
            or_(ProjetoPD.problema.isnot(None), ProjetoPD.solucao.isnot(None), ProjetoPD.licoes_aprendidas.isnot(None))
        )
        if busca:
            termo = f"%{busca}%"
            query = query.filter(
                or_(
                    ProjetoPD.nome.ilike(termo), ProjetoPD.problema.ilike(termo), ProjetoPD.solucao.ilike(termo),
                    ProjetoPD.licoes_aprendidas.ilike(termo), ProjetoPD.categoria.ilike(termo),
                    ProjetoPD.cliente.ilike(termo), ProjetoPD.produto.ilike(termo), ProjetoPD.fornecedor.ilike(termo),
                )
            )
        projetos = query.order_by(ProjetoPD.data_real_conclusao.desc().nullslast(), ProjetoPD.id.desc()).all()
        return render_template("pd_conhecimento.html", projetos=projetos, busca=busca)

    # ------------------------------------------------------------------
    # Usuários (papéis de acesso) — só ADMIN cadastra/edita usuários
    # ------------------------------------------------------------------
    @app.route("/usuarios")
    @requer_role("ADMIN")
    def usuarios_lista():
        usuarios = Usuario.query.order_by(Usuario.nome).all()
        return render_template("usuarios_lista.html", usuarios=usuarios)

    @app.route("/usuarios/novo", methods=["GET", "POST"])
    @requer_role("ADMIN")
    def usuarios_novo():
        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip()
            username = f.get("username", "").strip()
            senha = f.get("senha", "")
            role = f.get("role") or "PCP"
            setor = f.get("setor", "").strip() or None

            if not nome or not username or not senha:
                flash("Nome, usuário e senha são obrigatórios.", "danger")
                return render_template("usuarios_form.html", usuario=None, form=f)

            if Usuario.query.filter_by(username=username).first():
                flash("Já existe um usuário com esse nome de login.", "danger")
                return render_template("usuarios_form.html", usuario=None, form=f)

            if role not in ROLES:
                role = "PCP"

            novo = Usuario(nome=nome, username=username, role=role, setor=setor if role == "LIDER" else None)
            novo.set_senha(senha)
            db.session.add(novo)
            db.session.commit()
            flash(f"Usuário {nome} criado com sucesso.", "success")
            return redirect(url_for("usuarios_lista"))

        return render_template("usuarios_form.html", usuario=None, form={})

    @app.route("/usuarios/<int:usuario_id>/editar", methods=["GET", "POST"])
    @requer_role("ADMIN")
    def usuarios_editar(usuario_id):
        usuario = db.session.get(Usuario, usuario_id)
        if usuario is None:
            flash("Usuário não encontrado.", "danger")
            return redirect(url_for("usuarios_lista"))

        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip()
            username = f.get("username", "").strip()
            role = f.get("role") or usuario.role
            setor = f.get("setor", "").strip() or None
            senha = f.get("senha", "")

            if not nome or not username:
                flash("Nome e usuário são obrigatórios.", "danger")
                return render_template("usuarios_form.html", usuario=usuario, form=f)

            outro = Usuario.query.filter(Usuario.username == username, Usuario.id != usuario.id).first()
            if outro:
                flash("Já existe outro usuário com esse nome de login.", "danger")
                return render_template("usuarios_form.html", usuario=usuario, form=f)

            if role not in ROLES:
                role = usuario.role

            usuario.nome = nome
            usuario.username = username
            usuario.role = role
            usuario.setor = setor if role == "LIDER" else None
            usuario.ativo = bool(f.get("ativo"))
            if senha:
                usuario.set_senha(senha)

            # o usuário admin original nunca perde acesso total nem fica desativado,
            # mesmo que alguém manipule o formulário
            if usuario.username == "admin":
                usuario.role = "ADMIN"
                usuario.setor = None
                usuario.ativo = True

            db.session.commit()
            flash(f"Usuário {usuario.nome} atualizado com sucesso.", "success")
            return redirect(url_for("usuarios_lista"))

        return render_template("usuarios_form.html", usuario=usuario, form={})

    @app.route("/usuarios/<int:usuario_id>/desativar", methods=["POST"])
    @requer_role("ADMIN")
    def usuarios_desativar(usuario_id):
        usuario = db.session.get(Usuario, usuario_id)
        if usuario is None:
            flash("Usuário não encontrado.", "danger")
        elif usuario.username == "admin":
            flash("O usuário admin original não pode ser desativado.", "danger")
        else:
            usuario.ativo = not usuario.ativo
            db.session.commit()
            flash(
                f"Usuário {usuario.nome} {'reativado' if usuario.ativo else 'desativado'} com sucesso.",
                "info",
            )
        return redirect(url_for("usuarios_lista"))

    # ------------------------------------------------------------------
    # Zerar dados de teste (pedido do Bruno, 28/08/2026; ampliado pra
    # Qualidade + P&D em 03/09/2026 — "quero que exclua não só produção e
    # operação, mas também todos os dados de testes (P&D e Qualidade)") —
    # apaga TODOS os Pedido/ItemPedido (Gestão Produção), PedidoOperacao
    # (Gestão Operação), RncQualidade + InspecaoFinal/RdimMedicao/
    # RdimPecaDesvio (Qualidade) e ProjetoPD/TesteProjetoPD/VisitaReuniaoPD
    # (P&D) de uma vez, pra ele testar manualmente do zero. Só ADMIN, com
    # confirmação por texto digitado — e sempre com um backup (.xlsx)
    # disponível antes, já que é uma exclusão permanente. Cadastros
    # (usuários, estações, transportadoras) nunca são tocados.
    # ------------------------------------------------------------------
    _FRASE_CONFIRMACAO_ZERAR = "ZERAR TUDO"

    @app.route("/admin/zerar-dados")
    @requer_role("ADMIN")
    def admin_zerar_dados():
        return render_template(
            "admin_zerar_dados.html",
            total_pedidos=Pedido.query.count(),
            total_itens=ItemPedido.query.count(),
            total_pedidos_operacao=PedidoOperacao.query.count(),
            total_rnc=RncQualidade.query.count(),
            total_inspecoes_rdim=InspecaoFinal.query.count(),
            total_projetos_pd=ProjetoPD.query.count(),
            frase_confirmacao=_FRASE_CONFIRMACAO_ZERAR,
        )

    @app.route("/admin/zerar-dados/backup.xlsx")
    @requer_role("ADMIN")
    def admin_zerar_dados_backup():
        """Só gera e baixa o backup — não apaga nada. Pode ser clicado quantas
        vezes quiser antes (ou até sem intenção de zerar depois)."""
        wb = _construir_backup_pedidos_wb()
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nome = f"backup_pedidos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        resposta = Response(
            buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resposta.headers["Content-Disposition"] = f"attachment; filename={nome}"
        return resposta

    @app.route("/admin/zerar-dados", methods=["POST"])
    @requer_role("ADMIN")
    def admin_zerar_dados_confirmar():
        confirmacao = request.form.get("confirmacao", "").strip()
        if confirmacao != _FRASE_CONFIRMACAO_ZERAR:
            flash(f'Digite exatamente "{_FRASE_CONFIRMACAO_ZERAR}" pra confirmar. Nada foi apagado.', "danger")
            return redirect(url_for("admin_zerar_dados"))

        total_pedidos = Pedido.query.count()
        total_itens = ItemPedido.query.count()
        total_pedidos_operacao = PedidoOperacao.query.count()
        total_rnc = RncQualidade.query.count()
        total_inspecoes_rdim = InspecaoFinal.query.count()
        total_projetos_pd = ProjetoPD.query.count()

        # Ordem importa: filhos antes dos pais, por causa das foreign keys —
        # bulk delete (.query.delete()) não aciona cascade de ORM, só as
        # normais do banco, então cada FK precisa ser removida "na mão" na
        # ordem certa (RdimPecaDesvio/RdimMedicao -> InspecaoFinal ->
        # ItemPedido; Programacao -> ItemPedido; HistoricoAlteracao ->
        # Pedido; TesteProjetoPD/VisitaReuniaoPD -> ProjetoPD). RncQualidade
        # e ProjetoPD/PedidoOperacao são tabelas independentes (sem FK com o
        # resto), podem vir em qualquer ordem.
        RdimPecaDesvio.query.delete(synchronize_session=False)
        RdimMedicao.query.delete(synchronize_session=False)
        InspecaoFinal.query.delete(synchronize_session=False)
        Programacao.query.delete(synchronize_session=False)
        HistoricoAlteracao.query.delete(synchronize_session=False)
        ItemPedido.query.delete(synchronize_session=False)
        Pedido.query.delete(synchronize_session=False)
        PedidoOperacao.query.delete(synchronize_session=False)
        RncQualidade.query.delete(synchronize_session=False)
        TesteProjetoPD.query.delete(synchronize_session=False)
        VisitaReuniaoPD.query.delete(synchronize_session=False)
        ProjetoPD.query.delete(synchronize_session=False)

        # Marca que foi de propósito — sem isso, _seed_inicial e
        # _importar_gestao_operacao reimportariam a planilha antiga sozinhos
        # no próximo deploy, assim que virem as tabelas vazias. (RNC e P&D
        # não têm reimportação automática recorrente — _seed_rnc_qualidade já
        # roda uma única vez, guardado por ControleSistema próprio, e nunca
        # reimporta de novo mesmo com a tabela vazia; P&D não tem seed nenhum.)
        if ControleSistema.query.filter_by(chave=_CHAVE_DADOS_ZERADOS_MANUALMENTE).first() is None:
            db.session.add(ControleSistema(chave=_CHAVE_DADOS_ZERADOS_MANUALMENTE))

        db.session.commit()
        app.logger.warning(
            "ZERAR DADOS: %s executou a limpeza manual — %d pedidos, %d itens, %d pedidos de Gestão "
            "Operação, %d RNCs, %d inspeções RDIM e %d projetos de P&D apagados.",
            current_user.nome, total_pedidos, total_itens, total_pedidos_operacao,
            total_rnc, total_inspecoes_rdim, total_projetos_pd,
        )
        flash(
            f"Pronto: {total_pedidos} pedido(s), {total_itens} item(ns), {total_pedidos_operacao} "
            f"pedido(s) de Gestão Operação, {total_rnc} RNC(s), {total_inspecoes_rdim} inspeção(ões) "
            f"RDIM e {total_projetos_pd} projeto(s) de P&D foram apagados. Pode começar a testar do zero.",
            "success",
        )
        return redirect(url_for("admin_zerar_dados"))

    # ------------------------------------------------------------------
    # Estações — visão geral por setor + Kanban de produção
    # ------------------------------------------------------------------
    @app.route("/estacoes")
    @login_required
    def estacoes_lista():
        hoje = date.today()
        estacoes_por_nome = {e.nome: e for e in Estacao.query.all()}

        def _linha(e):
            fila = ItemPedido.query.filter(
                ItemPedido.estacao == e.nome, ItemPedido.status_producao != "FINALIZADO"
            ).count()
            criticos = ItemPedido.query.filter(
                ItemPedido.estacao == e.nome,
                ItemPedido.status_producao != "FINALIZADO",
                ItemPedido.liberacao_prevista.isnot(None),
                ItemPedido.liberacao_prevista < hoje,
            ).count()
            itens_lt = ItemPedido.query.filter(
                ItemPedido.estacao == e.nome,
                ItemPedido.inicio_producao.isnot(None),
                ItemPedido.termino_inspecao.isnot(None),
            ).all()
            lt_medio = (
                round(sum((i.termino_inspecao - i.inicio_producao).days for i in itens_lt) / len(itens_lt), 1)
                if itens_lt
                else None
            )
            return {"estacao": e, "rotulo": rotulo_estacao(e.nome), "fila": fila, "criticos": criticos, "lt_medio": lt_medio}

        # 3 colunas fixas (pedido do Bruno, 03/09/2026) — ver
        # ESTACOES_GRUPOS_MONITORAMENTO em models.py. Só entram estações
        # ativas; uma estação ativa que não conste em nenhum grupo cai numa
        # coluna extra "Outras estações" no final, pra nunca sumir em
        # silêncio do monitoramento (ex.: uma estação nova cadastrada depois
        # sem eu saber encaixar num dos 3 grupos).
        grupos = []
        nomes_agrupados = set()
        for grupo in ESTACOES_GRUPOS_MONITORAMENTO:
            linhas_grupo = []
            for nome in grupo["estacoes"]:
                nomes_agrupados.add(nome)
                e = estacoes_por_nome.get(nome)
                if e is not None and e.ativo:
                    linhas_grupo.append(_linha(e))
            grupos.append({"titulo": grupo["titulo"], "linhas": linhas_grupo})

        extras = [
            _linha(e)
            for nome, e in sorted(estacoes_por_nome.items())
            if nome not in nomes_agrupados and e.ativo
        ]
        if extras:
            grupos.append({"titulo": "Outras estações", "linhas": extras})

        return render_template("estacoes_lista.html", grupos=grupos)

    @app.route("/estacoes/<nome>")
    @login_required
    def estacao_kanban(nome):
        estacao = Estacao.query.filter_by(nome=nome).first()
        if estacao is None:
            flash("Estação não encontrada.", "danger")
            return redirect(url_for("estacoes_lista"))

        itens = ItemPedido.query.options(selectinload(ItemPedido.pedido)).filter(ItemPedido.estacao == nome).all()

        colunas = {chave: [] for chave in STATUS_CHAO_OPCOES}
        for item in itens:
            colunas[item.status_chao].append(item)

        # Pedido do Bruno (03/09/2026, em TODAS as estações): em Pendente e Em
        # produção, prazo mais curto (liberação prevista) sempre primeiro —
        # quem "aperta" mais fica visível sem precisar rolar a coluna. Item
        # sem liberação prevista cadastrada vai pro final da coluna (não tem
        # como saber se é urgente ou não).
        def _chave_prazo(item):
            return (item.liberacao_prevista is None, item.liberacao_prevista or date.max, item.id)

        for chave in ("PENDENTE", "EM_PRODUCAO"):
            colunas[chave].sort(key=_chave_prazo)

        # Finalizado continua no critério antigo (25/08/2026): item mais
        # recém concluído/movimentado no topo — depois de entregue, prazo
        # previsto deixa de ser o que importa pra essa coluna.
        colunas["FINALIZADO"].sort(key=lambda i: (i.atualizado_em or datetime.min, i.id), reverse=True)

        return render_template(
            "estacoes_kanban.html",
            estacao=estacao,
            rotulo=rotulo_estacao(estacao.nome),
            colunas=colunas,
            pode_editar=pode_editar_estacao(current_user, nome),
        )

    @app.route("/estacoes/<nome>/kanban/mover", methods=["POST"])
    @login_required
    def estacao_kanban_mover(nome):
        item_id = request.form.get("item_id", type=int)
        item = db.session.get(ItemPedido, item_id) if item_id else None
        if item is None:
            flash("Item não encontrado.", "danger")
            return redirect(url_for("estacao_kanban", nome=nome))

        if not pode_editar_estacao(current_user, item.estacao or nome):
            flash("Você não tem permissão para mover itens desta estação.", "danger")
            return redirect(url_for("estacao_kanban", nome=nome))

        antes = {c: getattr(item, c) for c in CAMPOS_HISTORICO_ITEM}
        hoje = date.today()
        etapa_atual = item.status_chao

        # Kanban simplificado (3 colunas): "avançar" carimba a data que falta
        # pra status_producao concluir sozinho a próxima etapa (mesma regra
        # de atualizar_status_automatico, sem duplicar lógica aqui).
        if etapa_atual == "PENDENTE":
            item.inicio_producao = item.inicio_producao or hoje
        elif etapa_atual == "EM_PRODUCAO":
            item.termino_inspecao = item.termino_inspecao or hoje
            item.liberacao_faturamento = item.liberacao_faturamento or hoje
        else:
            flash("Este item já está finalizado.", "info")
            return redirect(url_for("estacao_kanban", nome=nome))

        item.status_manual = False
        item.atualizar_status_automatico()

        depois = {c: getattr(item, c) for c in CAMPOS_HISTORICO_ITEM}
        _registrar_alteracoes("item_pedido", item.id, item.pedido_id, antes, depois, CAMPOS_HISTORICO_ITEM)
        db.session.commit()

        flash(
            f"\"{item.descricao_produto}\" avançou para {STATUS_CHAO_LABELS.get(item.status_chao, item.status_chao)}.",
            "success",
        )
        return redirect(url_for("estacao_kanban", nome=nome))

    # ------------------------------------------------------------------
    # Cadastros > Estações — CRUD simples (ADMIN/PCP)
    # ------------------------------------------------------------------
    @app.route("/cadastros/estacoes")
    @requer_role("ADMIN", "PCP")
    def cadastros_estacoes():
        estacoes = Estacao.query.order_by(Estacao.ordem_exibicao).all()
        return render_template("cadastros_estacoes.html", estacoes=estacoes)

    @app.route("/cadastros/estacoes/novo", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP")
    def cadastros_estacoes_novo():
        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip().upper()
            if not nome:
                flash("Informe o nome da estação.", "danger")
                return render_template("cadastros_estacoes_form.html", estacao=None, form=f)
            if Estacao.query.filter_by(nome=nome).first():
                flash("Já existe uma estação com esse nome.", "danger")
                return render_template("cadastros_estacoes_form.html", estacao=None, form=f)

            maior_ordem = db.session.query(func.max(Estacao.ordem_exibicao)).scalar() or 0
            meta = f.get("meta_lead_time_dias", "").strip()
            nova = Estacao(
                nome=nome,
                ordem_exibicao=maior_ordem + 1,
                meta_lead_time_dias=int(meta) if meta.isdigit() else None,
                ativo=True,
            )
            db.session.add(nova)
            db.session.commit()
            flash(f"Estação {nome} cadastrada com sucesso.", "success")
            return redirect(url_for("cadastros_estacoes"))

        return render_template("cadastros_estacoes_form.html", estacao=None, form={})

    @app.route("/cadastros/estacoes/<int:estacao_id>/editar", methods=["GET", "POST"])
    @requer_role("ADMIN", "PCP")
    def cadastros_estacoes_editar(estacao_id):
        estacao = db.session.get(Estacao, estacao_id)
        if estacao is None:
            flash("Estação não encontrada.", "danger")
            return redirect(url_for("cadastros_estacoes"))

        if request.method == "POST":
            f = request.form
            nome = f.get("nome", "").strip().upper()
            if not nome:
                flash("Informe o nome da estação.", "danger")
                return render_template("cadastros_estacoes_form.html", estacao=estacao, form=f)
            outra = Estacao.query.filter(Estacao.nome == nome, Estacao.id != estacao.id).first()
            if outra:
                flash("Já existe outra estação com esse nome.", "danger")
                return render_template("cadastros_estacoes_form.html", estacao=estacao, form=f)

            estacao.nome = nome
            meta = f.get("meta_lead_time_dias", "").strip()
            estacao.meta_lead_time_dias = int(meta) if meta.isdigit() else None
            estacao.ativo = bool(f.get("ativo"))
            db.session.commit()
            flash(f"Estação {estacao.nome} atualizada com sucesso.", "success")
            return redirect(url_for("cadastros_estacoes"))

        return render_template("cadastros_estacoes_form.html", estacao=estacao, form={})

    # ------------------------------------------------------------------
    # Programação — calendário mensal do PCP (pedido do Bruno, 31/08/2026):
    # 1 linha por semana (domingo a sábado) do mês, item já entra no dia certo
    # sozinho pela liberação prevista (azul) ou liberação real (verde), sem
    # precisar programar nada manualmente. Substitui o board antigo de
    # segunda-sexta com "+"/modal (ver _semanas_calendario_pcp acima).
    # ------------------------------------------------------------------
    @app.route("/programacao")
    @login_required
    def programacao_semana():
        hoje = date.today()
        ano = request.args.get("ano", hoje.year, type=int)
        mes = request.args.get("mes", hoje.month, type=int)
        if not (1 <= mes <= 12):
            ano, mes = hoje.year, hoje.month

        semanas = _semanas_calendario_pcp(ano, mes)
        inicio_range = semanas[0]["inicio"]
        fim_range = semanas[-1]["fim"]

        itens = (
            ItemPedido.query.options(selectinload(ItemPedido.pedido))
            .filter(
                or_(
                    ItemPedido.liberacao_real.between(inicio_range, fim_range),
                    and_(
                        ItemPedido.liberacao_real.is_(None),
                        ItemPedido.liberacao_prevista.between(inicio_range, fim_range),
                    ),
                )
            )
            .all()
        )

        for semana in semanas:
            # cada dia guarda um GRUPO por pedido, não por item — desde que
            # Liberação prevista/real virou campo único do pedido (cascateia
            # pra todos os itens ao salvar, pedido do Bruno 31/08/2026), todo
            # item do mesmo pedido cai sempre no mesmo dia, então o quadrante
            # mostra 1 card por pedido (resumo), com os itens dentro do
            # hover/clique — em vez de 1 card por item repetindo cliente/frete/etc.
            semana["dias"] = [{} for _ in range(7)]  # 0=domingo ... 6=sábado, cada um {pedido_id: grupo}
            semana["dias_datas"] = [semana["inicio"] + timedelta(days=i) for i in range(7)]

        for item in itens:
            data_efetiva = item.liberacao_real or item.liberacao_prevista
            for semana in semanas:
                if semana["inicio"] <= data_efetiva <= semana["fim"]:
                    coluna = (data_efetiva.weekday() + 1) % 7  # weekday(): segunda=0 -> domingo vira 0
                    grupos = semana["dias"][coluna]
                    grupo = grupos.get(item.pedido_id)
                    if grupo is None:
                        grupo = {"pedido": item.pedido, "itens": [], "data": data_efetiva}
                        grupos[item.pedido_id] = grupo
                    grupo["itens"].append(item)
                    break

        for semana in semanas:
            for coluna_idx, grupos in enumerate(semana["dias"]):
                lista = sorted(grupos.values(), key=lambda g: ((g["pedido"].cliente or ""), g["pedido"].pedido_venda or ""))
                for g in lista:
                    # se os itens desse pedido não estiverem 100% sincronizados
                    # (dado antigo, de antes do campo virar único por pedido), o
                    # card só fica verde quando TODOS já tiverem liberação real —
                    # senão fica azul (mais conservador).
                    g["confirmado"] = all(it.liberacao_real is not None for it in g["itens"])
                semana["dias"][coluna_idx] = lista

        total_sem_liberacao_prevista = (
            ItemPedido.query
            .filter(ItemPedido.liberacao_prevista.is_(None), ItemPedido.liberacao_real.is_(None))
            .filter(ItemPedido.status_producao != "FINALIZADO")
            .count()
        )

        mes_anterior_ano, mes_anterior_mes = _somar_meses(ano, mes, -1)
        mes_seguinte_ano, mes_seguinte_mes = _somar_meses(ano, mes, 1)

        return render_template(
            "programacao_semana.html",
            semanas=semanas,
            ano=ano, mes=mes,
            mes_label=f"{MESES_PT[mes - 1]} / {ano}",
            mes_anterior=dict(ano=mes_anterior_ano, mes=mes_anterior_mes),
            mes_seguinte=dict(ano=mes_seguinte_ano, mes=mes_seguinte_mes),
            hoje=hoje,
            total_sem_liberacao_prevista=total_sem_liberacao_prevista,
            DIAS_SEMANA_LABELS=["Domingo", "Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"],
        )

    @app.route("/programacao/novo", methods=["POST"])
    @requer_role("ADMIN", "PCP")
    def programacao_novo():
        f = request.form
        item_id = f.get("item_pedido_id", type=int)
        data_programada = _parse_data_form(f.get("data_programada"))
        item = db.session.get(ItemPedido, item_id) if item_id else None

        if item is None or data_programada is None:
            flash("Selecione um item e uma data válidos para programar.", "danger")
            return redirect(url_for("programacao_semana"))

        estacao_escolhida = f.get("estacao") or item.estacao or "—"

        nova = Programacao(
            item_pedido_id=item.id,
            data_programada=data_programada,
            estacao=estacao_escolhida,
            prioridade_producao=f.get("prioridade_producao") or None,
            observacao=f.get("observacao", "").strip() or None,
            responsavel_id=f.get("responsavel_id", type=int) or None,
            criado_por_id=current_user.id,
            status="ATIVA",
        )
        db.session.add(nova)

        # o Kanban de Estações mostra o item pela estação DELE (item.estacao), não
        # pela estação gravada na programação — então programar pra uma estação
        # diferente da atual também reatribui o item, pra ele aparecer no board certo
        if estacao_escolhida and estacao_escolhida != "—" and item.estacao != estacao_escolhida:
            antes = {c: getattr(item, c) for c in CAMPOS_HISTORICO_ITEM}
            item.estacao = estacao_escolhida
            depois = {c: getattr(item, c) for c in CAMPOS_HISTORICO_ITEM}
            _registrar_alteracoes("item_pedido", item.id, item.pedido_id, antes, depois, CAMPOS_HISTORICO_ITEM)

        db.session.commit()
        flash(f"\"{item.descricao_produto}\" programado para {data_programada.strftime('%d/%m/%Y')}.", "success")
        return redirect(url_for("programacao_semana", semana=data_programada.isoformat()))

    @app.route("/programacao/<int:programacao_id>/reprogramar", methods=["POST"])
    @requer_role("ADMIN", "PCP")
    def programacao_reprogramar(programacao_id):
        atual = db.session.get(Programacao, programacao_id)
        if atual is None or atual.status != "ATIVA":
            flash("Programação não encontrada ou já não está mais ativa.", "danger")
            return redirect(url_for("programacao_semana"))

        nova_data = _parse_data_form(request.form.get("data_programada"))
        if nova_data is None:
            flash("Informe uma nova data válida para reprogramar.", "danger")
            return redirect(url_for("programacao_semana"))

        atual.status = "REPROGRAMADA"
        nova = Programacao(
            item_pedido_id=atual.item_pedido_id,
            data_programada=nova_data,
            estacao=atual.estacao,
            prioridade_producao=atual.prioridade_producao,
            observacao=atual.observacao,
            responsavel_id=atual.responsavel_id,
            criado_por_id=current_user.id,
            status="ATIVA",
        )
        db.session.add(nova)
        db.session.commit()
        flash(f"Reprogramado para {nova_data.strftime('%d/%m/%Y')}.", "success")
        return redirect(url_for("programacao_semana", semana=nova_data.isoformat()))

    @app.route("/programacao/<int:programacao_id>/cancelar", methods=["POST"])
    @requer_role("ADMIN", "PCP")
    def programacao_cancelar(programacao_id):
        atual = db.session.get(Programacao, programacao_id)
        if atual is not None and atual.status == "ATIVA":
            atual.status = "CANCELADA"
            db.session.commit()
            flash("Programação cancelada.", "info")
        return redirect(url_for("programacao_semana"))


app = create_app()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
