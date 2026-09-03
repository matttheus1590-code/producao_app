"""Sincroniza Gestão Operação (`PedidoOperacao`) com uma planilha mais nova
que a usada no import inicial — pedido do Bruno em 28/08/2026, junto com o
novo recurso "Faturamento por Semana" na aba Resultados. Reutilizado de novo
em 03/09/2026 pra sincronizar com a planilha "03_09 Gestão de Fluxo Produtivo
2026" (mesmo layout de colunas, confirmado coluna a coluna contra COL antes
de reaproveitar).

Diferente de `importar_gestao_operacao.py` (que só roda sobre a tabela ainda
vazia, na primeira vez que o site sobe), este módulo roda POR CIMA de dados
que já existem — por isso cada chamada é protegida por um marcador próprio em
`ControleSistema` (ver `_sincronizar_gestao_operacao_28_08_2026` e
`_sincronizar_gestao_operacao_03_09_2026` em app.py), no mesmo espírito de
`sincronizar_planilha_producao.py` (sync de Gestão Produção). Cada marcador
roda só UMA vez.

Reaproveita 100% do parsing e do casamento por `pedido_venda` (exato +
aproximado por token numérico) já existentes em `importar_gestao_operacao.py`
— mesma aba "GESTAO OPERACAO". A diferença de comportamento:

  1. Pedido já existe no site (casou por pedido_venda, exato ou aproximado):
     ATUALIZADO — tanto os campos comerciais/PCP/logística/resultados (go_*)
     quanto os campos de identidade (cliente, vendedor, cidade, estado, país,
     frete, prioridade) — sempre que a planilha tem um valor. Célula em
     branco na planilha NUNCA apaga um valor que já existe no site (mesma
     regra do sync de Produção) — o pedido_venda em si nunca muda (é a
     própria chave de casamento).
  2. Pedido está na planilha mas não existe no site: CRIADO (mesma lógica de
     "novo" do importador original).
  3. Pedido existe no site mas não aparece na planilha: fica intocado.

Ampliação de 03/09/2026 — a planilha ganhou um bloco "Resultados" que
duplica 3 valores já existentes em colunas anteriores (VALOR NF EMITIDA,
CUSTO PRODUÇÃO REAL, CUSTO FRETE FINAL), só que bem mais completo (conferido
manualmente: nas linhas em que os dois lados têm valor, eles nunca
divergem). `_coalesce_numero` lê a coluna original primeiro e só usa a
duplicata de Resultados quando a original está vazia/ilegível — pedido do
Bruno: "quero que extraia o máximo de números e informações". Também
rastreamos (sem bloquear a sincronização) criticidade/OTD com valor fora do
esperado, pra aparecer no relatório em vez de sumir silenciosamente."""

import re

import openpyxl

from extensions import db
from importar_gestao_operacao import (
    COL,
    SHEET_NAME,
    _VALORES_INVALIDOS_PEDIDO_VENDA,
    _cell,
    _construir_indice_tokens,
    _match_pedido,
    _obter_ou_criar_transportadora,
    _parse_data,
    _parse_numero,
    _parse_otd,
    _parse_texto,
    _tokens_numericos,
    construir_mapa_transportadoras,
)
from models import PedidoOperacao

_CRITICIDADES_RECONHECIDAS = {"MEDIA", "MÉDIA", "BAIXA", "ALTA"}
_OTD_RECONHECIDOS = {"SIM", "S", "YES", "TRUE", "NÃO", "NAO", "N", "NO", "FALSE"}


def _coalesce_numero(ws, row, *campos):
    """_parse_numero da primeira coluna (dentre `campos`, na ordem dada) que
    tiver um valor válido — usado pros 3 pares "coluna original / duplicata
    mais completa no bloco Resultados" (ver docstring do módulo)."""
    for campo in campos:
        v = _parse_numero(_cell(ws, row, campo))
        if v is not None:
            return v
    return None


def _campos_go(ws, row, transportadora_obj):
    return dict(
        go_tipo_pedido=_parse_texto(_cell(ws, row, "tipo_pedido"), 60),
        go_contrato=_parse_texto(_cell(ws, row, "contrato"), 60),
        go_pedido_compra_cliente=_parse_texto(_cell(ws, row, "pedido_compra_cliente"), 60),
        go_proposta=_parse_texto(_cell(ws, row, "proposta"), 60),
        go_data_solicitada_entrega=_parse_data(_cell(ws, row, "data_solicitada_entrega")),
        go_status_pedido_info=_parse_texto(_cell(ws, row, "status_pedido_info"), 120),
        go_valor_pedido_operacao=_parse_numero(_cell(ws, row, "valor_total_pedido")),
        go_previsao_liberacao_pcp=_parse_data(_cell(ws, row, "previsao_liberacao_pcp")),
        go_data_efetiva_liberacao_pcp=_parse_data(_cell(ws, row, "data_efetiva_liberacao_pcp")),
        go_data_solicitada_cliente_retira=_parse_data(_cell(ws, row, "data_solicitada_cliente_retira")),
        go_custo_producao_real=_coalesce_numero(ws, row, "custo_producao_real", "custo_producao_real_resultados"),
        go_termino_semanal_pcp=_parse_texto(_cell(ws, row, "termino_semanal_pcp"), 40),
        go_data_emissao_nf=_parse_data(_cell(ws, row, "data_emissao_nf")),
        go_valor_nf_emitida=_coalesce_numero(ws, row, "valor_nf_emitida", "valor_nf_emitida_resultados"),
        go_numero_nf=_parse_texto(_cell(ws, row, "numero_nf"), 30),
        go_status_logistica=_parse_texto(_cell(ws, row, "status_logistica"), 60),
        go_data_pedido_expedido=_parse_data(_cell(ws, row, "data_pedido_expedido")),
        go_transportadora_id=transportadora_obj.id if transportadora_obj else None,
        go_custo_frete_previsto=_parse_numero(_cell(ws, row, "custo_frete_previsto")),
        go_custo_frete_final=_coalesce_numero(ws, row, "custo_frete_final", "custo_frete_final_resultados"),
        go_custo_frete_sobre_nota=_parse_numero(_cell(ws, row, "custo_frete_sobre_nota")),
        go_data_prevista_entrega=_parse_data(_cell(ws, row, "data_prevista_entrega")),
        go_data_real_entrega=_parse_data(_cell(ws, row, "data_real_entrega")),
        go_otd_realizado=_parse_otd(_cell(ws, row, "otd")),
        go_data_solicitada_cliente_final=_parse_data(_cell(ws, row, "data_solicitada_cliente_final")),
        go_data_entregue_cliente=_parse_data(_cell(ws, row, "data_entregue_cliente")),
        go_obs_operacao=_parse_texto(_cell(ws, row, "obs")),
        go_status_final_alinhamento=_parse_texto(_cell(ws, row, "status_final_alinhamento"), 60),
    )


def _prioridade_da_criticidade(v):
    criticidade = _parse_texto(v)
    if not criticidade:
        return None
    cu = criticidade.strip().upper()
    if cu in ("MEDIA", "MÉDIA"):
        return "MÉDIA"
    if cu in ("BAIXA", "ALTA"):
        return cu
    return None


def sincronizar_gestao_operacao(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    # -- índice de pedido_venda já cadastrados (exato + aproximado por token) --
    exatos = {}
    for p in PedidoOperacao.query.all():
        pv = (p.pedido_venda or "").strip()
        if pv:
            exatos[pv] = p
    indice_tokens = _construir_indice_tokens(exatos)

    # -- transportadoras: mesmo mapa canônico do importador original --
    nomes_transportadora = []
    for row in range(2, ws.max_row + 1):
        v = _cell(ws, row, "transportadora")
        if v:
            nomes_transportadora.append(v)
    mapa_transportadora = construir_mapa_transportadoras(nomes_transportadora)
    cache_transportadora = {}

    stats = {
        "linhas_lidas": 0,
        "pedidos_atualizados": 0,
        "pedidos_criados": 0,
        "campos_atualizados": 0,
        "exato": 0,
        "aproximado": 0,
        "sem_match_pedido_venda": 0,
        "criticidades_nao_reconhecidas": {},
        "otd_nao_reconhecidos": {},
    }

    # Não usa um LAST_ROW fixo (diferente do importador original) — planilhas
    # de sincronização mudam de tamanho a cada envio do Bruno, então lê até o
    # fim de verdade da planilha, ignorando linhas sem CLIENTE preenchido.
    for row in range(2, ws.max_row + 1):
        cliente = _parse_texto(_cell(ws, row, "cliente"))
        if not cliente:
            continue
        stats["linhas_lidas"] += 1

        pedido_venda_raw = _parse_texto(_cell(ws, row, "pedido_venda")) or ""
        pedido, tipo_match = _match_pedido(pedido_venda_raw, exatos, indice_tokens)
        if tipo_match == "exato":
            stats["exato"] += 1
        elif tipo_match == "aproximado":
            stats["aproximado"] += 1
        elif pedido_venda_raw:
            # tem "PEDIDO VENDA" preenchido na planilha mas não casou com
            # nenhum pedido_venda já cadastrado nem por token — linha vira
            # PedidoOperacao NOVO (comportamento já existente), só registra
            # pra aparecer no relatório pro Bruno.
            stats["sem_match_pedido_venda"] += 1

        criticidade_raw = _parse_texto(_cell(ws, row, "criticidade"))
        if criticidade_raw and criticidade_raw.strip().upper() not in _CRITICIDADES_RECONHECIDAS:
            chave = criticidade_raw.strip()
            stats["criticidades_nao_reconhecidas"][chave] = stats["criticidades_nao_reconhecidas"].get(chave, 0) + 1

        otd_raw = _parse_texto(_cell(ws, row, "otd"))
        if otd_raw and otd_raw.strip().upper() not in _OTD_RECONHECIDOS:
            chave = otd_raw.strip()
            stats["otd_nao_reconhecidos"][chave] = stats["otd_nao_reconhecidos"].get(chave, 0) + 1

        transportadora_obj = None
        transp_raw = _cell(ws, row, "transportadora")
        if transp_raw:
            disp = re.sub(r"\s+", " ", str(transp_raw).strip().upper())
            canonico = mapa_transportadora.get(disp, disp)
            transportadora_obj = _obter_ou_criar_transportadora(canonico, cache_transportadora)

        campos_identidade = dict(
            cliente=cliente,
            vendedor=_parse_texto(_cell(ws, row, "vendedor")),
            data_inclusao_pedido=_parse_data(_cell(ws, row, "data_inclusao_pedido")),
            pais=_parse_texto(_cell(ws, row, "pais")),
            estado=_parse_texto(_cell(ws, row, "estado")),
            cidade=_parse_texto(_cell(ws, row, "cidade")),
            frete=_parse_texto(_cell(ws, row, "modalidade_frete")),
            prioridade=_prioridade_da_criticidade(_cell(ws, row, "criticidade")),
        )
        campos_go = _campos_go(ws, row, transportadora_obj)

        if pedido is not None:
            # já existe — atualiza identidade + go_*, mas só quando a
            # planilha tem valor (célula em branco nunca apaga o que já está
            # no site) e nunca mexe no próprio pedido_venda (chave de casamento).
            mudou = False
            for campo, valor in {**campos_identidade, **campos_go}.items():
                if valor is None:
                    continue
                if getattr(pedido, campo) != valor:
                    setattr(pedido, campo, valor)
                    mudou = True
                    stats["campos_atualizados"] += 1
            if mudou:
                stats["pedidos_atualizados"] += 1
        else:
            pedido_venda_valido = (
                pedido_venda_raw if pedido_venda_raw.upper() not in _VALORES_INVALIDOS_PEDIDO_VENDA else None
            )
            novo = PedidoOperacao(
                cliente=cliente,
                vendedor=campos_identidade["vendedor"],
                pedido_venda=pedido_venda_valido,
                data_inclusao_pedido=campos_identidade["data_inclusao_pedido"],
                pais=campos_identidade["pais"] or "Brasil",
                estado=campos_identidade["estado"],
                cidade=campos_identidade["cidade"],
                frete=campos_identidade["frete"],
                prioridade=campos_identidade["prioridade"] or "MÉDIA",
                **campos_go,
            )
            db.session.add(novo)
            stats["pedidos_criados"] += 1
            # registra nos índices pra eventuais linhas seguintes com o mesmo pedido_venda
            if pedido_venda_raw:
                exatos.setdefault(pedido_venda_raw, novo)
                for tok in _tokens_numericos(pedido_venda_raw):
                    indice_tokens.setdefault(tok, []).append(novo)

    db.session.commit()
    return stats
