"""Importa (uma única vez) a planilha "Controle RNC - Qualidade" enviada pelo
Bruno em 31/08/2026 para dentro da tabela `rnc_qualidade` — tabela nova,
começa vazia, então essa importação é simplesmente um carregamento inicial
(sem "casar" com nada que já existe), mesmo espírito de `seed.py`
(importar_planilha) para a Gestão Produção original.

Aba lida: "Controle RNC - Agosto" (cabeçalho na linha 4, dados a partir da
linha 5). A aba "Dashboard" da planilha NÃO é importada — ela só continha
totais/KPIs já calculados a partir da própria aba de controle; no site esses
mesmos números são recalculados ao vivo a partir da tabela `rnc_qualidade`
(ver rota `qualidade_dashboard` em app.py), então não há nada pra guardar
dali. A aba "Listas" também não é lida em runtime — os valores dela viraram
as constantes RNC_*_OPCOES em models.py.

Os valores são gravados exatamente como estão na planilha (texto livre),
mesmo quando não batem 100% com as listas de sugestão (ex.: um RNC real com
Tipo de NC "Documental (OP)", valor que não está em RNC_TIPO_NC_OPCOES) —
fidelidade aos dados reais do Bruno é mais importante que normalizar pra uma
lista fixa; ele consegue corrigir manualmente dentro do site depois, se
quiser.
"""

from datetime import date, datetime

import openpyxl

from extensions import db
from models import RncQualidade

SHEET_NAME = "Controle RNC - Agosto"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

COL = {
    "numero_rnc": 1,
    "revisao": 2,
    "data_emissao": 3,
    "emitente": 4,
    "setor": 5,
    "origem": 6,
    "cliente_projeto": 7,
    "numero_pedido_contrato": 8,
    "produto_equipamento": 9,
    "numero_op": 10,
    "local_setor": 11,
    "data_identificacao": 12,
    "responsavel_identificacao": 13,
    "descricao_nc": 14,
    "qtd_nao_conforme": 15,
    "requisito_nao_atendido": 16,
    "tipo_nc": 17,
    "severidade": 18,
    "acao_contencao_imediata": 19,
    "porque_1": 20,
    "porque_2": 21,
    "porque_3": 22,
    "porque_4": 23,
    "porque_5": 24,
    "causa_raiz": 25,
    "ferramenta_analise": 26,
    "disposicao_produto": 27,
    "acao_corretiva_descricao": 28,
    "responsavel_acao_corretiva": 29,
    "prazo_acao_corretiva": 30,
    "data_realizacao": 31,
    "status_acao_corretiva": 32,
    "data_verificacao_eficacia": 33,
    "eficacia_acao": 34,
    "obs_verificacao": 35,
    "reincidencia": 36,
    "numero_rnc_relacionada": 37,
    "custo_estimado": 38,
    "status_geral": 39,
    "responsavel_qualidade": 40,
    "data_fechamento": 41,
    # 42 = "Dias em Aberto" da planilha (congelado no dia da exportação) —
    # não importado: no site esse valor é sempre recalculado ao vivo, ver
    # RncQualidade.dias_em_aberto.
    "evidencias_anexos": 43,
    "observacoes_gerais": 44,
}

CAMPOS_TEXTO = [
    "numero_rnc", "emitente", "setor", "origem", "cliente_projeto",
    "numero_pedido_contrato", "produto_equipamento", "numero_op", "local_setor",
    "responsavel_identificacao", "descricao_nc", "requisito_nao_atendido", "tipo_nc",
    "severidade", "acao_contencao_imediata", "porque_1", "porque_2", "porque_3",
    "porque_4", "porque_5", "causa_raiz", "ferramenta_analise", "disposicao_produto",
    "acao_corretiva_descricao", "responsavel_acao_corretiva", "status_acao_corretiva",
    "eficacia_acao", "obs_verificacao", "reincidencia", "numero_rnc_relacionada",
    "status_geral", "responsavel_qualidade", "evidencias_anexos", "observacoes_gerais",
]
CAMPOS_DATA = [
    "data_emissao", "data_identificacao", "prazo_acao_corretiva", "data_realizacao",
    "data_verificacao_eficacia", "data_fechamento",
]
CAMPOS_INTEIRO = ["revisao", "qtd_nao_conforme"]
CAMPOS_FLOAT = ["custo_estimado"]


def _limpar_texto(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _limpar_data(valor):
    """A maioria das células de data já vem como datetime (openpyxl com
    data_only=True), mas pelo menos uma célula real da planilha do Bruno
    veio como texto malformado ("20/08/202", faltando um dígito) — nesse
    caso melhor deixar em branco (o Bruno corrige manualmente) do que
    quebrar a importação inteira ou adivinhar o ano errado."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _limpar_inteiro(valor):
    if valor is None or valor == "":
        return None
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def _limpar_float(valor):
    if valor is None or valor == "":
        return None
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def importar_rnc_qualidade(xlsx_path):
    """Lê a planilha e cria um `RncQualidade` por linha preenchida. Retorna
    o total de RNCs criados. Não faz nenhum casamento/atualização — só roda
    quando a tabela `rnc_qualidade` está vazia (ver `_seed_rnc_qualidade` em
    app.py), então toda linha da planilha vira um registro novo."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    total = 0
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        numero_rnc = ws.cell(row=row, column=COL["numero_rnc"]).value
        cliente = ws.cell(row=row, column=COL["cliente_projeto"]).value
        # Linha "vazia" (fim da planilha, ou linha em branco no meio) —
        # ignora. Nº RNC ou Cliente/Projeto preenchido é o mínimo pra contar
        # como um RNC de verdade.
        if not _limpar_texto(numero_rnc) and not _limpar_texto(cliente):
            continue

        dados = {}
        for campo in CAMPOS_TEXTO:
            dados[campo] = _limpar_texto(ws.cell(row=row, column=COL[campo]).value)
        for campo in CAMPOS_DATA:
            dados[campo] = _limpar_data(ws.cell(row=row, column=COL[campo]).value)
        for campo in CAMPOS_INTEIRO:
            dados[campo] = _limpar_inteiro(ws.cell(row=row, column=COL[campo]).value)
        for campo in CAMPOS_FLOAT:
            dados[campo] = _limpar_float(ws.cell(row=row, column=COL[campo]).value)

        if not dados.get("status_geral"):
            # Planilha do Bruno não preenche essa coluna — todo RNC sem
            # data de fechamento é considerado "Aberto" por padrão.
            dados["status_geral"] = "Fechado" if dados.get("data_fechamento") else "Aberto"

        db.session.add(RncQualidade(**dados))
        total += 1

    db.session.commit()
    return total
