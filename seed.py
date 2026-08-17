"""Importa os pedidos da planilha original (aba 'GERAL TESTE') para o banco.

Rodado automaticamente na primeira inicialização do app (quando a tabela de
pedidos está vazia). Pode também ser executado manualmente:

    python seed.py data/controle_producao_base.xlsx
"""
import sys
from datetime import date, datetime

import openpyxl

from extensions import db
from models import ItemPedido, Pedido

SHEET_NAME = "GERAL TESTE"
PRIMEIRA_LINHA = 3
ULTIMA_LINHA = 1148

BR_UFS = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS",
    "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC",
    "SP", "SE", "TO",
}

# posição (1-indexada) de cada campo na aba original
COL = {
    "data_cliente": 1,       # A - DATA DO COMERCIAL / CLIENTE
    "data_inclusao": 2,      # B - DATA INCLUSÃO PEDIDO
    "prioridade": 5,         # E - PRIORIDADE
    "cliente": 6,            # F - Cliente (Nome Fantasia)
    "cnpj": 7,               # G - CNPJ
    "cidade": 8,             # H - Cidade
    "estado_pais": 9,        # I - ESTADO / PAIS
    "frete": 10,             # J - Modelo Frete
    "vendedor": 11,          # K - Vendedor
    "pedido_venda": 12,      # L - Pedido de Venda
    "descricao": 13,         # M - Descrição do Produto
    "quantidade": 14,        # N - Quantidade
    "rnc": 18,                # R - RNC
    "inicio_producao": 19,    # S - INICIO PRODUÇÃO
    "inicio_inspecao": 20,    # T - INICIO INSPEÇÃO / EMBALAGEM
    "termino_inspecao": 21,   # U - TÉRMINO DE INSPEÇÃO / EMBALAGEM
    "liberacao_faturamento": 22,  # V - LIBERAÇÃO DE FATURAMENTO
    "status": 24,              # X - STATUS PRODUÇÃO
    "estacao": 25,              # Y - ESTAÇÃO
    "custo_unitario": 26,       # Z - VENDA UNIDADE ITEM
    "obs": 29,                   # AC - OBS
    "liberacao_prevista": 30,    # AD - LIBERAÇÃO PREVISTA
}


def _texto(v, default=None, upper=False):
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    return s.upper() if upper else s


def _numero(v, default=0.0):
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default


def _parse_data(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _split_estado_pais(raw):
    if not raw:
        return None, "Brasil"
    v = str(raw).strip()
    vu = v.upper()
    if vu in BR_UFS:
        return vu, "Brasil"
    if vu == "EXP":
        return None, "Exterior"
    return None, v.title()


def importar_planilha(caminho, primeira_linha=PRIMEIRA_LINHA, ultima_linha=ULTIMA_LINHA):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[SHEET_NAME]

    total = 0
    for r in range(primeira_linha, ultima_linha + 1):
        cliente = ws.cell(row=r, column=COL["cliente"]).value
        descricao = ws.cell(row=r, column=COL["descricao"]).value
        if not cliente and not descricao:
            continue  # linha em branco

        estado, pais = _split_estado_pais(ws.cell(row=r, column=COL["estado_pais"]).value)

        pedido = Pedido(
            data_cliente=_parse_data(ws.cell(row=r, column=COL["data_cliente"]).value),
            data_inclusao_pedido=_parse_data(ws.cell(row=r, column=COL["data_inclusao"]).value),
            cliente=_texto(cliente, default="SEM CLIENTE"),
            cnpj=_texto(ws.cell(row=r, column=COL["cnpj"]).value),
            cidade=_texto(ws.cell(row=r, column=COL["cidade"]).value),
            estado=estado,
            pais=pais,
            frete=_texto(ws.cell(row=r, column=COL["frete"]).value, upper=True),
            vendedor=_texto(ws.cell(row=r, column=COL["vendedor"]).value),
            pedido_venda=_texto(ws.cell(row=r, column=COL["pedido_venda"]).value),
            prioridade=_texto(ws.cell(row=r, column=COL["prioridade"]).value, default="MÉDIA", upper=True),
            estacao=_texto(ws.cell(row=r, column=COL["estacao"]).value, upper=True),
            status_producao=_texto(ws.cell(row=r, column=COL["status"]).value, default="PENDENTE", upper=True),
            inicio_producao=_parse_data(ws.cell(row=r, column=COL["inicio_producao"]).value),
            inicio_inspecao=_parse_data(ws.cell(row=r, column=COL["inicio_inspecao"]).value),
            termino_inspecao=_parse_data(ws.cell(row=r, column=COL["termino_inspecao"]).value),
            liberacao_faturamento=_parse_data(ws.cell(row=r, column=COL["liberacao_faturamento"]).value),
            liberacao_prevista=_parse_data(ws.cell(row=r, column=COL["liberacao_prevista"]).value),
            rnc=_texto(ws.cell(row=r, column=COL["rnc"]).value),
            obs=_texto(ws.cell(row=r, column=COL["obs"]).value),
        )
        pedido.itens.append(
            ItemPedido(
                descricao_produto=_texto(descricao, default="SEM DESCRIÇÃO"),
                quantidade=_numero(ws.cell(row=r, column=COL["quantidade"]).value),
                custo_unitario=_numero(ws.cell(row=r, column=COL["custo_unitario"]).value),
            )
        )
        db.session.add(pedido)
        total += 1

    db.session.commit()
    return total


if __name__ == "__main__":
    # execução manual: `python seed.py caminho/para/planilha.xlsx`
    from app import create_app

    caminho = sys.argv[1] if len(sys.argv) > 1 else "data/controle_producao_base.xlsx"
    app = create_app()
    with app.app_context():
        total = importar_planilha(caminho)
        print(f"{total} pedidos importados com sucesso.")
